# dashboard/views.py
import json
import logging
from collections import defaultdict
from decimal import Decimal
from datetime import date, timedelta
from calendar import monthrange, monthcalendar
from urllib.parse import quote

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.staticfiles import finders
from django.core.paginator import Paginator
from django.db import transaction, models
from django.db.models import Sum, Count
from django.db.models import F
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.templatetags.static import static
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_http_methods, require_GET

from pywebpush import webpush, WebPushException

from trabajadores.models import Trabajador
from inventario.models import Insumo, InventarioTrabajador
from mantenimientos.models import (
    Mantenimiento,
    UsoInsumo,
    FotoMantenimiento,
    ChecklistMantenimiento,
)
from finanzas.models import (
    Ingreso,
    Egreso,
    MovimientoRecurrente,
    Factura,
    FacturaItem,
    ObligacionTrabajador,
    PagoTrabajador,
    LotePagoTrabajador,
    AnticipoTrabajador,
)
from clientes.models import Cliente
from contratos.models import Contrato
from contratos.programacion import (
    DIAS_SEMANA,
    cancelar_programacion_futura,
    generar_mantenimientos_contrato,
    normalizar_dias,
    validar_programacion,
)

try:
    from .models import PushSubscription
except Exception:
    PushSubscription = None

try:
    from .models import Notificacion
except Exception:
    Notificacion = None

try:
    from .models import ActividadSistema
except Exception:
    ActividadSistema = None

logger = logging.getLogger(__name__)


# -------------------
# Helpers de roles
# -------------------
def es_admin(user):
    if not user.is_authenticated:
        return False

    if user.is_superuser or user.is_staff:
        return True

    grupos = {g.name.strip().lower() for g in user.groups.all()}
    return (
        "administradores" in grupos
        or "administrador" in grupos
        or "admins" in grupos
        or "adimistradores" in grupos
    )


def es_trabajador(user):
    if not user.is_authenticated:
        return False
    grupos = {g.name.strip().lower() for g in user.groups.all()}
    return "trabajadores" in grupos or "trabajador" in grupos


# -------------------
# Fotos requeridas
# -------------------
FOTOS_REQUERIDAS = [
    "Inicio de Mantenimiento",
    "Fin de Mantenimiento",
    "Nivel PH y Cl",
]


def _nombre_foto_valido(nombre: str) -> bool:
    return nombre in FOTOS_REQUERIDAS


def _telefono_whatsapp_ecuador(telefono):
    """Normaliza teléfonos ecuatorianos para wa.me sin alterar el dato guardado."""
    digitos = "".join(ch for ch in str(telefono or "") if ch.isdigit())
    if digitos.startswith("00"):
        digitos = digitos[2:]
    if digitos.startswith("593"):
        return digitos
    if digitos.startswith("0") and len(digitos) >= 9:
        return "593" + digitos[1:]
    return digitos


# -------------------
# Helpers de fechas recurrentes
# -------------------
def _sumar_un_mes(fecha_base):
    nuevo_mes = fecha_base.month + 1
    nuevo_anio = fecha_base.year

    if nuevo_mes > 12:
        nuevo_mes = 1
        nuevo_anio += 1

    ultimo_dia = monthrange(nuevo_anio, nuevo_mes)[1]
    nuevo_dia = min(fecha_base.day, ultimo_dia)

    return date(nuevo_anio, nuevo_mes, nuevo_dia)


def _siguiente_fecha_recurrente(fecha_actual, frecuencia):
    if frecuencia == "semanal":
        return fecha_actual + timedelta(days=7)
    return _sumar_un_mes(fecha_actual)


def _inicio_fin_mes(anio, mes):
    primer_dia = date(anio, mes, 1)
    ultimo_dia = date(anio, mes, monthrange(anio, mes)[1])
    return primer_dia, ultimo_dia


def _mes_anterior(anio, mes):
    if mes == 1:
        return anio - 1, 12
    return anio, mes - 1


def _mes_siguiente(anio, mes):
    if mes == 12:
        return anio + 1, 1
    return anio, mes + 1


# -------------------
# Helpers ingresos / egresos manuales
# -------------------
def _ingreso_es_manual(ingreso):
    try:
        return bool(getattr(ingreso, "es_manual"))
    except Exception:
        pass

    cliente_id = getattr(ingreso, "cliente_id", None)
    contrato_id = getattr(ingreso, "contrato_id", None)
    return cliente_id is None and contrato_id is None


def _egreso_es_manual(egreso):
    try:
        return bool(getattr(egreso, "es_manual"))
    except Exception:
        pass

    mantenimiento_id = getattr(egreso, "mantenimiento_id", None)
    insumo_id = getattr(egreso, "insumo_id", None)
    return mantenimiento_id is None and insumo_id is None


def _crear_egreso_manual(concepto, categoria, total, fecha):
    kwargs = {
        "cantidad": 1,
        "costo_unitario": total,
        "fecha": fecha,
    }

    try:
        kwargs["mantenimiento"] = None
    except Exception:
        pass

    try:
        kwargs["insumo"] = None
    except Exception:
        pass

    if hasattr(Egreso, "concepto"):
        kwargs["concepto"] = concepto
    if hasattr(Egreso, "categoria"):
        kwargs["categoria"] = categoria or "Manual"

    egreso = Egreso.objects.create(**kwargs)
    return egreso


# -------------------
# Helpers notificaciones recurrentes
# -------------------
def _notificacion_recurrente_ya_existe_hoy(user, titulo, mensaje, url):
    if Notificacion is None:
        return False

    hoy = timezone.localdate()

    try:
        return Notificacion.objects.filter(
            user=user,
            titulo=titulo,
            mensaje=mensaje,
            url=url,
            creada_en__date=hoy,
        ).exists()
    except Exception:
        return False


def _notificacion_mantenimiento_hoy_ya_existe(user, mantenimiento):
    if Notificacion is None:
        return False

    hoy = timezone.localdate()
    titulo = "📅 Mantenimiento para hoy"
    mensaje = f"Hoy te toca el mantenimiento de {mantenimiento.cliente}."
    url = f"/dashboard/mantenimientos/{mantenimiento.pk}/"

    try:
        return Notificacion.objects.filter(
            user=user,
            titulo=titulo,
            mensaje=mensaje,
            url=url,
            creada_en__date=hoy,
        ).exists()
    except Exception:
        return False


def notificar_movimientos_recurrentes_proximos():
    """
    Recordatorio automático 1 día antes.
    Se ejecuta al entrar a vistas admin para avisar cobros/pagos próximos.
    """
    hoy = date.today()
    manana = hoy + timedelta(days=1)

    movimientos = MovimientoRecurrente.objects.filter(
        activo=True,
        proxima_fecha=manana
    ).order_by("proxima_fecha", "id")

    if not movimientos.exists():
        return

    admins = _admins_queryset()
    url = "/dashboard/finanzas/recurrentes/"

    for mov in movimientos:
        if mov.tipo == "ingreso":
            titulo = "💰 Cobro recurrente próximo"
            mensaje = (
                f"Mañana debes cobrar '{mov.concepto}' por ${mov.monto} "
                f"(fecha {mov.proxima_fecha})."
            )
        else:
            titulo = "💸 Pago recurrente próximo"
            mensaje = (
                f"Mañana debes pagar '{mov.concepto}' por ${mov.monto} "
                f"(fecha {mov.proxima_fecha})."
            )

        for admin_user in admins:
            if _notificacion_recurrente_ya_existe_hoy(admin_user, titulo, mensaje, url):
                continue

            _crear_notificacion(
                user=admin_user,
                titulo=titulo,
                mensaje=mensaje,
                url=url,
                enviar_push=True,
            )


def notificar_trabajadores_mantenimientos_hoy():
    """
    Envía notificación al trabajador el mismo día del mantenimiento.
    Evita duplicados por día/mantenimiento.
    """
    hoy = timezone.localdate()

    mantenimientos_hoy = (
        Mantenimiento.objects.filter(fecha=hoy)
        .select_related("cliente", "contrato")
        .prefetch_related("trabajadores", "trabajadores__user")
        .order_by("id")
    )

    for mantenimiento in mantenimientos_hoy:
        try:
            trabajadores = mantenimiento.trabajadores.all()
        except Exception:
            trabajadores = []

        for trabajador in trabajadores:
            user = getattr(trabajador, "user", None)
            if not user or not getattr(user, "is_active", False):
                continue

            if _notificacion_mantenimiento_hoy_ya_existe(user, mantenimiento):
                continue

            _crear_notificacion(
                user=user,
                titulo="📅 Mantenimiento para hoy",
                mensaje=f"Hoy te toca el mantenimiento de {mantenimiento.cliente}.",
                url=f"/dashboard/mantenimientos/{mantenimiento.pk}/",
                enviar_push=True,
            )


# -------------------
# Helpers calendario visual
# -------------------
def _build_calendario_mantenimientos(anio, mes, trabajador=None):
    primer_dia, ultimo_dia = _inicio_fin_mes(anio, mes)

    qs = (
        Mantenimiento.objects.filter(fecha__range=(primer_dia, ultimo_dia))
        .select_related("cliente", "contrato")
        .prefetch_related("trabajadores")
        .order_by("fecha", "estado", "id")
    )

    if trabajador is not None:
        qs = qs.filter(trabajadores=trabajador)

    mantenimientos = list(qs)

    por_fecha = {}
    for mantenimiento in mantenimientos:
        por_fecha.setdefault(mantenimiento.fecha, []).append(mantenimiento)

    semanas = []
    for semana in monthcalendar(anio, mes):
        fila = []
        for dia in semana:
            if dia == 0:
                fila.append({
                    "dia": 0,
                    "fecha": None,
                    "es_hoy": False,
                    "items": [],
                    "total": 0,
                    "realizados": 0,
                    "pendientes": 0,
                    "atrasados": 0,
                    "sin_asignar": 0,
                })
                continue

            fecha_actual = date(anio, mes, dia)
            items = por_fecha.get(fecha_actual, [])
            realizados = len([m for m in items if getattr(m, "estado", "") == "realizado"])
            pendientes = len([m for m in items if getattr(m, "estado", "") == "pendiente"])
            atrasados = len([
                m for m in items
                if getattr(m, "estado", "") == "pendiente" and fecha_actual < timezone.localdate()
            ])

            sin_asignar = 0
            for m in items:
                try:
                    if not m.trabajadores.exists():
                        sin_asignar += 1
                except Exception:
                    pass

            fila.append({
                "dia": dia,
                "fecha": fecha_actual,
                "es_hoy": fecha_actual == timezone.localdate(),
                "items": items[:4],
                "total": len(items),
                "realizados": realizados,
                "pendientes": pendientes,
                "atrasados": atrasados,
                "sin_asignar": sin_asignar,
            })
        semanas.append(fila)

    total_mes = len(mantenimientos)
    total_realizados = len([m for m in mantenimientos if getattr(m, "estado", "") == "realizado"])
    total_pendientes = len([m for m in mantenimientos if getattr(m, "estado", "") == "pendiente"])

    return {
        "anio": anio,
        "mes": mes,
        "semanas": semanas,
        "total_mes": total_mes,
        "total_realizados": total_realizados,
        "total_pendientes": total_pendientes,
    }


def _resolver_hora_mantenimiento(mantenimiento):
    """
    Intenta encontrar una hora programada en distintos nombres posibles de campo.
    Si no existe, devuelve None.
    """
    posibles_campos = [
        "hora",
        "hora_programada",
        "hora_visita",
        "hora_mantenimiento",
    ]

    for campo in posibles_campos:
        valor = getattr(mantenimiento, campo, None)
        if valor:
            return valor

    return None


def _bloque_horario_desde_hora(hora_obj):
    """
    Devuelve: manana / tarde / noche / sin_hora
    """
    if not hora_obj:
        return "sin_hora"

    try:
        hora = int(hora_obj.hour)
    except Exception:
        return "sin_hora"

    if 6 <= hora < 12:
        return "manana"
    if 12 <= hora < 18:
        return "tarde"
    if hora >= 18:
        return "noche"
    return "sin_hora"


def _build_agenda_semanal_mantenimientos(fecha_base, items):
    """
    Construye una agenda semanal de lunes a domingo.
    Agrupa por día y por bloques:
    - manana
    - tarde
    - noche
    - sin_hora
    """
    inicio_semana = fecha_base - timedelta(days=fecha_base.weekday())
    fin_semana = inicio_semana + timedelta(days=6)

    por_fecha = {}
    for mantenimiento in items:
        fecha_m = getattr(mantenimiento, "fecha", None)
        if not fecha_m:
            continue
        por_fecha.setdefault(fecha_m, []).append(mantenimiento)

    dias = []
    total_semana = 0
    total_pendientes = 0
    total_realizados = 0
    total_atrasados = 0
    total_sin_asignar = 0

    for i in range(7):
        fecha_actual = inicio_semana + timedelta(days=i)
        items_dia = sorted(
            por_fecha.get(fecha_actual, []),
            key=lambda m: (
                _resolver_hora_mantenimiento(m) is None,
                _resolver_hora_mantenimiento(m) or "",
                getattr(m, "estado", "") or "",
                getattr(m, "id", 0),
            )
        )

        bloques = {
            "manana": [],
            "tarde": [],
            "noche": [],
            "sin_hora": [],
        }

        pendientes = 0
        realizados = 0
        atrasados = 0
        sin_asignar = 0

        for m in items_dia:
            hora_m = _resolver_hora_mantenimiento(m)
            bloque = _bloque_horario_desde_hora(hora_m)
            bloques[bloque].append(m)

            estado_m = getattr(m, "estado", "") or ""
            if estado_m == "pendiente":
                pendientes += 1
            elif estado_m == "realizado":
                realizados += 1

            if estado_m == "pendiente" and fecha_actual < timezone.localdate():
                atrasados += 1

            try:
                if not m.trabajadores.exists():
                    sin_asignar += 1
            except Exception:
                pass

        total_semana += len(items_dia)
        total_pendientes += pendientes
        total_realizados += realizados
        total_atrasados += atrasados
        total_sin_asignar += sin_asignar

        dias.append({
            "fecha": fecha_actual,
            "dia_numero": fecha_actual.day,
            "nombre_corto": fecha_actual.strftime("%a"),
            "es_hoy": fecha_actual == timezone.localdate(),
            "total": len(items_dia),
            "pendientes": pendientes,
            "realizados": realizados,
            "atrasados": atrasados,
            "sin_asignar": sin_asignar,
            "bloques": [
                {"key": "manana", "label": "🌅 Mañana", "items": bloques["manana"]},
                {"key": "tarde", "label": "☀️ Tarde", "items": bloques["tarde"]},
                {"key": "noche", "label": "🌙 Noche", "items": bloques["noche"]},
                {"key": "sin_hora", "label": "🕘 Sin hora", "items": bloques["sin_hora"]},
            ],
        })

    return {
        "inicio": inicio_semana,
        "fin": fin_semana,
        "dias": dias,
        "total_semana": total_semana,
        "total_pendientes": total_pendientes,
        "total_realizados": total_realizados,
        "total_atrasados": total_atrasados,
        "total_sin_asignar": total_sin_asignar,
    }

# -------------------
# Helpers financiero
# -------------------
def _resumen_financiero_rango(fecha_inicio, fecha_fin):
    ingresos_total = Ingreso.objects.filter(
        fecha__range=(fecha_inicio, fecha_fin)
    ).aggregate(total=Sum("total"))["total"] or 0

    egresos_total = Egreso.objects.filter(
        fecha__range=(fecha_inicio, fecha_fin)
    ).aggregate(total=Sum("total"))["total"] or 0

    balance_total = ingresos_total - egresos_total

    return {
        "ingresos": float(ingresos_total),
        "egresos": float(egresos_total),
        "balance": float(balance_total),
    }


def _variacion_porcentual(actual, anterior):
    try:
        actual = float(actual or 0)
        anterior = float(anterior or 0)
        if anterior == 0:
            if actual == 0:
                return 0.0
            return 100.0
        return round(((actual - anterior) / anterior) * 100, 2)
    except Exception:
        return 0.0



def _tasa_cumplimiento(realizados, total):
    try:
        realizados = int(realizados or 0)
        total = int(total or 0)
        if total <= 0:
            return 0.0
        return round((realizados / total) * 100, 2)
    except Exception:
        return 0.0


def _tendencia_desde_variacion(valor):
    try:
        valor = float(valor or 0)
    except Exception:
        valor = 0.0

    if valor > 0:
        return "up"
    if valor < 0:
        return "down"
    return "flat"


def _estado_variacion_financiera(valor, invertido=False):
    try:
        valor = float(valor or 0)
    except Exception:
        valor = 0.0

    if valor == 0:
        return "neutral"

    if invertido:
        return "positivo" if valor < 0 else "negativo"
    return "positivo" if valor > 0 else "negativo"


def _mes_label(anio, mes):
    return f"{mes:02d}/{anio}"


def _iterar_meses_hacia_atras(fecha_base, cantidad=6):
    meses = []
    anio = fecha_base.year
    mes = fecha_base.month

    for _ in range(max(int(cantidad or 0), 1)):
        meses.append((anio, mes))
        anio, mes = _mes_anterior(anio, mes)

    meses.reverse()
    return meses


def _build_serie_financiera_meses(fecha_base, cantidad=6):
    labels = []
    ingresos = []
    egresos = []
    balances = []

    for anio, mes in _iterar_meses_hacia_atras(fecha_base, cantidad=cantidad):
        inicio_mes, fin_mes = _inicio_fin_mes(anio, mes)
        resumen = _resumen_financiero_rango(inicio_mes, fin_mes)

        labels.append(_mes_label(anio, mes))
        ingresos.append(float(resumen["ingresos"]))
        egresos.append(float(resumen["egresos"]))
        balances.append(float(resumen["balance"]))

    return {
        "labels": labels,
        "ingresos": ingresos,
        "egresos": egresos,
        "balance": balances,
    }


def _build_serie_operativa_meses(fecha_base, cantidad=6):
    labels = []
    realizados = []
    pendientes = []
    atrasados = []

    for anio, mes in _iterar_meses_hacia_atras(fecha_base, cantidad=cantidad):
        inicio_mes, fin_mes = _inicio_fin_mes(anio, mes)

        qs_mes = Mantenimiento.objects.filter(fecha__range=(inicio_mes, fin_mes))
        labels.append(_mes_label(anio, mes))
        realizados.append(qs_mes.filter(estado="realizado").count())
        pendientes.append(qs_mes.filter(estado="pendiente").count())
        atrasados.append(qs_mes.filter(estado="pendiente", fecha__lt=timezone.localdate()).count())

    return {
        "labels": labels,
        "realizados": realizados,
        "pendientes": pendientes,
        "atrasados": atrasados,
    }


def _top_clientes_facturacion(fecha_inicio, fecha_fin, limite=5):
    acumulado = {}

    ingresos = (
        Ingreso.objects.filter(fecha__range=(fecha_inicio, fecha_fin), cliente__isnull=False)
        .select_related("cliente", "contrato")
        .order_by("fecha", "id")
    )

    for ingreso in ingresos:
        cliente = getattr(ingreso, "cliente", None)
        if not cliente:
            continue

        key = getattr(cliente, "pk", None) or str(cliente)
        if key not in acumulado:
            acumulado[key] = {
                "cliente_id": getattr(cliente, "pk", None),
                "cliente": str(cliente),
                "total": Decimal("0"),
                "movimientos": 0,
            }

        acumulado[key]["total"] += Decimal(getattr(ingreso, "total", 0) or 0)
        acumulado[key]["movimientos"] += 1

    items = sorted(
        acumulado.values(),
        key=lambda x: (-float(x["total"]), -x["movimientos"], x["cliente"])
    )[:limite]

    for idx, item in enumerate(items, start=1):
        item["posicion"] = idx
        item["total_float"] = float(item["total"])
        item["ticket_promedio"] = round(float(item["total"]) / item["movimientos"], 2) if item["movimientos"] else 0.0

    return items


def _top_contratos_facturacion(fecha_inicio, fecha_fin, limite=5):
    acumulado = {}

    ingresos = (
        Ingreso.objects.filter(fecha__range=(fecha_inicio, fecha_fin), contrato__isnull=False)
        .select_related("cliente", "contrato")
        .order_by("fecha", "id")
    )

    for ingreso in ingresos:
        contrato = getattr(ingreso, "contrato", None)
        if not contrato:
            continue

        key = getattr(contrato, "pk", None) or str(contrato)
        cliente = getattr(ingreso, "cliente", None)
        if key not in acumulado:
            acumulado[key] = {
                "contrato_id": getattr(contrato, "pk", None),
                "contrato": str(contrato),
                "cliente": str(cliente) if cliente else "-",
                "total": Decimal("0"),
                "movimientos": 0,
            }

        acumulado[key]["total"] += Decimal(getattr(ingreso, "total", 0) or 0)
        acumulado[key]["movimientos"] += 1

    items = sorted(
        acumulado.values(),
        key=lambda x: (-float(x["total"]), -x["movimientos"], x["contrato"])
    )[:limite]

    for idx, item in enumerate(items, start=1):
        item["posicion"] = idx
        item["total_float"] = float(item["total"])

    return items


def _top_trabajadores_mes(fecha_inicio, fecha_fin, limite=5):
    acumulado = {}

    mantenimientos = (
        Mantenimiento.objects.filter(fecha__range=(fecha_inicio, fecha_fin))
        .select_related("cliente", "contrato")
        .prefetch_related("trabajadores", "trabajadores__user")
        .order_by("fecha", "id")
    )

    for mantenimiento in mantenimientos:
        estado = getattr(mantenimiento, "estado", "") or ""
        for trabajador in mantenimiento.trabajadores.all():
            username = str(getattr(getattr(trabajador, "user", None), "username", "") or f"Trabajador #{getattr(trabajador, 'pk', '')}")
            key = getattr(trabajador, "pk", None) or username

            if key not in acumulado:
                acumulado[key] = {
                    "trabajador_id": getattr(trabajador, "pk", None),
                    "username": username,
                    "total": 0,
                    "realizados": 0,
                    "pendientes": 0,
                }

            acumulado[key]["total"] += 1
            if estado == "realizado":
                acumulado[key]["realizados"] += 1
            else:
                acumulado[key]["pendientes"] += 1

    items = sorted(
        acumulado.values(),
        key=lambda x: (-x["realizados"], -x["total"], x["username"])
    )[:limite]

    for idx, item in enumerate(items, start=1):
        item["posicion"] = idx
        item["cumplimiento"] = _tasa_cumplimiento(item["realizados"], item["total"])

    return items


def _top_insumos_mes(fecha_inicio, fecha_fin, limite=5):
    acumulado = {}

    usos = (
        UsoInsumo.objects.filter(mantenimiento__fecha__range=(fecha_inicio, fecha_fin))
        .select_related("insumo", "mantenimiento", "mantenimiento__cliente")
        .order_by("id")
    )

    for uso in usos:
        insumo = getattr(uso, "insumo", None)
        nombre = str(insumo) if insumo else "Insumo"
        key = getattr(insumo, "pk", None) or nombre

        if key not in acumulado:
            acumulado[key] = {
                "insumo_id": getattr(insumo, "pk", None) if insumo else None,
                "insumo": nombre,
                "cantidad_total": 0,
                "mantenimientos": set(),
            }

        acumulado[key]["cantidad_total"] += int(getattr(uso, "cantidad", 0) or 0)
        mantenimiento_id = getattr(uso, "mantenimiento_id", None)
        if mantenimiento_id:
            acumulado[key]["mantenimientos"].add(mantenimiento_id)

    items = []
    for item in acumulado.values():
        items.append({
            "insumo_id": item["insumo_id"],
            "insumo": item["insumo"],
            "cantidad_total": item["cantidad_total"],
            "mantenimientos_count": len(item["mantenimientos"]),
        })

    items = sorted(
        items,
        key=lambda x: (-x["cantidad_total"], -x["mantenimientos_count"], x["insumo"])
    )[:limite]

    for idx, item in enumerate(items, start=1):
        item["posicion"] = idx

    return items


def _kpis_analitica_pro(hoy):
    inicio_mes, fin_mes = _inicio_fin_mes(hoy.year, hoy.month)
    anio_ant, mes_ant = _mes_anterior(hoy.year, hoy.month)
    inicio_mes_ant, fin_mes_ant = _inicio_fin_mes(anio_ant, mes_ant)

    resumen_mes_actual = _resumen_financiero_rango(inicio_mes, fin_mes)
    resumen_mes_anterior = _resumen_financiero_rango(inicio_mes_ant, fin_mes_ant)

    ingresos_mes_qs = Ingreso.objects.filter(fecha__range=(inicio_mes, fin_mes))
    ingresos_mes_count = ingresos_mes_qs.count()
    ticket_promedio_mes = round(float(resumen_mes_actual["ingresos"]) / ingresos_mes_count, 2) if ingresos_mes_count else 0.0

    clientes_activos_ids = set(
        Mantenimiento.objects.filter(fecha__range=(inicio_mes, fin_mes), cliente__isnull=False)
        .values_list("cliente_id", flat=True)
    )
    clientes_activos_ids.update(
        Ingreso.objects.filter(fecha__range=(inicio_mes, fin_mes), cliente__isnull=False)
        .values_list("cliente_id", flat=True)
    )
    clientes_activos_ids = {cid for cid in clientes_activos_ids if cid}

    mantenimientos_mes_qs = Mantenimiento.objects.filter(fecha__range=(inicio_mes, fin_mes))
    mantenimientos_realizados_mes = mantenimientos_mes_qs.filter(estado="realizado").count()
    mantenimientos_pendientes_mes = mantenimientos_mes_qs.filter(estado="pendiente").count()
    mantenimientos_atrasados_mes = Mantenimiento.objects.filter(fecha__lt=hoy, estado="pendiente").count()
    total_mantenimientos_mes = mantenimientos_mes_qs.count()

    cumplimiento_mes = _tasa_cumplimiento(mantenimientos_realizados_mes, total_mantenimientos_mes)

    return {
        "inicio_mes": inicio_mes,
        "fin_mes": fin_mes,
        "inicio_mes_anterior": inicio_mes_ant,
        "fin_mes_anterior": fin_mes_ant,
        "resumen_mes_actual": resumen_mes_actual,
        "resumen_mes_anterior": resumen_mes_anterior,
        "variacion_ingresos_mes": _variacion_porcentual(resumen_mes_actual["ingresos"], resumen_mes_anterior["ingresos"]),
        "variacion_egresos_mes": _variacion_porcentual(resumen_mes_actual["egresos"], resumen_mes_anterior["egresos"]),
        "variacion_balance_mes": _variacion_porcentual(resumen_mes_actual["balance"], resumen_mes_anterior["balance"]),
        "ticket_promedio_mes": ticket_promedio_mes,
        "ingresos_mes_count": ingresos_mes_count,
        "clientes_activos_mes": len(clientes_activos_ids),
        "mantenimientos_realizados_mes": mantenimientos_realizados_mes,
        "mantenimientos_pendientes_mes": mantenimientos_pendientes_mes,
        "mantenimientos_atrasados_mes": mantenimientos_atrasados_mes,
        "total_mantenimientos_mes": total_mantenimientos_mes,
        "cumplimiento_mes": cumplimiento_mes,
        "top_clientes_facturacion": _top_clientes_facturacion(inicio_mes, fin_mes),
        "top_contratos_facturacion": _top_contratos_facturacion(inicio_mes, fin_mes),
        "top_trabajadores_mes": _top_trabajadores_mes(inicio_mes, fin_mes),
        "top_insumos_mes": _top_insumos_mes(inicio_mes, fin_mes),
        "serie_financiera_6m": _build_serie_financiera_meses(hoy, cantidad=6),
        "serie_operativa_6m": _build_serie_operativa_meses(hoy, cantidad=6),
    }


# -------------------
# Automatización de movimientos recurrentes
# -------------------
def procesar_movimientos_recurrentes():
    hoy = date.today()

    movimientos = MovimientoRecurrente.objects.filter(
        activo=True,
        proxima_fecha__lte=hoy
    ).order_by("proxima_fecha", "id")

    total_ingresos = 0
    total_egresos = 0
    movimientos_procesados = 0

    for mov in movimientos:
        generado_este_movimiento = 0

        while mov.activo and mov.proxima_fecha <= hoy:
            fecha_mov = mov.proxima_fecha

            if mov.tipo == "ingreso":
                Ingreso.objects.create(
                    concepto=mov.concepto,
                    total=mov.monto,
                    fecha=fecha_mov
                )

                total_ingresos += 1
                generado_este_movimiento += 1

                _notificar_admins(
                    titulo="💰 Ingreso recurrente generado",
                    mensaje=f"Se generó el ingreso recurrente '{mov.concepto}' por ${mov.monto} (fecha {fecha_mov}).",
                    url="/dashboard/finanzas/flujo/",
                    enviar_push=True,
                )

            elif mov.tipo == "egreso":
                _crear_egreso_manual(
                    concepto=mov.concepto,
                    categoria="Recurrente",
                    total=mov.monto,
                    fecha=fecha_mov,
                )

                total_egresos += 1
                generado_este_movimiento += 1

                _notificar_admins(
                    titulo="💸 Pago recurrente generado",
                    mensaje=f"Se registró el egreso recurrente '{mov.concepto}' por ${mov.monto} (fecha {fecha_mov}).",
                    url="/dashboard/finanzas/flujo/",
                    enviar_push=True,
                )

            mov.proxima_fecha = _siguiente_fecha_recurrente(fecha_mov, mov.frecuencia)
            mov.save(update_fields=["proxima_fecha"])

        if generado_este_movimiento > 0:
            movimientos_procesados += 1

    return {
        "movimientos_procesados": movimientos_procesados,
        "ingresos_generados": total_ingresos,
        "egresos_generados": total_egresos,
        "total_generados": total_ingresos + total_egresos,
    }


# -------------------
# Login / Logout
# -------------------
def login_view(request):
    ctx = {"error": ""}

    next_url = (request.GET.get("next", "") or "").strip()
    if request.method == "POST":
        next_url = (request.POST.get("next", next_url) or "").strip()

    def safe_redirect_target(url: str) -> str:
        if url and url_has_allowed_host_and_scheme(
            url=url,
            allowed_hosts={request.get_host()},
            require_https=not settings.DEBUG,
        ):
            return url
        return "/dashboard/inicio/"

    if request.method == "POST":
        username = (request.POST.get("username", "") or "").strip()
        password = (request.POST.get("password", "") or "").strip()
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            messages.success(request, "Bienvenido.")
            return redirect(safe_redirect_target(next_url))

        ctx["error"] = "Usuario o contraseña incorrectos"
        ctx["next"] = next_url
        messages.error(request, ctx["error"])
        return render(request, "dashboard/login.html", ctx)

    ctx["next"] = next_url
    return render(request, "dashboard/login.html", ctx)


def logout_view(request):
    logout(request)
    return redirect("/login/")


# -------------------
# /dashboard/sw.js
# -------------------
def sw_js_view(request):
    path = finders.find("dashboard/sw-dashboard.js")

    if path:
        with open(path, "rb") as f:
            content = f.read()
        resp = HttpResponse(content, content_type="application/javascript; charset=utf-8")
    else:
        resp = HttpResponse(
            "/* ERROR: dashboard/sw-dashboard.js no encontrado en staticfiles */",
            content_type="application/javascript; charset=utf-8",
        )

    resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp["Pragma"] = "no-cache"
    resp["Expires"] = "0"
    resp["Service-Worker-Allowed"] = "/dashboard/"
    return resp


# -------------------
# Manifest servido por Django
# -------------------
def manifest_json_view(request):
    data = {
        "name": "Piscinas App",
        "short_name": "Piscinas",
        "description": "Gestión de mantenimientos, operativo y finanzas.",
        "id": "/dashboard/",
        "start_url": "/dashboard/inicio/",
        "scope": "/dashboard/",
        "display": "standalone",
        "background_color": "#ffffff",
        "theme_color": "#0d6efd",
        "orientation": "portrait",
        "icons": [
            {
                "src": static("dashboard/icons/icon-192.png"),
                "sizes": "192x192",
                "type": "image/png",
                "purpose": "any",
            },
            {
                "src": static("dashboard/icons/icon-192-maskable.png"),
                "sizes": "192x192",
                "type": "image/png",
                "purpose": "maskable",
            },
            {
                "src": static("dashboard/icons/icon-512.png"),
                "sizes": "512x512",
                "type": "image/png",
                "purpose": "any",
            },
            {
                "src": static("dashboard/icons/icon-512-maskable.png"),
                "sizes": "512x512",
                "type": "image/png",
                "purpose": "maskable",
            },
        ],
    }

    resp = JsonResponse(data)
    resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp["Pragma"] = "no-cache"
    resp["Expires"] = "0"
    return resp


# ==========================================================
# PUSH / NOTIFICACIONES HELPERS
# ==========================================================
def _clean_str(value) -> str:
    return (value or "").replace("\n", "").replace("\r", "").strip()


def _normalize_subscription_payload(data: dict) -> dict:
    if not isinstance(data, dict):
        return {}

    if "subscription" in data and isinstance(data["subscription"], dict):
        data = data["subscription"]

    endpoint = _clean_str(data.get("endpoint"))
    keys = data.get("keys") if isinstance(data.get("keys"), dict) else {}

    p256dh = _clean_str(keys.get("p256dh"))
    auth = _clean_str(keys.get("auth"))

    if not endpoint or not p256dh or not auth:
        return {}

    return {
        "endpoint": endpoint,
        "p256dh": p256dh,
        "auth": auth,
        "raw": data,
    }


def _push_status_code_from_exception(ex):
    try:
        return getattr(getattr(ex, "response", None), "status_code", None)
    except Exception:
        return None


def _send_push_to_user(user, title, body, url="/dashboard/notificaciones/", tag=None):
    if PushSubscription is None:
        return {"ok": False, "error": "Modelo PushSubscription no disponible"}

    vapid_private_key = (getattr(settings, "VAPID_PRIVATE_KEY", "") or "").strip()
    if not vapid_private_key:
        return {"ok": False, "error": "VAPID_PRIVATE_KEY vacío en settings/env"}

    vapid_subject = (
        getattr(settings, "VAPID_SUBJECT", "") or "mailto:admin@piscinas-app.local"
    ).strip()

    subs = PushSubscription.objects.filter(user=user).order_by("-updated_at", "-created_at")
    if not subs.exists():
        return {"ok": False, "error": "Usuario sin suscripciones push"}

    payload = json.dumps(
        {
            "title": title,
            "body": body,
            "url": url or "/dashboard/notificaciones/",
            "tag": tag or f"notif-{user.pk}",
        }
    )

    sent = 0
    failed = 0

    for s in subs:
        subscription_info = {
            "endpoint": s.endpoint,
            "keys": {"p256dh": s.p256dh, "auth": s.auth},
        }

        try:
            webpush(
                subscription_info=subscription_info,
                data=payload,
                vapid_private_key=vapid_private_key,
                vapid_claims={"sub": vapid_subject},
                content_encoding="aes128gcm",
                ttl=60,
            )
            sent += 1

        except WebPushException as ex:
            failed += 1
            status = _push_status_code_from_exception(ex)

            logger.warning(
                "Push falló user=%s sub_id=%s status=%s error=%s",
                getattr(user, "username", "unknown"),
                s.id,
                status,
                str(ex),
            )

            if status in (404, 410):
                try:
                    s.delete()
                except Exception:
                    logger.exception("No se pudo borrar sub expirada id=%s", s.id)

        except Exception:
            failed += 1
            logger.exception(
                "Error inesperado enviando push a user=%s",
                getattr(user, "username", "unknown"),
            )

    return {"ok": sent > 0, "sent": sent, "failed": failed}


def _crear_notificacion(user, titulo, mensaje, url="/dashboard/notificaciones/", enviar_push=False):
    if not user or not getattr(user, "is_authenticated", False):
        return None

    notif = None

    if Notificacion is not None:
        try:
            notif = Notificacion.objects.create(
                user=user,
                titulo=titulo,
                mensaje=mensaje,
                url=url,
                leida=False,
            )
        except Exception:
            logger.exception(
                "No se pudo crear Notificacion para user=%s",
                getattr(user, "username", "unknown"),
            )

    if enviar_push:
        try:
            _send_push_to_user(
                user=user,
                title=titulo,
                body=mensaje,
                url=url or "/dashboard/notificaciones/",
                tag=f"notif-{user.pk}",
            )
        except Exception:
            logger.exception(
                "No se pudo enviar push para user=%s",
                getattr(user, "username", "unknown"),
            )

    return notif


def _registrar_actividad(user, titulo, descripcion, url=""):
    if ActividadSistema is None:
        return None

    try:
        return ActividadSistema.objects.create(
            user=user if getattr(user, "is_authenticated", False) else None,
            titulo=titulo,
            descripcion=descripcion,
            url=url or "",
        )
    except Exception:
        logger.exception(
            "No se pudo registrar ActividadSistema para user=%s",
            getattr(user, "username", "unknown"),
        )
        return None


def _admins_queryset():
    ids_admins = []

    for user in User.objects.filter(is_active=True):
        try:
            if es_admin(user):
                ids_admins.append(user.id)
        except Exception:
            pass

    return User.objects.filter(id__in=ids_admins)


def _notificar_admins(titulo, mensaje, url="/dashboard/notificaciones/", enviar_push=False, excluir_user_id=None):
    admins = _admins_queryset()

    if excluir_user_id:
        admins = admins.exclude(id=excluir_user_id)

    for admin_user in admins:
        _crear_notificacion(
            user=admin_user,
            titulo=titulo,
            mensaje=mensaje,
            url=url,
            enviar_push=enviar_push,
        )


# ==========================================================
# Helpers operativo admin
# ==========================================================
def _mantenimiento_match_busqueda(mantenimiento, q: str) -> bool:
    if not q:
        return True

    q = q.strip().lower()
    if not q:
        return True

    bloques = [
        str(getattr(mantenimiento, "cliente", "") or ""),
        str(getattr(mantenimiento, "contrato", "") or ""),
        str(getattr(mantenimiento, "estado", "") or ""),
        str(getattr(mantenimiento, "fecha", "") or ""),
        str(getattr(mantenimiento, "observaciones", "") or ""),
    ]

    try:
        for trabajador in mantenimiento.trabajadores.all():
            bloques.append(str(getattr(getattr(trabajador, "user", None), "username", "") or ""))
    except Exception:
        pass

    texto = " ".join(bloques).lower()
    return q in texto


def _filtrar_mantenimientos_por_busqueda(items, q: str):
    if not q:
        return list(items)
    return [m for m in items if _mantenimiento_match_busqueda(m, q)]


def _resumen_trabajadores_desde_listas(dia_list, atrasados, proximos):
    resumen = {}

    def asegurar_trabajador(trabajador):
        trabajador_id = getattr(trabajador, "id", None)
        username = str(getattr(getattr(trabajador, "user", None), "username", "") or "Sin usuario")

        if trabajador_id not in resumen:
            resumen[trabajador_id] = {
                "trabajadores__id": trabajador_id,
                "trabajadores__user__username": username,
                "dia": 0,
                "atrasados": 0,
                "proximos": 0,
            }
        return resumen[trabajador_id]

    for mantenimiento in dia_list:
        try:
            for trabajador in mantenimiento.trabajadores.all():
                asegurar_trabajador(trabajador)["dia"] += 1
        except Exception:
            pass

    for mantenimiento in atrasados:
        try:
            for trabajador in mantenimiento.trabajadores.all():
                asegurar_trabajador(trabajador)["atrasados"] += 1
        except Exception:
            pass

    for mantenimiento in proximos:
        try:
            for trabajador in mantenimiento.trabajadores.all():
                asegurar_trabajador(trabajador)["proximos"] += 1
        except Exception:
            pass

    return sorted(
        resumen.values(),
        key=lambda x: (-x["atrasados"], -x["dia"], -x["proximos"], x["trabajadores__user__username"]),
    )


def _sin_asignar_count(items):
    total = 0
    for m in items:
        try:
            if not m.trabajadores.exists():
                total += 1
        except Exception:
            pass
    return total


def _clasificar_estado_trabajador(carga_hoy, atrasados, proximos):
    carga_total = carga_hoy + atrasados + proximos

    if atrasados > 0 or carga_hoy >= 3 or carga_total >= 6:
        return "saturado"
    if carga_hoy >= 2 or carga_total >= 3:
        return "media"
    return "libre"


# ==========================================================
# PUSH
# ==========================================================
@login_required
@require_http_methods(["GET"])
def vapid_public_key_view(request):
    key = _clean_str(getattr(settings, "VAPID_PUBLIC_KEY", ""))
    if not key:
        return JsonResponse(
            {"publicKey": "", "warning": "VAPID_PUBLIC_KEY vacío"},
            status=200,
        )
    return JsonResponse({"publicKey": key})


@login_required
@require_http_methods(["GET"])
def push_status_view(request):
    if PushSubscription is None:
        return JsonResponse(
            {"ok": False, "enabled": False, "count": 0, "error": "Modelo PushSubscription no disponible"},
            status=500,
        )

    qs = PushSubscription.objects.filter(user=request.user).order_by("-updated_at", "-created_at")
    return JsonResponse(
        {
            "ok": True,
            "enabled": qs.exists(),
            "count": qs.count(),
        }
    )


@login_required
@require_http_methods(["POST"])
def save_subscription_view(request):
    if PushSubscription is None:
        return JsonResponse(
            {"ok": False, "error": "Modelo PushSubscription no disponible"},
            status=500,
        )

    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON inválido"}, status=400)

    norm = _normalize_subscription_payload(data)
    if not norm:
        return JsonResponse({"ok": False, "error": "Payload incompleto"}, status=400)

    user_agent = (request.META.get("HTTP_USER_AGENT", "") or "").strip()

    try:
        with transaction.atomic():
            obj, created = PushSubscription.objects.update_or_create(
                endpoint=norm["endpoint"],
                defaults={
                    "user": request.user,
                    "p256dh": norm["p256dh"],
                    "auth": norm["auth"],
                    "user_agent": user_agent,
                },
            )
    except Exception as ex:
        logger.exception("Error guardando push subscription para user=%s", request.user.pk)
        return JsonResponse(
            {"ok": False, "error": f"No se pudo guardar la suscripción: {str(ex)}"},
            status=500,
        )

    logger.info(
        "Push subscription guardada user=%s endpoint=%s created=%s",
        request.user.username,
        norm["endpoint"][:80],
        created,
    )

    return JsonResponse(
        {
            "ok": True,
            "created": created,
            "updated": not created,
            "id": obj.id,
            "user": request.user.username,
        },
        status=201 if created else 200,
    )


@login_required
@require_http_methods(["POST"])
def delete_subscription_view(request):
    if PushSubscription is None:
        return JsonResponse(
            {"ok": False, "error": "Modelo PushSubscription no disponible"},
            status=500,
        )

    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON inválido"}, status=400)

    norm = _normalize_subscription_payload(data)
    endpoint = norm.get("endpoint") if norm else _clean_str(data.get("endpoint"))

    if not endpoint:
        return JsonResponse({"ok": False, "error": "Endpoint requerido"}, status=400)

    deleted, _ = PushSubscription.objects.filter(
        user=request.user,
        endpoint=endpoint,
    ).delete()

    return JsonResponse({"ok": True, "deleted": bool(deleted)})


@login_required
@require_http_methods(["GET", "POST"])
def push_test_view(request):
    if request.method == "GET":
        return JsonResponse(
            {"ok": True, "msg": "push_test_view OK. Usa POST para enviar push."}
        )

    result = _send_push_to_user(
        user=request.user,
        title="✅ Prueba Piscinas App",
        body=f"Hola {request.user.username}, tu Push está funcionando 🎉",
        url="/dashboard/notificaciones/",
        tag=f"piscinas-{request.user.username}",
    )

    if not result.get("ok"):
        return JsonResponse(result, status=400)

    return JsonResponse(result)


# -------------------
# Centro de Acciones del administrador
# -------------------
def _centro_acciones_contexto():
    """Construye prioridades diarias sin duplicar la lógica financiera."""
    hoy = timezone.localdate()
    proximos_tres_dias = hoy + timedelta(days=3)
    limite_programacion = hoy + timedelta(days=7)

    facturas_base = list(
        Factura.objects
        .exclude(estado=Factura.ESTADO_ANULADA)
        .select_related("cliente", "contrato")
        .prefetch_related("pagos")
        .order_by("fecha_vencimiento", "id")
    )
    facturas_pendientes = [f for f in facturas_base if f.saldo > 0]

    cobros_vencidos = [
        f for f in facturas_pendientes
        if f.fecha_vencimiento and f.fecha_vencimiento < hoy
    ]
    cobros_hoy = [
        f for f in facturas_pendientes
        if (
            (f.fecha_cobro_desde and f.fecha_cobro_desde <= hoy <= f.fecha_vencimiento)
            or f.fecha_vencimiento == hoy
        )
    ]
    ids_prioridad = {f.pk for f in cobros_vencidos + cobros_hoy}
    cobros_proximos = [
        f for f in facturas_pendientes
        if f.pk not in ids_prioridad
        and f.fecha_cobro_desde
        and hoy < f.fecha_cobro_desde <= proximos_tres_dias
    ]

    facturas_por_emitir = [
        f for f in facturas_base
        if f.requiere_factura
        and not f.factura_enviada
        and f.fecha_facturacion_programada
        and f.fecha_facturacion_programada <= hoy
    ]

    obligaciones = list(
        ObligacionTrabajador.objects
        .exclude(estado=ObligacionTrabajador.ESTADO_ANULADO)
        .select_related("trabajador", "trabajador__user", "contrato", "contrato__cliente")
        .prefetch_related("pagos")
        .filter(fecha_pago_programada__lte=hoy)
        .order_by("fecha_pago_programada", "trabajador_id", "id")
    )
    obligaciones_pendientes = [o for o in obligaciones if o.saldo > 0]
    nomina_por_trabajador = {}
    for obligacion in obligaciones_pendientes:
        clave = obligacion.trabajador_id
        fila = nomina_por_trabajador.setdefault(
            clave,
            {
                "trabajador": obligacion.trabajador,
                "saldo": Decimal("0.00"),
                "cantidad": 0,
                "fecha_mas_antigua": obligacion.fecha_pago_programada,
            },
        )
        fila["saldo"] += obligacion.saldo
        fila["cantidad"] += 1
        if obligacion.fecha_pago_programada < fila["fecha_mas_antigua"]:
            fila["fecha_mas_antigua"] = obligacion.fecha_pago_programada
    nomina_pendiente = sorted(
        nomina_por_trabajador.values(),
        key=lambda item: (item["fecha_mas_antigua"], str(item["trabajador"])),
    )

    mantenimientos_hoy_qs = (
        Mantenimiento.objects
        .filter(fecha=hoy)
        .select_related("cliente", "contrato")
        .prefetch_related("trabajadores")
        .order_by("estado", "id")
    )
    mantenimientos_hoy = list(mantenimientos_hoy_qs)
    mantenimientos_atrasados = list(
        Mantenimiento.objects
        .filter(fecha__lt=hoy, estado="pendiente")
        .select_related("cliente", "contrato")
        .prefetch_related("trabajadores")
        .order_by("fecha", "id")[:10]
    )
    mantenimientos_sin_asignar = list(
        Mantenimiento.objects
        .filter(fecha__lte=hoy, estado="pendiente", trabajadores__isnull=True)
        .select_related("cliente", "contrato")
        .order_by("fecha", "id")
        .distinct()[:10]
    )

    contratos_programacion = list(
        Contrato.objects
        .filter(activo=True, generacion_automatica=True)
        .filter(models.Q(programado_hasta__isnull=True) | models.Q(programado_hasta__lte=limite_programacion))
        .select_related("cliente", "tecnico_designado", "tecnico_designado__user")
        .order_by("programado_hasta", "id")[:10]
    )

    actividad_reciente = []
    if ActividadSistema is not None:
        actividad_reciente = list(
            ActividadSistema.objects.select_related("user").all()[:8]
        )

    total_acciones = (
        len(cobros_vencidos)
        + len(cobros_hoy)
        + len(facturas_por_emitir)
        + len(nomina_pendiente)
        + len(mantenimientos_atrasados)
        + len(mantenimientos_sin_asignar)
        + len(contratos_programacion)
    )

    return {
        "hoy": hoy,
        "cobros_vencidos": cobros_vencidos[:6],
        "cobros_hoy": cobros_hoy[:6],
        "cobros_proximos": cobros_proximos[:5],
        "cantidad_cobros_vencidos": len(cobros_vencidos),
        "cantidad_cobros_hoy": len(cobros_hoy),
        "facturas_por_emitir": facturas_por_emitir[:6],
        "cantidad_facturas_por_emitir": len(facturas_por_emitir),
        "nomina_pendiente": nomina_pendiente[:6],
        "cantidad_trabajadores_pendientes": len(nomina_pendiente),
        "total_nomina_pendiente": sum((x["saldo"] for x in nomina_pendiente), Decimal("0.00")),
        "mantenimientos_hoy": mantenimientos_hoy[:8],
        "cantidad_mantenimientos_hoy": len(mantenimientos_hoy),
        "cantidad_mantenimientos_pendientes_hoy": sum(1 for m in mantenimientos_hoy if m.estado == "pendiente"),
        "cantidad_mantenimientos_realizados_hoy": sum(1 for m in mantenimientos_hoy if m.estado == "realizado"),
        "mantenimientos_atrasados": mantenimientos_atrasados,
        "cantidad_mantenimientos_atrasados": Mantenimiento.objects.filter(fecha__lt=hoy, estado="pendiente").count(),
        "mantenimientos_sin_asignar": mantenimientos_sin_asignar,
        "cantidad_mantenimientos_sin_asignar": Mantenimiento.objects.filter(fecha__lte=hoy, estado="pendiente", trabajadores__isnull=True).distinct().count(),
        "contratos_programacion": contratos_programacion,
        "cantidad_contratos_programacion": len(contratos_programacion),
        "actividad_reciente": actividad_reciente,
        "total_acciones": total_acciones,
    }


# -------------------
# INICIO por rol (menú)
# -------------------
@login_required
def inicio_view(request):
    ctx = {"VAPID_PUBLIC_KEY": getattr(settings, "VAPID_PUBLIC_KEY", "")}

    if es_admin(request.user):
        ctx["es_admin"] = True
        notificar_movimientos_recurrentes_proximos()
        notificar_trabajadores_mantenimientos_hoy()
        try:
            from finanzas.alertas_financieras import generar_alertas_financieras
            generar_alertas_financieras(enviar_push=True)
        except Exception:
            logger.exception("No se pudieron actualizar las alertas financieras al abrir el Centro de Acciones.")
        ctx.update(_centro_acciones_contexto())
        return render(request, "dashboard/home_admin.html", ctx)

    if es_trabajador(request.user):
        ctx["es_admin"] = False
        notificar_trabajadores_mantenimientos_hoy()
        return render(request, "dashboard/home_trabajador.html", ctx)

    return render(request, "dashboard/no_autorizado.html", status=403)


# -------------------
# /dashboard/home/ = pantalla REAL por rol
# -------------------
@login_required
def home_view(request):
    return dashboard_view(request)


# -------------------
# /dashboard/ = alias de la pantalla REAL por rol
# -------------------
@login_required
def dashboard_view(request):
    base_ctx = {"VAPID_PUBLIC_KEY": getattr(settings, "VAPID_PUBLIC_KEY", "")}

    if es_admin(request.user):
        hoy = date.today()
        notificar_movimientos_recurrentes_proximos()
        notificar_trabajadores_mantenimientos_hoy()

        total_ingresos = Ingreso.objects.aggregate(total=Sum("total"))["total"] or 0
        total_egresos = Egreso.objects.aggregate(total=Sum("total"))["total"] or 0
        balance = total_ingresos - total_egresos

        ingresos_hoy = Ingreso.objects.filter(fecha=hoy).aggregate(total=Sum("total"))["total"] or 0
        egresos_hoy = Egreso.objects.filter(fecha=hoy).aggregate(total=Sum("total"))["total"] or 0
        balance_hoy = ingresos_hoy - egresos_hoy

        primer_dia_mes_actual, ultimo_dia_mes_actual = _inicio_fin_mes(hoy.year, hoy.month)
        anio_mes_anterior, mes_mes_anterior = _mes_anterior(hoy.year, hoy.month)
        primer_dia_mes_anterior, ultimo_dia_mes_anterior = _inicio_fin_mes(anio_mes_anterior, mes_mes_anterior)

        resumen_mes_actual = _resumen_financiero_rango(primer_dia_mes_actual, ultimo_dia_mes_actual)
        resumen_mes_anterior = _resumen_financiero_rango(primer_dia_mes_anterior, ultimo_dia_mes_anterior)

        variacion_ingresos_mes = _variacion_porcentual(
            resumen_mes_actual["ingresos"],
            resumen_mes_anterior["ingresos"],
        )
        variacion_egresos_mes = _variacion_porcentual(
            resumen_mes_actual["egresos"],
            resumen_mes_anterior["egresos"],
        )
        variacion_balance_mes = _variacion_porcentual(
            resumen_mes_actual["balance"],
            resumen_mes_anterior["balance"],
        )

        analitica_pro = _kpis_analitica_pro(hoy)

        recurrentes_proximos_3_dias = list(
            MovimientoRecurrente.objects.filter(
                activo=True,
                proxima_fecha__gte=hoy,
                proxima_fecha__lte=hoy + timedelta(days=3)
            ).order_by("proxima_fecha", "id")[:10]
        )

        actualizar_facturas_vencidas()
        total_facturas_pendientes = Factura.objects.filter(estado=Factura.ESTADO_PENDIENTE).count()
        total_facturas_vencidas = Factura.objects.filter(estado=Factura.ESTADO_VENCIDA).count()
        total_facturas_pagadas = Factura.objects.filter(estado=Factura.ESTADO_PAGADA).count()

        mantenimientos_hoy_qs = (
            Mantenimiento.objects.filter(fecha=hoy)
            .select_related("cliente", "contrato")
            .prefetch_related("trabajadores")
            .order_by("estado", "id")
        )

        total_mantenimientos_hoy = mantenimientos_hoy_qs.count()
        realizados_hoy = mantenimientos_hoy_qs.filter(estado="realizado").count()
        pendientes_hoy = mantenimientos_hoy_qs.filter(estado="pendiente").count()

        trabajadores_activos_hoy = (
            mantenimientos_hoy_qs
            .filter(trabajadores__isnull=False)
            .values("trabajadores")
            .distinct()
            .count()
        )

        if total_mantenimientos_hoy > 0:
            cumplimiento_hoy = round((realizados_hoy / total_mantenimientos_hoy) * 100, 2)
        else:
            cumplimiento_hoy = 0

        if cumplimiento_hoy >= 80:
            rendimiento_estado_clase = "success"
        elif cumplimiento_hoy >= 50:
            rendimiento_estado_clase = "warning"
        else:
            rendimiento_estado_clase = "danger"

        mantenimientos_atrasados_qs = (
            Mantenimiento.objects.filter(
                fecha__lt=hoy,
                estado="pendiente",
            )
            .select_related("cliente", "contrato")
            .prefetch_related("trabajadores")
            .order_by("fecha", "id")
        )

        pendientes_sin_asignar_qs = Mantenimiento.objects.filter(
            estado="pendiente",
            trabajadores__isnull=True,
        ).distinct()

        atrasados_sin_asignar_qs = Mantenimiento.objects.filter(
            fecha__lt=hoy,
            estado="pendiente",
            trabajadores__isnull=True,
        ).distinct()

        total_atrasados = mantenimientos_atrasados_qs.count()
        total_pendientes_sin_asignar = pendientes_sin_asignar_qs.count()
        total_atrasados_sin_asignar = atrasados_sin_asignar_qs.count()

        pendientes_hoy_items = list(
            mantenimientos_hoy_qs.filter(estado="pendiente")[:5]
        )

        sin_asignar_hoy_items = list(
            mantenimientos_hoy_qs.filter(estado="pendiente", trabajadores__isnull=True).distinct()[:5]
        )

        atrasados_urgentes_items = list(
            mantenimientos_atrasados_qs[:5]
        )

        requiere_atencion_items = []
        requiere_atencion_ids = set()

        for m in list(atrasados_sin_asignar_qs[:3]):
            if m.id not in requiere_atencion_ids:
                m.es_atrasado = True
                m.sin_asignar = True
                requiere_atencion_items.append(m)
                requiere_atencion_ids.add(m.id)

        for m in sin_asignar_hoy_items:
            if m.id not in requiere_atencion_ids:
                m.es_atrasado = m.fecha < hoy
                m.sin_asignar = True
                requiere_atencion_items.append(m)
                requiere_atencion_ids.add(m.id)

        for m in atrasados_urgentes_items:
            if m.id not in requiere_atencion_ids and len(requiere_atencion_items) < 5:
                m.es_atrasado = True
                try:
                    m.sin_asignar = not m.trabajadores.exists()
                except Exception:
                    m.sin_asignar = False
                requiere_atencion_items.append(m)
                requiere_atencion_ids.add(m.id)

        total_requieren_atencion = len(requiere_atencion_items)

        atrasados = list(mantenimientos_atrasados_qs)
        dia_list = list(mantenimientos_hoy_qs)
        proximos = list(
            Mantenimiento.objects.filter(fecha__gt=hoy, estado="pendiente")
            .select_related("cliente", "contrato")
            .prefetch_related("trabajadores")
            .order_by("fecha", "id")[:50]
        )

        resumen_trabajadores = _resumen_trabajadores_desde_listas(dia_list, atrasados, proximos)

        top_trabajadores = []
        trabajadores_libres = []
        trabajadores_media = []
        trabajadores_saturados = []

        for item in resumen_trabajadores:
            carga_hoy = item.get("dia", 0)
            atrasados_t = item.get("atrasados", 0)
            proximos_t = item.get("proximos", 0)
            carga_total = carga_hoy + atrasados_t + proximos_t
            estado = _clasificar_estado_trabajador(carga_hoy, atrasados_t, proximos_t)

            trabajador_data = {
                "id": item.get("trabajadores__id"),
                "username": item.get("trabajadores__user__username"),
                "carga_hoy": carga_hoy,
                "atrasados": atrasados_t,
                "proximos": proximos_t,
                "carga_total": carga_total,
                "estado": estado,
            }

            top_trabajadores.append(trabajador_data)

            if estado == "libre":
                trabajadores_libres.append(trabajador_data)
            elif estado == "media":
                trabajadores_media.append(trabajador_data)
            else:
                trabajadores_saturados.append(trabajador_data)

        top_trabajadores = top_trabajadores[:5]
        trabajadores_libres = trabajadores_libres[:5]
        trabajadores_media = trabajadores_media[:5]
        trabajadores_saturados = trabajadores_saturados[:5]

        total_trabajadores_libres = len([
            t for t in resumen_trabajadores
            if _clasificar_estado_trabajador(t.get("dia", 0), t.get("atrasados", 0), t.get("proximos", 0)) == "libre"
        ])
        total_trabajadores_media = len([
            t for t in resumen_trabajadores
            if _clasificar_estado_trabajador(t.get("dia", 0), t.get("atrasados", 0), t.get("proximos", 0)) == "media"
        ])
        total_trabajadores_saturados = len([
            t for t in resumen_trabajadores
            if _clasificar_estado_trabajador(t.get("dia", 0), t.get("atrasados", 0), t.get("proximos", 0)) == "saturado"
        ])

        trabajador_recomendado = None
        candidatos_recomendados = []

        for item in resumen_trabajadores:
            carga_hoy = item.get("dia", 0)
            atrasados_t = item.get("atrasados", 0)
            proximos_t = item.get("proximos", 0)
            estado = _clasificar_estado_trabajador(carga_hoy, atrasados_t, proximos_t)

            if estado != "saturado":
                candidatos_recomendados.append({
                    "id": item.get("trabajadores__id"),
                    "username": item.get("trabajadores__user__username"),
                    "carga_hoy": carga_hoy,
                    "atrasados": atrasados_t,
                    "proximos": proximos_t,
                    "carga_total": carga_hoy + atrasados_t + proximos_t,
                })

        if candidatos_recomendados:
            mejor = min(
                candidatos_recomendados,
                key=lambda x: (x["atrasados"], x["carga_hoy"], x["carga_total"], x["username"])
            )

            razones = []
            if mejor["carga_hoy"] == 0:
                razones.append("sin mantenimientos hoy")
            elif mejor["carga_hoy"] == 1:
                razones.append("solo tiene 1 mantenimiento hoy")
            else:
                razones.append(f"tiene {mejor['carga_hoy']} mantenimientos hoy")

            if mejor["atrasados"] == 0:
                razones.append("sin atrasados")
            else:
                razones.append(f"{mejor['atrasados']} atrasados")

            razones.append(f"carga total {mejor['carga_total']}")

            mejor["motivo"] = " · ".join(razones)
            trabajador_recomendado = mejor

        actividades_recientes = []
        if ActividadSistema is not None:
            actividades_recientes = list(ActividadSistema.objects.select_related("user").all()[:5])

        grafico_finanzas = {
            "labels": ["Ingresos", "Egresos", "Balance"],
            "data": [
                float(resumen_mes_actual.get("ingresos", 0) or 0),
                float(resumen_mes_actual.get("egresos", 0) or 0),
                float(resumen_mes_actual.get("balance", 0) or 0),
            ],
        }
        grafico_operativo = {
            "labels": ["Realizados hoy", "Pendientes hoy", "Atrasados", "Sin asignar"],
            "data": [
                int(realizados_hoy),
                int(pendientes_hoy),
                int(total_atrasados),
                int(total_pendientes_sin_asignar),
            ],
        }
        grafico_tendencia_financiera = {
            "labels": analitica_pro["serie_financiera_6m"]["labels"],
            "datasets": [
                {"label": "Ingresos", "data": analitica_pro["serie_financiera_6m"]["ingresos"]},
                {"label": "Egresos", "data": analitica_pro["serie_financiera_6m"]["egresos"]},
                {"label": "Balance", "data": analitica_pro["serie_financiera_6m"]["balance"]},
            ],
        }
        grafico_tendencia_operativa = {
            "labels": analitica_pro["serie_operativa_6m"]["labels"],
            "datasets": [
                {"label": "Realizados", "data": analitica_pro["serie_operativa_6m"]["realizados"]},
                {"label": "Pendientes", "data": analitica_pro["serie_operativa_6m"]["pendientes"]},
                {"label": "Atrasados", "data": analitica_pro["serie_operativa_6m"]["atrasados"]},
            ],
        }

        ctx = {
            **base_ctx,
            "modo": "admin",
            "hoy": hoy,
            "total_ingresos": float(total_ingresos),
            "total_egresos": float(total_egresos),
            "balance": float(balance),
            "ingresos_hoy": float(ingresos_hoy),
            "egresos_hoy": float(egresos_hoy),
            "balance_hoy": float(balance_hoy),
            "resumen_mes_actual": resumen_mes_actual,
            "resumen_mes_anterior": resumen_mes_anterior,
            "variacion_ingresos_mes": variacion_ingresos_mes,
            "variacion_egresos_mes": variacion_egresos_mes,
            "variacion_balance_mes": variacion_balance_mes,
            "recurrentes_proximos_3_dias": recurrentes_proximos_3_dias,
            "total_facturas_pendientes": total_facturas_pendientes,
            "total_facturas_vencidas": total_facturas_vencidas,
            "total_facturas_pagadas": total_facturas_pagadas,
            "total_mantenimientos_hoy": total_mantenimientos_hoy,
            "realizados_hoy": realizados_hoy,
            "pendientes_hoy": pendientes_hoy,
            "trabajadores_activos_hoy": trabajadores_activos_hoy,
            "cumplimiento_hoy": cumplimiento_hoy,
            "rendimiento_estado_clase": rendimiento_estado_clase,
            "total_atrasados": total_atrasados,
            "total_pendientes_sin_asignar": total_pendientes_sin_asignar,
            "total_atrasados_sin_asignar": total_atrasados_sin_asignar,
            "pendientes_hoy_items": pendientes_hoy_items,
            "sin_asignar_hoy_items": sin_asignar_hoy_items,
            "atrasados_urgentes_items": atrasados_urgentes_items,
            "requiere_atencion_items": requiere_atencion_items,
            "total_requieren_atencion": total_requieren_atencion,
            "top_trabajadores": top_trabajadores,
            "trabajadores_libres": trabajadores_libres,
            "trabajadores_media": trabajadores_media,
            "trabajadores_saturados": trabajadores_saturados,
            "total_trabajadores_libres": total_trabajadores_libres,
            "total_trabajadores_media": total_trabajadores_media,
            "total_trabajadores_saturados": total_trabajadores_saturados,
            "trabajador_recomendado": trabajador_recomendado,
            "hay_alertas_operativas": (
                total_atrasados > 0
                or total_pendientes_sin_asignar > 0
                or total_atrasados_sin_asignar > 0
            ),
            "actividades_recientes": actividades_recientes,
            "grafico_finanzas": json.dumps(grafico_finanzas),
            "grafico_operativo": json.dumps(grafico_operativo),
            "grafico_tendencia_financiera": json.dumps(grafico_tendencia_financiera),
            "grafico_tendencia_operativa": json.dumps(grafico_tendencia_operativa),

            # Analítica PRO
            "kpi_ticket_promedio_mes": analitica_pro["ticket_promedio_mes"],
            "kpi_clientes_activos_mes": analitica_pro["clientes_activos_mes"],
            "kpi_ingresos_mes_count": analitica_pro["ingresos_mes_count"],
            "kpi_mantenimientos_realizados_mes": analitica_pro["mantenimientos_realizados_mes"],
            "kpi_mantenimientos_pendientes_mes": analitica_pro["mantenimientos_pendientes_mes"],
            "kpi_mantenimientos_atrasados_mes": analitica_pro["mantenimientos_atrasados_mes"],
            "kpi_total_mantenimientos_mes": analitica_pro["total_mantenimientos_mes"],
            "kpi_cumplimiento_mes": analitica_pro["cumplimiento_mes"],

            "tendencia_ingresos_mes": _tendencia_desde_variacion(analitica_pro["variacion_ingresos_mes"]),
            "tendencia_egresos_mes": _tendencia_desde_variacion(analitica_pro["variacion_egresos_mes"]),
            "tendencia_balance_mes": _tendencia_desde_variacion(analitica_pro["variacion_balance_mes"]),
            "estado_variacion_ingresos_mes": _estado_variacion_financiera(analitica_pro["variacion_ingresos_mes"]),
            "estado_variacion_egresos_mes": _estado_variacion_financiera(analitica_pro["variacion_egresos_mes"], invertido=True),
            "estado_variacion_balance_mes": _estado_variacion_financiera(analitica_pro["variacion_balance_mes"]),

            "top_clientes_facturacion": analitica_pro["top_clientes_facturacion"],
            "top_contratos_facturacion": analitica_pro["top_contratos_facturacion"],
            "top_trabajadores_mes": analitica_pro["top_trabajadores_mes"],
            "top_insumos_mes": analitica_pro["top_insumos_mes"],
            "serie_financiera_6m": analitica_pro["serie_financiera_6m"],
            "serie_operativa_6m": analitica_pro["serie_operativa_6m"],

            "es_admin": True,
        }
        return render(request, "dashboard/dashboard.html", ctx)

    if es_trabajador(request.user):
        hoy = date.today()
        notificar_trabajadores_mantenimientos_hoy()

        try:
            trabajador = request.user.trabajador
        except Exception:
            return render(request, "dashboard/no_autorizado.html", status=403)

        anio_cal = int(request.GET.get("anio_cal", hoy.year))
        mes_cal = int(request.GET.get("mes_cal", hoy.month))
        anio_cal_ant, mes_cal_ant = _mes_anterior(anio_cal, mes_cal)
        anio_cal_sig, mes_cal_sig = _mes_siguiente(anio_cal, mes_cal)

        fecha_seleccionada_str = (request.GET.get("fecha_seleccionada", "") or "").strip()
        fecha_seleccionada = parse_date(fecha_seleccionada_str) if fecha_seleccionada_str else None

        vista_actual = (request.GET.get("vista", "calendario") or "calendario").strip().lower()
        ver_mas_proximos = request.GET.get("ver_mas_proximos") == "1"

        calendario_trabajador = _build_calendario_mantenimientos(
            anio_cal,
            mes_cal,
            trabajador=trabajador
        )

        if fecha_seleccionada:
            mantenimientos_hoy = (
                Mantenimiento.objects.filter(
                    fecha=fecha_seleccionada,
                    trabajadores=trabajador
                )
                .select_related("cliente", "contrato")
                .order_by("estado", "fecha", "id")
            )

            mantenimientos_atrasados = Mantenimiento.objects.none()
            mantenimientos_proximos = Mantenimiento.objects.none()
            total_proximos_reales = 0
            mostrar_boton_ver_mas_proximos = False

        else:
            mantenimientos_hoy = (
                Mantenimiento.objects.filter(fecha=hoy, trabajadores=trabajador)
                .select_related("cliente", "contrato")
                .order_by("estado", "fecha", "id")
            )

            qs_mantenimientos_proximos = (
                Mantenimiento.objects.filter(fecha__gt=hoy, trabajadores=trabajador)
                .select_related("cliente", "contrato")
                .order_by("fecha", "id")
            )
            total_proximos_reales = qs_mantenimientos_proximos.count()

            if ver_mas_proximos:
                mantenimientos_proximos = qs_mantenimientos_proximos
            else:
                mantenimientos_proximos = qs_mantenimientos_proximos[:10]

            mostrar_boton_ver_mas_proximos = total_proximos_reales > 10

            mantenimientos_atrasados = (
                Mantenimiento.objects.filter(
                    fecha__lt=hoy,
                    estado="pendiente",
                    trabajadores=trabajador
                )
                .select_related("cliente", "contrato")
                .order_by("fecha", "id")[:20]
            )

        # ==========================================
        # NUEVO: agenda semanal trabajador
        # ==========================================
        fecha_base_agenda = fecha_seleccionada or hoy
        inicio_agenda = fecha_base_agenda - timedelta(days=fecha_base_agenda.weekday())
        fin_agenda = inicio_agenda + timedelta(days=6)

        agenda_items = list(
            Mantenimiento.objects.filter(
                fecha__range=(inicio_agenda, fin_agenda),
                trabajadores=trabajador
            )
            .select_related("cliente", "contrato")
            .prefetch_related("trabajadores")
            .order_by("fecha", "estado", "id")
        )

        agenda_semanal = _build_agenda_semanal_mantenimientos(
            fecha_base=fecha_base_agenda,
            items=agenda_items,
        )

        semana_anterior = inicio_agenda - timedelta(days=7)
        semana_siguiente = inicio_agenda + timedelta(days=7)

        ctx = {
            **base_ctx,
            "modo": "trabajador",
            "hoy": hoy,
            "vista_actual": vista_actual,
            "mantenimientos_hoy": mantenimientos_hoy,
            "mantenimientos_proximos": mantenimientos_proximos,
            "total_proximos_reales": total_proximos_reales,
            "mostrar_boton_ver_mas_proximos": mostrar_boton_ver_mas_proximos,
            "ver_mas_proximos": ver_mas_proximos,
            "mantenimientos_atrasados": mantenimientos_atrasados,
            "anio_cal": anio_cal,
            "mes_cal": mes_cal,
            "anio_cal_ant": anio_cal_ant,
            "mes_cal_ant": mes_cal_ant,
            "anio_cal_sig": anio_cal_sig,
            "mes_cal_sig": mes_cal_sig,
            "calendario_trabajador": calendario_trabajador,
            "fecha_seleccionada": fecha_seleccionada,
            "fecha_seleccionada_str": fecha_seleccionada_str,
            "agenda_semanal": agenda_semanal,
            "agenda_semana_anterior": semana_anterior,
            "agenda_semana_siguiente": semana_siguiente,
            "es_admin": False,
        }
        return render(request, "dashboard/dashboard_trabajador.html", ctx)


# -------------------
# Pantalla de notificaciones
# -------------------
@login_required
def notificaciones_view(request):
    subs_count = 0
    push_enabled = False

    if PushSubscription is not None:
        subs_count = PushSubscription.objects.filter(user=request.user).count()
        push_enabled = subs_count > 0

    notificaciones = []
    no_modelo_notificaciones = False

    if Notificacion is not None:
        notificaciones = list(
            Notificacion.objects.filter(user=request.user)
            .order_by("-creada_en")[:20]
        )

        ids_no_leidas = [n.id for n in notificaciones if not n.leida]
        if ids_no_leidas:
            ahora = timezone.now()
            Notificacion.objects.filter(
                id__in=ids_no_leidas,
                user=request.user,
                leida=False,
            ).update(
                leida=True,
                leida_en=ahora,
            )

            for n in notificaciones:
                if n.id in ids_no_leidas:
                    n.leida = True
                    n.leida_en = ahora
    else:
        no_modelo_notificaciones = True

    return render(
        request,
        "dashboard/notificaciones.html",
        {
            "subs_count": subs_count,
            "push_enabled": push_enabled,
            "notificaciones": notificaciones,
            "no_modelo_notificaciones": no_modelo_notificaciones,
            "es_admin": es_admin(request.user),
            "base_template": "dashboard/base_admin.html" if es_admin(request.user) else "dashboard/base_trabajador.html",
        },
    )


@login_required
@require_GET
def notificaciones_json_view(request):
    if es_admin(request.user):
        try:
            from finanzas.alertas_financieras import generar_alertas_financieras
            generar_alertas_financieras(enviar_push=True)
        except Exception:
            logger.exception("No se pudieron actualizar las alertas financieras.")

    if Notificacion is None:
        return JsonResponse({
            "ok": True,
            "items": [],
            "unread_count": 0,
        })

    qs = Notificacion.objects.filter(user=request.user).order_by("-creada_en")[:10]

    items = []
    for n in qs:
        items.append({
            "id": n.id,
            "titulo": n.titulo,
            "mensaje": n.mensaje,
            "url": n.url or "/dashboard/notificaciones/",
            "leida": n.leida,
            "creada_en": n.creada_en.strftime("%d/%m/%Y %H:%M"),
        })

    unread_count = Notificacion.objects.filter(
        user=request.user,
        leida=False
    ).count()

    return JsonResponse({
        "ok": True,
        "items": items,
        "unread_count": unread_count,
    })


@login_required
def notificaciones_historial_view(request):
    if Notificacion is None:
        page_obj = None
    else:
        qs = Notificacion.objects.filter(user=request.user).order_by("-creada_en")
        paginator = Paginator(qs, 15)
        page_number = request.GET.get("page")
        page_obj = paginator.get_page(page_number)

    return render(
        request,
        "dashboard/notificaciones_historial.html",
        {
            "page_obj": page_obj,
            "es_admin": es_admin(request.user),
            "base_template": "dashboard/base_admin.html" if es_admin(request.user) else "dashboard/base_trabajador.html",
        },
    )


@login_required
@require_http_methods(["POST"])
def marcar_notificacion_leida_view(request, pk):
    if Notificacion is None:
        return JsonResponse({"ok": False, "error": "Modelo no disponible"}, status=500)

    notificacion = get_object_or_404(Notificacion, pk=pk, user=request.user)

    if not notificacion.leida:
        notificacion.leida = True
        notificacion.leida_en = timezone.now()
        notificacion.save(update_fields=["leida", "leida_en"])

    return JsonResponse({"ok": True})


@login_required
@require_http_methods(["POST"])
def notificacion_eliminar_view(request, pk):
    if Notificacion is None:
        return JsonResponse({"ok": False, "error": "Modelo no disponible"}, status=500)

    notificacion = get_object_or_404(Notificacion, pk=pk, user=request.user)
    notificacion.delete()

    return JsonResponse({"ok": True})


@login_required
@require_http_methods(["POST"])
def notificaciones_eliminar_todas_view(request):
    if Notificacion is None:
        return JsonResponse({"ok": False, "error": "Modelo no disponible"}, status=500)

    Notificacion.objects.filter(user=request.user).delete()
    return JsonResponse({"ok": True})


@login_required
@require_http_methods(["POST"])
def marcar_todas_leidas_view(request):
    if Notificacion is None:
        return JsonResponse({"ok": False, "error": "Modelo no disponible"}, status=500)

    ahora = timezone.now()

    Notificacion.objects.filter(
        user=request.user,
        leida=False,
    ).update(
        leida=True,
        leida_en=ahora,
    )

    return JsonResponse({"ok": True})


@login_required
def actividad_historial_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    if ActividadSistema is None:
        page_obj = None
    else:
        qs = ActividadSistema.objects.select_related("user").all()
        paginator = Paginator(qs, 20)
        page_number = request.GET.get("page")
        page_obj = paginator.get_page(page_number)

    return render(
        request,
        "dashboard/actividad_historial.html",
        {
            "page_obj": page_obj,
            "es_admin": True,
        },
    )


# -------------------
# Historial mantenimientos
# -------------------
@login_required
def mantenimiento_historial_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    hoy = date.today()

    q = (request.GET.get("q", "") or "").strip()
    estado = (request.GET.get("estado", "") or "").strip().lower()
    filtro = (request.GET.get("filtro", "") or "").strip().lower()
    cliente_id = (request.GET.get("cliente", "") or "").strip()
    trabajador_id = (request.GET.get("trabajador", "") or "").strip()
    fecha_desde_str = (request.GET.get("fecha_desde", "") or "").strip()
    fecha_hasta_str = (request.GET.get("fecha_hasta", "") or "").strip()

    fecha_desde = parse_date(fecha_desde_str) if fecha_desde_str else None
    fecha_hasta = parse_date(fecha_hasta_str) if fecha_hasta_str else None

    qs = (
        Mantenimiento.objects
        .select_related("cliente", "contrato")
        .prefetch_related("trabajadores")
        .order_by("-fecha", "-id")
    )

    if filtro == "hoy":
        qs = qs.filter(fecha=hoy)
    elif filtro == "pendientes":
        qs = qs.filter(estado="pendiente")
        estado = "pendiente"
    elif filtro == "realizados":
        qs = qs.filter(estado="realizado")
        estado = "realizado"
    elif filtro == "atrasados":
        qs = qs.filter(fecha__lt=hoy, estado="pendiente")
    elif filtro == "sin_asignar":
        qs = qs.filter(trabajadores__isnull=True).distinct()

    if estado in ["pendiente", "realizado"]:
        qs = qs.filter(estado=estado)

    if cliente_id.isdigit():
        qs = qs.filter(cliente_id=int(cliente_id))

    if trabajador_id.isdigit():
        qs = qs.filter(trabajadores__id=int(trabajador_id))

    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)

    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    qs = qs.distinct()
    items = list(qs)

    if q:
        items = _filtrar_mantenimientos_por_busqueda(items, q)

    total_historial = len(items)
    total_realizados_historial = len([m for m in items if getattr(m, "estado", "") == "realizado"])
    total_pendientes_historial = len([m for m in items if getattr(m, "estado", "") == "pendiente"])
    total_atrasados_historial = len([
        m for m in items
        if getattr(m, "estado", "") == "pendiente" and getattr(m, "fecha", hoy) < hoy
    ])
    total_sin_asignar_historial = len([
        m for m in items
        if not m.trabajadores.exists()
    ])

    paginator = Paginator(items, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    clientes_filtro = []
    clientes_ids = set()
    for m in Mantenimiento.objects.select_related("cliente").all().order_by("cliente_id"):
        cid = getattr(m, "cliente_id", None)
        if cid and cid not in clientes_ids:
            clientes_ids.add(cid)
            clientes_filtro.append({
                "id": cid,
                "nombre": str(getattr(m, "cliente", "")),
            })

    trabajadores_filtro = list(
        Trabajador.objects.select_related("user").all().order_by("user__username")
    )

    query_params = request.GET.copy()
    if "page" in query_params:
        query_params.pop("page")
    querystring = query_params.urlencode()

    return render(
        request,
        "dashboard/mantenimientos_historial.html",
        {
            "page_obj": page_obj,
            "q": q,
            "estado": estado,
            "filtro": filtro,
            "cliente_id": cliente_id,
            "trabajador_id": trabajador_id,
            "fecha_desde": fecha_desde_str,
            "fecha_hasta": fecha_hasta_str,
            "clientes_filtro": clientes_filtro,
            "trabajadores_filtro": trabajadores_filtro,
            "total_historial": total_historial,
            "total_realizados_historial": total_realizados_historial,
            "total_pendientes_historial": total_pendientes_historial,
            "total_atrasados_historial": total_atrasados_historial,
            "total_sin_asignar_historial": total_sin_asignar_historial,
            "querystring": querystring,
            "es_admin": True,
        },
    )


# -------------------
# Operativo Admin
# -------------------
@login_required
def admin_operativo_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    hoy = date.today()
    filtro = (request.GET.get("filtro", "") or "").strip().lower()
    q = (request.GET.get("q", "") or "").strip()
    ver_mas_proximos = (request.GET.get("ver_mas_proximos", "") or "").strip() == "1"
    vista_actual = (request.GET.get("vista", "calendario") or "calendario").strip().lower()

    fecha_seleccionada_str = (request.GET.get("fecha_seleccionada", "") or "").strip()
    fecha_seleccionada = parse_date(fecha_seleccionada_str) if fecha_seleccionada_str else None

    anio_cal = int(request.GET.get("anio_cal", hoy.year))
    mes_cal = int(request.GET.get("mes_cal", hoy.month))
    anio_cal_ant, mes_cal_ant = _mes_anterior(anio_cal, mes_cal)
    anio_cal_sig, mes_cal_sig = _mes_siguiente(anio_cal, mes_cal)

    base_qs = (
        Mantenimiento.objects
        .select_related("cliente", "contrato")
        .prefetch_related("trabajadores")
        .order_by("fecha", "estado", "id")
    )

    if fecha_seleccionada:
        dia_list = list(
            base_qs.filter(fecha=fecha_seleccionada)
        )
        atrasados = []
        proximos = []
        etiqueta_periodo = f"Agenda del {fecha_seleccionada.strftime('%d/%m/%Y')}"

    elif filtro == "atrasados":
        dia_list = []
        atrasados = list(
            base_qs.filter(
                fecha__lt=hoy,
                estado="pendiente"
            )
        )
        proximos = []
        etiqueta_periodo = "Mantenimientos atrasados"

    elif filtro == "sin_asignar":
        dia_list = list(
            base_qs.filter(
                fecha=hoy,
                estado="pendiente",
                trabajadores__isnull=True
            ).distinct()
        )
        atrasados = list(
            base_qs.filter(
                fecha__lt=hoy,
                estado="pendiente",
                trabajadores__isnull=True
            ).distinct()
        )
        proximos = []
        etiqueta_periodo = "Mantenimientos sin asignar"

    elif filtro == "pendientes_hoy":
        dia_list = list(
            base_qs.filter(
                fecha=hoy,
                estado="pendiente"
            )
        )
        atrasados = []
        proximos = []
        etiqueta_periodo = "Pendientes de hoy"

    elif filtro == "urgentes":
        dia_list = list(
            base_qs.filter(
                fecha=hoy,
                estado="pendiente",
                trabajadores__isnull=True
            ).distinct()
        )
        atrasados = list(
            base_qs.filter(
                fecha__lt=hoy,
                estado="pendiente"
            )
        )
        proximos = []
        etiqueta_periodo = "Requieren atención inmediata"

    else:
        dia_list = list(
            base_qs.filter(fecha=hoy)
        )
        atrasados = list(
            base_qs.filter(
                fecha__lt=hoy,
                estado="pendiente"
            )
        )

        limite_proximos = None if ver_mas_proximos else 10
        qs_proximos = base_qs.filter(
            fecha__gt=hoy,
            estado="pendiente"
        )

        if limite_proximos is not None:
            qs_proximos = qs_proximos[:limite_proximos]

        proximos = list(qs_proximos)
        etiqueta_periodo = "Operativo de hoy"

    if q:
        dia_list = _filtrar_mantenimientos_por_busqueda(dia_list, q)
        atrasados = _filtrar_mantenimientos_por_busqueda(atrasados, q)
        proximos = _filtrar_mantenimientos_por_busqueda(proximos, q)

    resumen_trabajadores = _resumen_trabajadores_desde_listas(dia_list, atrasados, proximos)

    sin_asignar_dia = _sin_asignar_count(dia_list)
    sin_asignar_atrasados = _sin_asignar_count(atrasados)
    sin_asignar_proximos = _sin_asignar_count(proximos)

    total_sin_asignar = (
        sin_asignar_dia +
        sin_asignar_atrasados +
        sin_asignar_proximos
    )

    calendario_operativo = _build_calendario_mantenimientos(anio_cal, mes_cal)

    total_proximos_reales = base_qs.filter(fecha__gt=hoy, estado="pendiente").count()
    mostrar_boton_ver_mas_proximos = (
        not filtro and not ver_mas_proximos and not fecha_seleccionada and total_proximos_reales > 10
    )

    # ==========================================
    # NUEVO: agenda semanal PRO
    # ==========================================
    fecha_base_agenda = fecha_seleccionada or hoy
    inicio_agenda = fecha_base_agenda - timedelta(days=fecha_base_agenda.weekday())
    fin_agenda = inicio_agenda + timedelta(days=6)

    agenda_items = list(
        base_qs.filter(fecha__range=(inicio_agenda, fin_agenda))
    )

    if q:
        agenda_items = _filtrar_mantenimientos_por_busqueda(agenda_items, q)

    agenda_semanal = _build_agenda_semanal_mantenimientos(
        fecha_base=fecha_base_agenda,
        items=agenda_items,
    )

    semana_anterior = inicio_agenda - timedelta(days=7)
    semana_siguiente = inicio_agenda + timedelta(days=7)

    return render(
        request,
        "dashboard/admin_operativo.html",
        {
            "hoy": hoy,
            "q": q,
            "modo_actual": filtro,
            "vista_actual": vista_actual,
            "etiqueta_periodo": etiqueta_periodo,
            "dia_list": dia_list,
            "atrasados": atrasados,
            "proximos": proximos,
            "resumen_trabajadores": resumen_trabajadores,
            "sin_asignar_dia": sin_asignar_dia,
            "sin_asignar_atrasados": sin_asignar_atrasados,
            "sin_asignar_proximos": sin_asignar_proximos,
            "total_sin_asignar": total_sin_asignar,
            "anio_cal": anio_cal,
            "mes_cal": mes_cal,
            "anio_cal_ant": anio_cal_ant,
            "mes_cal_ant": mes_cal_ant,
            "anio_cal_sig": anio_cal_sig,
            "mes_cal_sig": mes_cal_sig,
            "calendario_operativo": calendario_operativo,
            "ver_mas_proximos": ver_mas_proximos,
            "mostrar_boton_ver_mas_proximos": mostrar_boton_ver_mas_proximos,
            "total_proximos_reales": total_proximos_reales,
            "fecha_seleccionada": fecha_seleccionada,
            "fecha_seleccionada_str": fecha_seleccionada_str,
            "agenda_semanal": agenda_semanal,
            "agenda_fecha_base": fecha_base_agenda,
            "agenda_semana_anterior": semana_anterior,
            "agenda_semana_siguiente": semana_siguiente,
            "es_admin": True,
        },
    )


# -------------------
# Detalle mantenimiento
# -------------------
@login_required
def mantenimiento_detalle_view(request, pk):
    """Detalle unificado del mantenimiento.

    Inspección, limpieza, consumos, fotografías y cierre se procesan en un
    único POST. Las actividades son opcionales; únicamente las tres fotos
    requeridas son obligatorias para finalizar.
    """
    mantenimiento = get_object_or_404(Mantenimiento, pk=pk)

    if es_admin(request.user):
        permitido = True
    elif es_trabajador(request.user):
        try:
            trabajador = request.user.trabajador
            permitido = mantenimiento.trabajadores.filter(pk=trabajador.pk).exists()
        except Exception:
            permitido = False
    else:
        permitido = False

    if not permitido:
        return render(request, "dashboard/no_autorizado.html", status=403)

    es_usuario_admin = es_admin(request.user)
    trabajador_actual = None
    if es_trabajador(request.user):
        try:
            trabajador_actual = request.user.trabajador
        except Exception:
            trabajador_actual = None

    if trabajador_actual:
        inventario_mantenimiento = list(
            InventarioTrabajador.objects.filter(
                trabajador=trabajador_actual,
                stock__gt=0,
                insumo__activo=True,
                insumo__puede_mantenimiento=True,
            ).select_related("insumo").order_by("insumo__nombre")
        )
        insumos = [x.insumo for x in inventario_mantenimiento]
    else:
        inventario_mantenimiento = []
        insumos = list(
            Insumo.objects.filter(activo=True, puede_mantenimiento=True).order_by("nombre")
        )

    trabajadores_mantenimiento = list(
        mantenimiento.trabajadores.filter(activo=True)
        .select_related("user")
        .order_by("user__username")
    )
    esta_realizado = mantenimiento.estado == "realizado"
    checklist, _ = ChecklistMantenimiento.objects.get_or_create(mantenimiento=mantenimiento)

    cliente_operativo = mantenimiento.cliente
    direccion_operativa = (cliente_operativo.direccion or "").strip()
    maps_url = (cliente_operativo.enlace_google_maps or "").strip()
    if not maps_url and direccion_operativa:
        maps_url = "https://www.google.com/maps/search/?api=1&query=" + quote(direccion_operativa)
    whatsapp_disponible = bool(
        trabajador_actual
        and _telefono_whatsapp_ecuador(cliente_operativo.telefono)
    )

    if request.method == "POST":
        accion = (request.POST.get("accion") or "").strip()
        next_url = (request.POST.get("next", "") or "").strip()

        def safe_return_url():
            if next_url and url_has_allowed_host_and_scheme(
                url=next_url,
                allowed_hosts={request.get_host()},
                require_https=not settings.DEBUG,
            ):
                return next_url
            return f"/dashboard/mantenimientos/{mantenimiento.pk}/"

        if accion == "marcar_pendiente":
            mantenimiento.estado = "pendiente"
            mantenimiento.borrador_guardado = True
            mantenimiento.save(update_fields=["estado", "borrador_guardado"])

            actor = request.user.username
            _notificar_admins(
                titulo="🟡 Mantenimiento pendiente",
                mensaje=f"El mantenimiento de {mantenimiento.cliente} fue marcado como pendiente por {actor}.",
                url=f"/dashboard/mantenimientos/{mantenimiento.pk}/",
                enviar_push=False,
                excluir_user_id=request.user.id if es_usuario_admin else None,
            )
            _registrar_actividad(
                user=request.user,
                titulo="Mantenimiento pendiente",
                descripcion=f"{actor} marcó como pendiente el mantenimiento de {mantenimiento.cliente}.",
                url=f"/dashboard/mantenimientos/{mantenimiento.pk}/",
            )
            messages.success(request, f"Mantenimiento de {mantenimiento.cliente} marcado como pendiente.")
            return redirect(safe_return_url())

        if accion not in {"guardar_borrador_unificado", "finalizar_unificado"}:
            messages.error(request, "Acción de mantenimiento no válida.")
            return redirect(safe_return_url())

        if esta_realizado:
            messages.error(
                request,
                "Este mantenimiento está realizado y bloqueado para cambios. Debes volverlo a pendiente para editarlo.",
            )
            return redirect(safe_return_url())

        finalizar = accion == "finalizar_unificado"
        consumos_creados = []
        fotos_subidas = []

        # Validación previa: al finalizar comprobamos las evidencias antes de
        # descontar inventario o escribir archivos. Así evitamos cualquier
        # efecto parcial incluso fuera de la transacción de base de datos.
        if finalizar:
            nombres_existentes = {
                f.descripcion
                for f in mantenimiento.fotos.all()
                if _nombre_foto_valido(f.descripcion)
            }
            nombres_nuevos = set()
            if request.FILES.get("foto_inicio"):
                nombres_nuevos.add("Inicio de Mantenimiento")
            if request.FILES.get("foto_fin"):
                nombres_nuevos.add("Fin de Mantenimiento")
            if request.FILES.get("foto_nivel"):
                nombres_nuevos.add("Nivel PH y Cl")
            faltantes_previos = [
                nombre for nombre in FOTOS_REQUERIDAS
                if nombre not in nombres_existentes | nombres_nuevos
            ]
            if faltantes_previos:
                messages.error(
                    request,
                    "Para finalizar debes cargar las 3 fotografías obligatorias. "
                    f"Faltan: {', '.join(faltantes_previos)}.",
                )
                return redirect(safe_return_url())

        try:
            with transaction.atomic():
                # ---------------------------------------------------------
                # 1) Inspección y limpieza (opcionales)
                # ---------------------------------------------------------
                campos_bool = [
                    "aspirado",
                    "cepillado",
                    "recoleccion_basura",
                    "limpieza_filtros",
                    "retrolavado_arena",
                    "limpieza_filos",
                ]
                for campo in campos_bool:
                    setattr(checklist, campo, request.POST.get(campo) == "on")

                checklist.bomba_estado = request.POST.get("bomba_estado", "")
                checklist.bomba_novedad = request.POST.get("bomba_novedad", "").strip()
                checklist.filtro_estado = request.POST.get("filtro_estado", "")
                checklist.filtro_novedad = request.POST.get("filtro_novedad", "").strip()
                checklist.nivel_agua = request.POST.get("nivel_agua", "")

                # Los campos químicos del checklist quedan obsoletos desde
                # Inventario Inteligente. Se limpian al volver a guardar un
                # mantenimiento pendiente para no duplicar información.
                checklist.cloro_granulado = False
                checklist.tricloro = False
                checklist.alguicida = False
                checklist.metasilicato = False
                checklist.floculante = False
                checklist.save()

                # Estado del agua se conserva en el campo compatible ya
                # existente, pero ahora forma parte de Inspección Inicial.
                mantenimiento.estado_agua_rapido = request.POST.get("estado_agua", "")
                mantenimiento.equipo_rapido = ""
                mantenimiento.recomendaciones_rapidas = []
                mantenimiento.observaciones = request.POST.get(
                    "observaciones", mantenimiento.observaciones
                ).strip()
                mantenimiento.borrador_guardado = not finalizar
                mantenimiento.save(
                    update_fields=[
                        "estado_agua_rapido",
                        "equipo_rapido",
                        "recomendaciones_rapidas",
                        "observaciones",
                        "borrador_guardado",
                    ]
                )

                # ---------------------------------------------------------
                # 2) Productos utilizados. Se pueden agregar varias filas
                #    dentro del mismo formulario.
                # ---------------------------------------------------------
                ids_insumo = request.POST.getlist("producto_insumo_id")
                cantidades = request.POST.getlist("producto_cantidad")
                unidades = request.POST.getlist("producto_unidad")

                if trabajador_actual:
                    trabajador_consumo = trabajador_actual
                else:
                    trabajador_consumo_id = (request.POST.get("trabajador_consumo_id") or "").strip()
                    trabajador_consumo = None
                    if any(x.strip() for x in ids_insumo):
                        if not trabajador_consumo_id:
                            raise ValueError("Selecciona el trabajador que utilizó los productos.")
                        trabajador_consumo = get_object_or_404(
                            Trabajador, pk=trabajador_consumo_id, activo=True
                        )
                        if not mantenimiento.trabajadores.filter(pk=trabajador_consumo.pk).exists():
                            raise ValueError(
                                "El trabajador seleccionado no está asignado a este mantenimiento."
                            )

                for index, insumo_id in enumerate(ids_insumo):
                    insumo_id = (insumo_id or "").strip()
                    if not insumo_id:
                        continue

                    cantidad_ingresada = cantidades[index] if index < len(cantidades) else ""
                    unidad = unidades[index] if index < len(unidades) else "kg"
                    insumo = get_object_or_404(
                        Insumo,
                        pk=insumo_id,
                        activo=True,
                        puede_mantenimiento=True,
                    )

                    cantidad_base = convertir_a_base(insumo, cantidad_ingresada, unidad)
                    movimiento = consumir_trabajador(
                        insumo=insumo,
                        trabajador=trabajador_consumo,
                        cantidad_base=cantidad_base,
                        mantenimiento=mantenimiento,
                        usuario=request.user,
                    )
                    uso = UsoInsumo.objects.create(
                        mantenimiento=mantenimiento,
                        insumo=insumo,
                        trabajador=trabajador_consumo,
                        cantidad=cantidad_base,
                        cantidad_ingresada=Decimal(str(cantidad_ingresada).replace(",", ".")),
                        unidad_registro=unidad,
                        costo_unitario=movimiento.costo_unitario,
                        costo_total=movimiento.total_costo,
                    )
                    consumos_creados.append(uso)

                # ---------------------------------------------------------
                # 3) Fotografías requeridas. Las nuevas se guardan junto a
                #    todo el formulario; las ya existentes se respetan.
                # ---------------------------------------------------------
                mapa_fotos = [
                    ("Inicio de Mantenimiento", request.FILES.get("foto_inicio")),
                    ("Fin de Mantenimiento", request.FILES.get("foto_fin")),
                    ("Nivel PH y Cl", request.FILES.get("foto_nivel")),
                ]
                existentes = {
                    f.descripcion: f
                    for f in mantenimiento.fotos.all()
                    if _nombre_foto_valido(f.descripcion)
                }
                for tipo_foto, imagen in mapa_fotos:
                    if not imagen or tipo_foto in existentes:
                        continue
                    FotoMantenimiento.objects.create(
                        mantenimiento=mantenimiento,
                        imagen=imagen,
                        descripcion=tipo_foto,
                    )
                    fotos_subidas.append(tipo_foto)
                    existentes[tipo_foto] = True

                # ---------------------------------------------------------
                # 4) Finalización. Solo las 3 fotografías son obligatorias.
                #    Si falta alguna, toda la transacción se revierte,
                #    incluyendo consumos de inventario añadidos en este POST.
                # ---------------------------------------------------------
                if finalizar:
                    faltantes = [nombre for nombre in FOTOS_REQUERIDAS if nombre not in existentes]
                    if faltantes:
                        raise ValueError(
                            "Para finalizar debes cargar las 3 fotografías obligatorias. "
                            f"Faltan: {', '.join(faltantes)}."
                        )
                    mantenimiento.estado = "realizado"
                    mantenimiento.borrador_guardado = False
                    mantenimiento.save(update_fields=["estado", "borrador_guardado"])

        except (ValueError, InventarioTrabajador.DoesNotExist) as exc:
            messages.error(request, str(exc) or "No se pudo guardar el mantenimiento.")
            return redirect(safe_return_url())
        except Exception:
            logger.exception("Error guardando mantenimiento unificado id=%s", mantenimiento.pk)
            messages.error(
                request,
                "Ocurrió un error al guardar. No se aplicaron cambios parciales; inténtalo nuevamente.",
            )
            return redirect(safe_return_url())

        actor = request.user.username
        if finalizar:
            _notificar_admins(
                titulo="✅ Mantenimiento realizado",
                mensaje=f"El mantenimiento de {mantenimiento.cliente} fue finalizado por {actor}.",
                url=f"/dashboard/mantenimientos/{mantenimiento.pk}/",
                enviar_push=True,
                excluir_user_id=request.user.id if es_usuario_admin else None,
            )
            _registrar_actividad(
                user=request.user,
                titulo="Mantenimiento realizado",
                descripcion=(
                    f"{actor} finalizó el mantenimiento de {mantenimiento.cliente}. "
                    f"Nuevos consumos: {len(consumos_creados)}; fotos nuevas: {len(fotos_subidas)}."
                ),
                url=f"/dashboard/mantenimientos/{mantenimiento.pk}/",
            )
            messages.success(request, f"Mantenimiento de {mantenimiento.cliente} finalizado correctamente.")
        else:
            _registrar_actividad(
                user=request.user,
                titulo="Borrador de mantenimiento guardado",
                descripcion=(
                    f"{actor} guardó el borrador de {mantenimiento.cliente}. "
                    f"Nuevos consumos: {len(consumos_creados)}; fotos nuevas: {len(fotos_subidas)}."
                ),
                url=f"/dashboard/mantenimientos/{mantenimiento.pk}/",
            )
            messages.success(request, "Borrador guardado. Todo lo registrado quedó conservado.")

        return redirect(safe_return_url())

    lista_usos = mantenimiento.usos_insumos.select_related("insumo", "trabajador__user").all()
    lista_egresos = mantenimiento.egresos.all() if hasattr(mantenimiento, "egresos") else []
    total_egresos = sum((Decimal(u.subtotal()) for u in lista_usos), Decimal("0.00"))

    fotos_qs = mantenimiento.fotos.all()
    fotos_por_nombre = {
        f.descripcion: f for f in fotos_qs if _nombre_foto_valido(f.descripcion)
    }
    fotos = [fotos_por_nombre[nombre] for nombre in FOTOS_REQUERIDAS if nombre in fotos_por_nombre]

    historial_cliente_reciente = []
    if es_usuario_admin:
        historial_cliente_reciente = (
            Mantenimiento.objects
            .filter(cliente=mantenimiento.cliente)
            .exclude(pk=mantenimiento.pk)
            .select_related("cliente", "contrato")
            .prefetch_related("trabajadores")
            .order_by("-fecha", "-id")[:5]
        )

    cantidad_fotos = len(fotos)
    cantidad_usos = lista_usos.count()
    checklist_limpieza_completados = sum(
        bool(valor)
        for valor in [
            checklist.aspirado,
            checklist.cepillado,
            checklist.recoleccion_basura,
            checklist.limpieza_filtros,
            checklist.retrolavado_arena,
            checklist.limpieza_filos,
        ]
    )
    checklist_inspeccion_completados = sum(
        bool(valor)
        for valor in [
            checklist.bomba_estado,
            checklist.filtro_estado,
            checklist.nivel_agua,
            mantenimiento.estado_agua_rapido,
        ]
    )
    checklist_completados = checklist_limpieza_completados + checklist_inspeccion_completados
    checklist_total = 10
    checklist_porcentaje = round((checklist_completados / checklist_total) * 100)

    puede_cerrar = cantidad_fotos == 3
    foto_inicio = fotos_por_nombre.get("Inicio de Mantenimiento")
    foto_fin = fotos_por_nombre.get("Fin de Mantenimiento")
    foto_nivel = fotos_por_nombre.get("Nivel PH y Cl")

    return render(
        request,
        "dashboard/mantenimiento_detalle.html",
        {
            "m": mantenimiento,
            "insumos": insumos,
            "lista_usos": lista_usos,
            "lista_egresos": lista_egresos,
            "inventario_mantenimiento": inventario_mantenimiento,
            "trabajadores_mantenimiento": trabajadores_mantenimiento,
            "trabajador_actual": trabajador_actual,
            "total_egresos": total_egresos,
            "es_admin": es_usuario_admin,
            "fotos": fotos,
            "cantidad_fotos": cantidad_fotos,
            "cantidad_usos": cantidad_usos,
            "puede_cerrar": puede_cerrar,
            "esta_realizado": esta_realizado,
            "foto_inicio": foto_inicio,
            "foto_fin": foto_fin,
            "foto_nivel": foto_nivel,
            "historial_cliente_reciente": historial_cliente_reciente,
            "checklist": checklist,
            "checklist_limpieza_completados": checklist_limpieza_completados,
            "checklist_inspeccion_completados": checklist_inspeccion_completados,
            "checklist_completados": checklist_completados,
            "checklist_total": checklist_total,
            "checklist_porcentaje": checklist_porcentaje,
            "cliente_operativo": cliente_operativo,
            "maps_url": maps_url,
            "whatsapp_disponible": whatsapp_disponible,
        },
    )


@login_required
@require_GET
def mantenimiento_whatsapp_cliente_view(request, pk):
    """Abre WhatsApp para un técnico asignado y registra el intento de contacto."""
    mantenimiento = get_object_or_404(
        Mantenimiento.objects.select_related("cliente"),
        pk=pk,
    )

    if not es_trabajador(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    try:
        trabajador = request.user.trabajador
    except Exception:
        return render(request, "dashboard/no_autorizado.html", status=403)

    if not mantenimiento.trabajadores.filter(pk=trabajador.pk).exists():
        return render(request, "dashboard/no_autorizado.html", status=403)

    telefono = _telefono_whatsapp_ecuador(mantenimiento.cliente.telefono)
    if not telefono:
        messages.error(request, "El cliente no tiene un número de WhatsApp registrado.")
        return redirect("mantenimiento_detalle", pk=mantenimiento.pk)

    nombre_trabajador = (request.user.get_full_name() or "").strip() or request.user.username
    mensaje = (
        f"Hola, buenos días. Soy {nombre_trabajador}, técnico de JVAQUA. "
        "Me encuentro aquí en su domicilio para realizar el mantenimiento de su piscina."
    )

    _registrar_actividad(
        user=request.user,
        titulo="Acceso rápido de WhatsApp utilizado",
        descripcion=(
            f"{nombre_trabajador} abrió WhatsApp para contactar a "
            f"{mantenimiento.cliente} desde el mantenimiento #{mantenimiento.pk}."
        ),
        url=f"/dashboard/mantenimientos/{mantenimiento.pk}/",
    )

    return redirect(f"https://wa.me/{telefono}?text={quote(mensaje)}")


@login_required
def mi_cuenta_trabajador_view(request):
    if not es_trabajador(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    trabajador = get_object_or_404(Trabajador.objects.select_related("user"), user=request.user)
    hoy = timezone.localdate()
    try:
        anio = int(request.GET.get("anio") or hoy.year)
        mes = int(request.GET.get("mes") or hoy.month)
        if mes < 1 or mes > 12:
            raise ValueError
    except (TypeError, ValueError):
        anio, mes = hoy.year, hoy.month

    obligaciones = list(
        ObligacionTrabajador.objects.filter(
            trabajador=trabajador,
            fecha_pago_programada__year=anio,
            fecha_pago_programada__month=mes,
        )
        .exclude(estado=ObligacionTrabajador.ESTADO_ANULADO)
        .select_related("contrato__cliente").prefetch_related("pagos")
        .order_by("fecha_pago_programada", "id")
    )
    generado = sum((o.valor_acordado for o in obligaciones), Decimal("0.00"))
    pagado = sum((o.monto_pagado for o in obligaciones), Decimal("0.00"))
    pendiente = sum((o.saldo for o in obligaciones), Decimal("0.00"))
    anticipos = list(AnticipoTrabajador.objects.filter(trabajador=trabajador).order_by("-fecha", "-id")[:30])
    anticipos_pendientes = sum((a.monto for a in anticipos if not a.descontado), Decimal("0.00"))
    proximas = [o.fecha_pago_programada for o in obligaciones if o.saldo > 0 and o.fecha_pago_programada]
    proximo_pago = min(proximas) if proximas else None
    contratos = list(Contrato.objects.filter(tecnico_designado=trabajador, activo=True).select_related("cliente").order_by("cliente__nombre"))
    mantenimientos_mes = Mantenimiento.objects.filter(trabajadores=trabajador, fecha__year=anio, fecha__month=mes)
    realizados_mes = mantenimientos_mes.filter(estado="realizado").count()
    pendientes_mes = mantenimientos_mes.filter(estado="pendiente").count()
    pagos_recientes = list(PagoTrabajador.objects.filter(obligacion__trabajador=trabajador, activo=True).select_related("obligacion__contrato__cliente").order_by("-fecha", "-id")[:30])
    lotes = list(LotePagoTrabajador.objects.filter(trabajador=trabajador, activo=True).order_by("-fecha", "-id")[:20])
    inventario_personal = list(
        InventarioTrabajador.objects.filter(trabajador=trabajador, stock__gt=0)
        .select_related("insumo").order_by("insumo__nombre")
    )
    movimientos_inventario_personal = list(
        MovimientoInventario.objects.filter(trabajador=trabajador)
        .select_related("insumo", "mantenimiento__cliente")
        .order_by("-creado_en", "-id")[:30]
    )
    return render(request, "dashboard/mi_cuenta_trabajador.html", {
        "trabajador": trabajador, "anio": anio, "mes": mes, "obligaciones": obligaciones,
        "generado": generado, "pagado": pagado, "pendiente": pendiente,
        "anticipos": anticipos, "anticipos_pendientes": anticipos_pendientes,
        "proximo_pago": proximo_pago, "contratos": contratos,
        "realizados_mes": realizados_mes, "pendientes_mes": pendientes_mes,
        "pagos_recientes": pagos_recientes, "lotes": lotes,
        "inventario_personal": inventario_personal,
        "movimientos_inventario_personal": movimientos_inventario_personal,
        "VAPID_PUBLIC_KEY": getattr(settings, "VAPID_PUBLIC_KEY", ""),
        "meses": [(1,"Enero"),(2,"Febrero"),(3,"Marzo"),(4,"Abril"),(5,"Mayo"),(6,"Junio"),(7,"Julio"),(8,"Agosto"),(9,"Septiembre"),(10,"Octubre"),(11,"Noviembre"),(12,"Diciembre")],
    })


@login_required
def foto_mantenimiento_eliminar_view(request, pk):
    foto = get_object_or_404(FotoMantenimiento, pk=pk)
    mantenimiento = foto.mantenimiento

    if es_admin(request.user):
        permitido = True
    elif es_trabajador(request.user):
        try:
            trabajador = request.user.trabajador
            permitido = mantenimiento.trabajadores.filter(pk=trabajador.pk).exists()
        except Exception:
            permitido = False
    else:
        permitido = False

    if not permitido:
        return render(request, "dashboard/no_autorizado.html", status=403)

    if mantenimiento.estado == "realizado":
        messages.error(request, "Este mantenimiento está realizado y bloqueado para cambios. Debes volverlo a pendiente para editarlo.")
        return redirect(f"/dashboard/mantenimientos/{mantenimiento.pk}/")

    if request.method == "POST":
        actor = request.user.username
        cliente_nombre = str(mantenimiento.cliente)
        foto_id = foto.pk
        foto_nombre = foto.descripcion or "foto"

        try:
            if foto.imagen:
                foto.imagen.delete(save=False)
        except Exception:
            logger.exception("No se pudo borrar el archivo físico de la foto id=%s", foto_id)

        foto.delete()

        _notificar_admins(
            titulo="🗑 Foto eliminada",
            mensaje=f"{actor} eliminó la foto '{foto_nombre}' del mantenimiento de {cliente_nombre}.",
            url=f"/dashboard/mantenimientos/{mantenimiento.pk}/",
            enviar_push=False,
            excluir_user_id=request.user.id if es_admin(request.user) else None,
        )
        _registrar_actividad(
            user=request.user,
            titulo="Foto eliminada",
            descripcion=f"{actor} eliminó la foto '{foto_nombre}' del mantenimiento de {cliente_nombre}.",
            url=f"/dashboard/mantenimientos/{mantenimiento.pk}/",
        )

        messages.success(request, "Foto eliminada correctamente.")
        return redirect(f"/dashboard/mantenimientos/{mantenimiento.pk}/")

    return redirect(f"/dashboard/mantenimientos/{mantenimiento.pk}/")


@login_required
def usoinsumo_eliminar_view(request, pk):
    uso = get_object_or_404(UsoInsumo, pk=pk)
    mantenimiento = uso.mantenimiento

    if es_admin(request.user):
        permitido = True
    elif es_trabajador(request.user):
        try:
            trabajador = request.user.trabajador
            permitido = mantenimiento.trabajadores.filter(pk=trabajador.pk).exists()
        except Exception:
            permitido = False
    else:
        permitido = False

    if not permitido:
        return render(request, "dashboard/no_autorizado.html", status=403)

    if mantenimiento.estado == "realizado":
        messages.error(request, "Este mantenimiento está realizado y bloqueado para cambios. Debes volverlo a pendiente para editarlo.")
        return redirect(f"/dashboard/mantenimientos/{mantenimiento.pk}/")

    if request.method == "POST":
        insumo = uso.insumo
        insumo_nombre = getattr(insumo, "nombre", "Insumo")
        cantidad = uso.cantidad

        if uso.trabajador_id:
            revertir_consumo(uso=uso, usuario=request.user)
        elif getattr(uso, "egreso_id", None):
            # Registro histórico anterior al inventario por trabajador.
            insumo.stock = Decimal(insumo.stock) + Decimal(uso.cantidad)
            insumo.save(update_fields=["stock"])

        if getattr(uso, "egreso_id", None):
            uso.egreso.delete()

        uso.delete()

        actor = request.user.username
        _notificar_admins(
            titulo="🗑 Insumo eliminado",
            mensaje=f"{actor} eliminó un uso de insumo en {mantenimiento.cliente}.",
            url=f"/dashboard/mantenimientos/{mantenimiento.pk}/",
            enviar_push=False,
            excluir_user_id=request.user.id if es_admin(request.user) else None,
        )
        _registrar_actividad(
            user=request.user,
            titulo="Insumo eliminado",
            descripcion=f"{actor} eliminó {insumo_nombre} x {cantidad} del mantenimiento de {mantenimiento.cliente}.",
            url=f"/dashboard/mantenimientos/{mantenimiento.pk}/",
        )

        messages.success(request, "Uso de insumo eliminado y stock devuelto.")
        return redirect(f"/dashboard/mantenimientos/{mantenimiento.pk}/")

    return render(
        request,
        "dashboard/usoinsumo_confirmar_eliminar.html",
        {"uso": uso, "es_admin": es_admin(request.user)},
    )


@login_required
def usoinsumo_editar_view(request, pk):
    uso = get_object_or_404(UsoInsumo, pk=pk)
    mantenimiento = uso.mantenimiento

    if es_admin(request.user):
        permitido = True
    elif es_trabajador(request.user):
        try:
            trabajador = request.user.trabajador
            permitido = mantenimiento.trabajadores.filter(pk=trabajador.pk).exists()
        except Exception:
            permitido = False
    else:
        permitido = False

    if not permitido:
        return render(request, "dashboard/no_autorizado.html", status=403)

    if mantenimiento.estado == "realizado":
        messages.error(request, "Este mantenimiento está realizado y bloqueado para cambios. Debes volverlo a pendiente para editarlo.")
        return redirect(f"/dashboard/mantenimientos/{mantenimiento.pk}/")

    if request.method == "POST":
        nueva_cantidad_str = request.POST.get("cantidad", "").strip()
        unidad = request.POST.get("unidad", uso.unidad_registro or "kg")
        try:
            nueva_base = convertir_a_base(uso.insumo, nueva_cantidad_str, unidad)
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect(f"/dashboard/usos/{uso.pk}/editar/")

        anterior = Decimal(uso.cantidad)
        if uso.trabajador_id:
            try:
                with transaction.atomic():
                    revertir_consumo(uso=uso, usuario=request.user)
                    movimiento = consumir_trabajador(
                        insumo=uso.insumo,
                        trabajador=uso.trabajador,
                        cantidad_base=nueva_base,
                        mantenimiento=mantenimiento,
                        usuario=request.user,
                        observacion=f"Edición consumo #{uso.pk}",
                    )
                    uso.cantidad = nueva_base
                    uso.cantidad_ingresada = Decimal(str(nueva_cantidad_str).replace(",", "."))
                    uso.unidad_registro = unidad
                    uso.costo_unitario = movimiento.costo_unitario
                    uso.costo_total = movimiento.total_costo
                    uso.save(update_fields=["cantidad", "cantidad_ingresada", "unidad_registro", "costo_unitario", "costo_total"])
            except (ValueError, InventarioTrabajador.DoesNotExist) as exc:
                messages.error(request, str(exc) if str(exc) else "Stock insuficiente.")
                return redirect(f"/dashboard/usos/{uso.pk}/editar/")
        else:
            # Compatibilidad para registros históricos sin trabajador.
            diff = nueva_base - anterior
            if diff > 0 and Decimal(uso.insumo.stock) < diff:
                messages.error(request, f"Stock general insuficiente. Disponible: {uso.insumo.stock}")
                return redirect(f"/dashboard/usos/{uso.pk}/editar/")
            uso.insumo.stock = Decimal(uso.insumo.stock) - diff
            uso.insumo.save(update_fields=["stock"])
            uso.cantidad = nueva_base
            uso.cantidad_ingresada = Decimal(str(nueva_cantidad_str).replace(",", "."))
            uso.unidad_registro = unidad
            uso.costo_unitario = uso.insumo.costo or 0
            uso.costo_total = (nueva_base * Decimal(uso.insumo.costo or 0)).quantize(Decimal("0.01"))
            uso.save()

        actor = request.user.username
        _registrar_actividad(
            user=request.user,
            titulo="Insumo actualizado",
            descripcion=f"{actor} actualizó {uso.insumo.nombre} de {anterior} a {nueva_base} {uso.insumo.unidad_corta} en {mantenimiento.cliente}.",
            url=f"/dashboard/mantenimientos/{mantenimiento.pk}/",
        )
        messages.success(request, "Consumo actualizado correctamente.")
        return redirect(f"/dashboard/mantenimientos/{mantenimiento.pk}/")

    return render(
        request,
        "dashboard/usoinsumo_editar.html",
        {"uso": uso, "es_admin": es_admin(request.user)},
    )


@login_required
def asignar_trabajadores_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    mantenimiento = get_object_or_404(Mantenimiento, pk=pk)
    trabajadores = list(
        Trabajador.objects.select_related("user").all().order_by("user__username")
    )
    hoy = date.today()

    trabajadores_info = []
    for trabajador in trabajadores:
        qs_base = (
            Mantenimiento.objects.filter(trabajadores=trabajador)
            .select_related("cliente", "contrato")
            .prefetch_related("trabajadores")
        )

        qs_mismo_dia = qs_base.filter(fecha=mantenimiento.fecha).exclude(pk=mantenimiento.pk).order_by("fecha", "id")
        carga_hoy = qs_mismo_dia.count()
        atrasados = qs_base.filter(fecha__lt=hoy, estado="pendiente").exclude(pk=mantenimiento.pk).count()
        proximos = qs_base.filter(fecha__gt=hoy, estado="pendiente").exclude(pk=mantenimiento.pk).count()

        clientes_mismo_dia = [str(getattr(mh, "cliente", "") or "") for mh in qs_mismo_dia[:5]]

        carga_total = carga_hoy + atrasados + proximos

        if carga_hoy == 0 and carga_total <= 1:
            carga_label = "Más libre"
            carga_badge = "success"
            choque_label = "Sin choque"
            choque_badge = "success"
        elif carga_hoy == 1:
            carga_label = "Carga media"
            carga_badge = "warning"
            choque_label = "Ocupación media"
            choque_badge = "warning"
        elif carga_hoy >= 2:
            carga_label = "Carga alta"
            carga_badge = "danger"
            choque_label = "Posible choque"
            choque_badge = "danger"
        else:
            carga_label = "Carga media"
            carga_badge = "warning"
            choque_label = "Ocupación media"
            choque_badge = "warning"

        ya_asignado = mantenimiento.trabajadores.filter(pk=trabajador.pk).exists()

        trabajadores_info.append({
            "obj": trabajador,
            "ya_asignado": ya_asignado,
            "carga_hoy": carga_hoy,
            "atrasados": atrasados,
            "proximos": proximos,
            "carga_total": carga_total,
            "carga_label": carga_label,
            "carga_badge": carga_badge,
            "choque_label": choque_label,
            "choque_badge": choque_badge,
            "clientes_mismo_dia": clientes_mismo_dia,
            "es_recomendado": False,
        })

    trabajadores_info.sort(
        key=lambda x: (
            x["ya_asignado"] is False,
            x["carga_hoy"],
            x["atrasados"],
            x["carga_total"],
            getattr(getattr(x["obj"], "user", None), "username", ""),
        )
    )

    recomendacion = None
    candidatos = [item for item in trabajadores_info if not item["ya_asignado"]]
    if not candidatos:
        candidatos = trabajadores_info

    if candidatos:
        recomendado = min(
            candidatos,
            key=lambda x: (
                x["carga_hoy"],
                x["atrasados"],
                x["carga_total"],
                getattr(getattr(x["obj"], "user", None), "username", ""),
            ),
        )
        recomendado["es_recomendado"] = True

        razones = []
        if recomendado["carga_hoy"] == 0:
            razones.append("no tiene mantenimientos ese día")
        else:
            razones.append(f"solo tiene {recomendado['carga_hoy']} ese día")

        if recomendado["atrasados"] == 0:
            razones.append("no tiene atrasados")
        else:
            razones.append(f"tiene {recomendado['atrasados']} atrasados")

        razones.append(f"carga total visible: {recomendado['carga_total']}")

        recomendacion = {
            "trabajador": recomendado["obj"],
            "motivo": " · ".join(razones),
        }

    if request.method == "POST":
        ids = request.POST.getlist("trabajadores")
        mantenimiento.trabajadores.set(ids)

        asignados = list(mantenimiento.trabajadores.select_related("user").all())
        nombres_asignados = []

        for trabajador in asignados:
            if getattr(trabajador, "user", None):
                nombres_asignados.append(trabajador.user.username)
                _crear_notificacion(
                    user=trabajador.user,
                    titulo="🛠 Nuevo mantenimiento asignado",
                    mensaje=f"Se te asignó el mantenimiento de {mantenimiento.cliente} para {mantenimiento.fecha}.",
                    url=f"/dashboard/mantenimientos/{mantenimiento.pk}/",
                    enviar_push=True,
                )

        actor = request.user.username
        detalle = ", ".join(nombres_asignados) if nombres_asignados else "sin trabajadores"
        _registrar_actividad(
            user=request.user,
            titulo="Trabajadores asignados",
            descripcion=f"{actor} actualizó la asignación del mantenimiento de {mantenimiento.cliente}: {detalle}.",
            url=f"/dashboard/mantenimientos/{mantenimiento.pk}/",
        )

        messages.success(request, "Trabajadores asignados correctamente.")
        return redirect("/dashboard/operativo/")

    return render(
        request,
        "dashboard/asignar_trabajadores.html",
        {
            "m": mantenimiento,
            "trabajadores": trabajadores,
            "trabajadores_info": trabajadores_info,
            "recomendacion": recomendacion,
            "es_admin": True,
        },
    )



# -------------------
# Facturación automática - helpers
# -------------------
def _contrato_activo_para_facturacion(contrato):
    if contrato is None:
        return False

    if hasattr(contrato, "activo"):
        return bool(contrato.activo)

    if hasattr(contrato, "estado"):
        estado = str(getattr(contrato, "estado", "") or "").strip().lower()
        if estado in ["activo", "vigente"]:
            return True
        if estado in ["inactivo", "cancelado", "finalizado"]:
            return False

    return True


def _obtener_cliente_de_contrato(contrato):
    return getattr(contrato, "cliente", None)


def _obtener_monto_contrato(contrato):
    posibles_campos = [
        "monto_mensual",
        "precio_mensual",
        "valor_mensual",
        "mensualidad",
        "monto",
        "precio",
        "valor",
        "costo",
    ]

    for campo in posibles_campos:
        if hasattr(contrato, campo):
            valor = getattr(contrato, campo, None)
            if valor is not None:
                try:
                    return Decimal(valor)
                except Exception:
                    pass

    return Decimal("0.00")


def _descripcion_factura_contrato(contrato, anio, mes):
    base = f"Servicio de mantenimiento de piscina {mes:02d}/{anio}"

    if hasattr(contrato, "nombre") and contrato.nombre:
        return f"{base} - {contrato.nombre}"

    if hasattr(contrato, "tipo") and contrato.tipo:
        return f"{base} - {contrato.tipo}"

    return base


def _dias_vencimiento_factura():
    return 5


def _crear_factura_para_contrato(contrato, anio, mes):
    cliente = _obtener_cliente_de_contrato(contrato)
    if not cliente:
        return None, False, "Contrato sin cliente"

    monto = _obtener_monto_contrato(contrato)
    if monto <= 0:
        return None, False, "Contrato sin monto válido"

    fecha_emision = date(anio, mes, 1)
    fecha_vencimiento = fecha_emision + timedelta(days=_dias_vencimiento_factura())

    factura, creada = Factura.objects.get_or_create(
        contrato=contrato,
        periodo_anio=anio,
        periodo_mes=mes,
        defaults={
            "cliente": cliente,
            "fecha_emision": fecha_emision,
            "fecha_vencimiento": fecha_vencimiento,
            "subtotal": monto,
            "impuesto": Decimal("0.00"),
            "total": monto,
            "estado": Factura.ESTADO_PENDIENTE,
            "observaciones": "",
        }
    )

    if creada:
        FacturaItem.objects.create(
            factura=factura,
            descripcion=_descripcion_factura_contrato(contrato, anio, mes),
            cantidad=Decimal("1.00"),
            precio_unitario=monto,
            subtotal=monto,
        )
        factura.actualizar_totales()

    return factura, creada, None


def generar_facturas_automaticas(anio=None, mes=None):
    hoy = timezone.localdate()
    anio = anio or hoy.year
    mes = mes or hoy.month

    creadas = 0
    existentes = 0
    errores = []

    from contratos.models import Contrato

    contratos = Contrato.objects.select_related("cliente").all().order_by("id")

    for contrato in contratos:
        if not _contrato_activo_para_facturacion(contrato):
            continue

        try:
            _, creada, error = _crear_factura_para_contrato(contrato, anio, mes)

            if error:
                errores.append(f"Contrato #{contrato.pk}: {error}")
                continue

            if creada:
                creadas += 1
            else:
                existentes += 1

        except Exception as ex:
            errores.append(f"Contrato #{contrato.pk}: {ex}")

    return {
        "creadas": creadas,
        "existentes": existentes,
        "errores": errores,
        "anio": anio,
        "mes": mes,
    }


def actualizar_facturas_vencidas():
    hoy = timezone.localdate()
    facturas = Factura.objects.filter(
        estado=Factura.ESTADO_PENDIENTE,
        fecha_vencimiento__lt=hoy,
    )

    total = 0
    for factura in facturas:
        factura.estado = Factura.ESTADO_VENCIDA
        factura.save(update_fields=["estado", "actualizada_en"])
        total += 1

    return total


# -------------------
# Facturación automática - vistas
# -------------------
@login_required
@require_GET
def factura_list_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    actualizar_facturas_vencidas()

    estado = (request.GET.get("estado", "") or "").strip().lower()
    periodo = (request.GET.get("periodo", "") or "").strip()

    facturas = Factura.objects.all().order_by("-periodo_anio", "-periodo_mes", "-id")

    if estado in [
        Factura.ESTADO_PENDIENTE,
        Factura.ESTADO_PAGADA,
        Factura.ESTADO_VENCIDA,
        Factura.ESTADO_ANULADA,
    ]:
        facturas = facturas.filter(estado=estado)

    if periodo:
        try:
            anio_txt, mes_txt = periodo.split("-")
            facturas = facturas.filter(
                periodo_anio=int(anio_txt),
                periodo_mes=int(mes_txt),
            )
        except Exception:
            pass

    total_facturado = facturas.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
    total_pagado = facturas.filter(estado=Factura.ESTADO_PAGADA).aggregate(total=Sum("total"))["total"] or Decimal("0.00")
    total_pendiente = facturas.filter(
        estado__in=[Factura.ESTADO_PENDIENTE, Factura.ESTADO_VENCIDA]
    ).aggregate(total=Sum("total"))["total"] or Decimal("0.00")

    return render(
        request,
        "dashboard/factura_list.html",
        {
            "facturas": facturas,
            "estado_actual": estado,
            "periodo_actual": periodo,
            "total_facturado": total_facturado,
            "total_pagado": total_pagado,
            "total_pendiente": total_pendiente,
            "mantenimientos_futuros": mantenimientos_futuros,
            "proximo_mantenimiento": proximo_mantenimiento,
            "es_admin": True,
        },
    )


@login_required
@require_GET
def factura_detalle_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    actualizar_facturas_vencidas()

    factura = get_object_or_404(
        Factura.objects.select_related("cliente", "contrato", "ingreso_generado").prefetch_related("items"),
        pk=pk
    )

    return render(
        request,
        "dashboard/factura_detalle.html",
        {
            "factura": factura,
            "items": factura.items.all(),
            "es_admin": True,
        },
    )


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def factura_generar_mes_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    hoy = timezone.localdate()

    try:
        anio = int(request.POST.get("anio") or hoy.year)
        mes = int(request.POST.get("mes") or hoy.month)
    except Exception:
        anio = hoy.year
        mes = hoy.month

    resultado = generar_facturas_automaticas(anio=anio, mes=mes)

    _registrar_actividad(
        user=request.user,
        titulo="Facturas generadas",
        descripcion=(
            f"{request.user.username} generó facturas del período {mes:02d}/{anio}: "
            f"{resultado['creadas']} creadas, {resultado['existentes']} existentes."
        ),
        url=f"/dashboard/finanzas/facturas/?periodo={anio}-{mes:02d}",
    )

    if resultado["creadas"]:
        messages.success(
            request,
            f"Facturación generada: {resultado['creadas']} factura(s) creadas para {mes:02d}/{anio}."
        )
    else:
        messages.info(
            request,
            f"No se crearon facturas nuevas para {mes:02d}/{anio}. Ya existían."
        )

    if resultado["errores"]:
        messages.warning(
            request,
            f"Se encontraron {len(resultado['errores'])} contrato(s) con observaciones."
        )

    return redirect(f"/dashboard/finanzas/facturas/?periodo={anio}-{mes:02d}")


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def factura_marcar_pagada_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    factura = get_object_or_404(Factura, pk=pk)

    if factura.estado == Factura.ESTADO_ANULADA:
        messages.warning(request, "No puedes marcar como pagada una factura anulada.")
        return redirect(f"/dashboard/finanzas/facturas/{factura.pk}/")

    fecha_pago_txt = (request.POST.get("fecha_pago", "") or "").strip()
    fecha_pago = timezone.localdate()

    if fecha_pago_txt:
        try:
            fecha_pago = date.fromisoformat(fecha_pago_txt)
        except Exception:
            pass

    factura.marcar_como_pagada(fecha_pago=fecha_pago)

    _registrar_actividad(
        user=request.user,
        titulo="Factura pagada",
        descripcion=f"{request.user.username} marcó como pagada la factura {factura.numero}.",
        url=f"/dashboard/finanzas/facturas/{factura.pk}/",
    )

    messages.success(request, f"Factura {factura.numero} marcada como pagada.")
    return redirect(f"/dashboard/finanzas/facturas/{factura.pk}/")


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def factura_anular_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    factura = get_object_or_404(Factura, pk=pk)

    if factura.estado == Factura.ESTADO_PAGADA:
        messages.warning(request, "No se puede anular una factura que ya fue pagada.")
        return redirect(f"/dashboard/finanzas/facturas/{factura.pk}/")

    factura.estado = Factura.ESTADO_ANULADA
    factura.save(update_fields=["estado", "actualizada_en"])

    _registrar_actividad(
        user=request.user,
        titulo="Factura anulada",
        descripcion=f"{request.user.username} anuló la factura {factura.numero}.",
        url=f"/dashboard/finanzas/facturas/{factura.pk}/",
    )

    messages.success(request, f"Factura {factura.numero} anulada correctamente.")
    return redirect(f"/dashboard/finanzas/facturas/{factura.pk}/")

@login_required
def flujo_mensual_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    hoy = timezone.localdate()

    try:
        anio = int(request.GET.get("anio", hoy.year))
        mes = int(request.GET.get("mes", hoy.month))
    except ValueError:
        anio, mes = hoy.year, hoy.month

    primer_dia = date(anio, mes, 1)
    ultimo_dia = date(anio, mes, monthrange(anio, mes)[1])

    ingresos_qs = Ingreso.objects.filter(fecha__range=(primer_dia, ultimo_dia)).order_by("-fecha", "-id")
    egresos_qs = Egreso.objects.filter(fecha__range=(primer_dia, ultimo_dia)).order_by("-fecha", "-id")

    ingresos_manuales_qs = ingresos_qs.filter(cliente__isnull=True, contrato__isnull=True)
    egresos_manuales_qs = egresos_qs.filter(mantenimiento__isnull=True, insumo__isnull=True)

    total_ingresos = ingresos_qs.aggregate(total=Sum("total"))["total"] or Decimal("0")
    total_egresos = egresos_qs.aggregate(total=Sum("total"))["total"] or Decimal("0")

    total_ingresos_manuales = ingresos_manuales_qs.aggregate(total=Sum("total"))["total"] or Decimal("0")
    total_egresos_manuales = egresos_manuales_qs.aggregate(total=Sum("total"))["total"] or Decimal("0")

    total_ingresos_automaticos = total_ingresos - total_ingresos_manuales
    total_egresos_automaticos = total_egresos - total_egresos_manuales

    balance = total_ingresos - total_egresos

    # Resumen del día (movido desde dashboard)
    ingresos_hoy_qs = Ingreso.objects.filter(fecha=hoy).order_by("-id")
    egresos_hoy_qs = Egreso.objects.filter(fecha=hoy).order_by("-id")

    total_ingresos_hoy = ingresos_hoy_qs.aggregate(total=Sum("total"))["total"] or Decimal("0")
    total_egresos_hoy = egresos_hoy_qs.aggregate(total=Sum("total"))["total"] or Decimal("0")
    balance_hoy = total_ingresos_hoy - total_egresos_hoy

    total_ingresos_hoy_manuales = ingresos_hoy_qs.filter(cliente__isnull=True, contrato__isnull=True).aggregate(total=Sum("total"))["total"] or Decimal("0")
    total_egresos_hoy_manuales = egresos_hoy_qs.filter(mantenimiento__isnull=True, insumo__isnull=True).aggregate(total=Sum("total"))["total"] or Decimal("0")

    total_ingresos_hoy_automaticos = total_ingresos_hoy - total_ingresos_hoy_manuales
    total_egresos_hoy_automaticos = total_egresos_hoy - total_egresos_hoy_manuales

    actualizar_facturas_vencidas()
    facturas_mes = Factura.objects.filter(
        periodo_anio=anio,
        periodo_mes=mes,
    ).select_related("cliente", "contrato", "ingreso_generado").order_by("-id")

    total_facturas_mes = facturas_mes.aggregate(total=Sum("total"))["total"] or Decimal("0")
    total_facturas_pagadas_mes = facturas_mes.filter(
        estado=Factura.ESTADO_PAGADA
    ).aggregate(total=Sum("total"))["total"] or Decimal("0")
    total_facturas_pendientes_mes = facturas_mes.filter(
        estado__in=[Factura.ESTADO_PENDIENTE, Factura.ESTADO_VENCIDA]
    ).aggregate(total=Sum("total"))["total"] or Decimal("0")
    total_facturas_vencidas = facturas_mes.filter(estado=Factura.ESTADO_VENCIDA).count()
    total_facturas_pagadas_count = facturas_mes.filter(estado=Factura.ESTADO_PAGADA).count()
    total_facturas_pendientes_count = facturas_mes.filter(
        estado__in=[Factura.ESTADO_PENDIENTE, Factura.ESTADO_VENCIDA]
    ).count()

    # Recurrentes próximos (movidos desde dashboard)
    recurrentes_proximos_3_dias = list(
        MovimientoRecurrente.objects.filter(
            activo=True,
            proxima_fecha__gte=hoy,
            proxima_fecha__lte=hoy + timedelta(days=3)
        ).order_by("proxima_fecha", "id")[:10]
    )

    # Mes anterior
    mes_anterior = mes - 1 or 12
    anio_anterior = anio - 1 if mes == 1 else anio

    primer_dia_ant = date(anio_anterior, mes_anterior, 1)
    ultimo_dia_ant = date(anio_anterior, mes_anterior, monthrange(anio_anterior, mes_anterior)[1])

    ingresos_ant = Ingreso.objects.filter(fecha__range=(primer_dia_ant, ultimo_dia_ant))
    egresos_ant = Egreso.objects.filter(fecha__range=(primer_dia_ant, ultimo_dia_ant))

    total_ingresos_ant = ingresos_ant.aggregate(total=Sum("total"))["total"] or Decimal("0")
    total_egresos_ant = egresos_ant.aggregate(total=Sum("total"))["total"] or Decimal("0")
    balance_ant = total_ingresos_ant - total_egresos_ant

    def variacion(actual, anterior):
        if anterior == 0:
            return 100 if actual > 0 else 0
        return ((actual - anterior) / anterior) * 100

    variacion_ingresos_mes = variacion(total_ingresos, total_ingresos_ant)
    variacion_egresos_mes = variacion(total_egresos, total_egresos_ant)
    variacion_balance_mes = variacion(balance, balance_ant)

    # Resumen diario
    resumen_diario = []
    balance_acumulado = Decimal("0")

    dias = monthrange(anio, mes)[1]
    for dia in range(1, dias + 1):
        fecha = date(anio, mes, dia)

        ingresos_dia = ingresos_qs.filter(fecha=fecha).aggregate(total=Sum("total"))["total"] or Decimal("0")
        egresos_dia = egresos_qs.filter(fecha=fecha).aggregate(total=Sum("total"))["total"] or Decimal("0")

        balance_dia = ingresos_dia - egresos_dia
        balance_acumulado += balance_dia

        resumen_diario.append({
            "dia": dia,
            "ingresos": ingresos_dia,
            "egresos": egresos_dia,
            "balance_acumulado": balance_acumulado,
        })

    top_ingresos = ingresos_qs.order_by("-total")[:5]
    top_egresos = egresos_qs.order_by("-total")[:5]

    fecha_inicio_reporte = primer_dia.strftime("%Y-%m-%d")
    fecha_fin_reporte = ultimo_dia.strftime("%Y-%m-%d")

    return render(request, "dashboard/flujo_mensual.html", {
        "anio": anio,
        "mes": mes,
        "hoy": hoy,
        "primer_dia": primer_dia,
        "ultimo_dia": ultimo_dia,

        "ingresos_qs": ingresos_qs,
        "egresos_qs": egresos_qs,
        "ingresos_hoy_qs": ingresos_hoy_qs[:8],
        "egresos_hoy_qs": egresos_hoy_qs[:8],

        "ingresos_manuales_qs": ingresos_manuales_qs,
        "egresos_manuales_qs": egresos_manuales_qs,

        "total_ingresos": total_ingresos,
        "total_egresos": total_egresos,
        "total_ingresos_hoy": total_ingresos_hoy,
        "total_egresos_hoy": total_egresos_hoy,

        "total_ingresos_manuales": total_ingresos_manuales,
        "total_egresos_manuales": total_egresos_manuales,
        "total_ingresos_hoy_manuales": total_ingresos_hoy_manuales,
        "total_egresos_hoy_manuales": total_egresos_hoy_manuales,

        "total_ingresos_automaticos": total_ingresos_automaticos,
        "total_egresos_automaticos": total_egresos_automaticos,
        "total_ingresos_hoy_automaticos": total_ingresos_hoy_automaticos,
        "total_egresos_hoy_automaticos": total_egresos_hoy_automaticos,

        "balance": balance,
        "balance_hoy": balance_hoy,

        "facturas_mes": facturas_mes,
        "total_facturas_mes": total_facturas_mes,
        "total_facturas_pagadas_mes": total_facturas_pagadas_mes,
        "total_facturas_pendientes_mes": total_facturas_pendientes_mes,
        "total_facturas_vencidas": total_facturas_vencidas,
        "total_facturas_pagadas_count": total_facturas_pagadas_count,
        "total_facturas_pendientes_count": total_facturas_pendientes_count,

        "recurrentes_proximos_3_dias": recurrentes_proximos_3_dias,

        "resumen_mes_actual": {
            "ingresos": total_ingresos,
            "egresos": total_egresos,
            "balance": balance,
        },
        "resumen_mes_anterior": {
            "ingresos": total_ingresos_ant,
            "egresos": total_egresos_ant,
            "balance": balance_ant,
        },

        "variacion_ingresos_mes": variacion_ingresos_mes,
        "variacion_egresos_mes": variacion_egresos_mes,
        "variacion_balance_mes": variacion_balance_mes,

        "resumen_diario": resumen_diario,

        "top_ingresos": top_ingresos,
        "top_egresos": top_egresos,

        "fecha_inicio_reporte": fecha_inicio_reporte,
        "fecha_fin_reporte": fecha_fin_reporte,
        "es_admin": True,
    })


@login_required
def egreso_manual_crear_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    if request.method != "POST":
        return redirect("/dashboard/finanzas/flujo/")

    concepto = (request.POST.get("concepto", "") or "").strip()
    categoria = (request.POST.get("categoria", "") or "").strip()
    total_str = (request.POST.get("total", "") or "").strip()
    fecha_str = (request.POST.get("fecha", "") or "").strip()
    next_url = (request.POST.get("next", "") or "").strip()

    if not concepto:
        messages.error(request, "Debes escribir un concepto para el egreso.")
        return redirect(next_url or "/dashboard/finanzas/flujo/")

    try:
        total = float(total_str)
        if total <= 0:
            raise ValueError
    except Exception:
        messages.error(request, "Total inválido para el egreso.")
        return redirect(next_url or "/dashboard/finanzas/flujo/")

    fecha = parse_date(fecha_str)
    if not fecha:
        messages.error(request, "Fecha inválida para el egreso.")
        return redirect(next_url or "/dashboard/finanzas/flujo/")

    _crear_egreso_manual(
        concepto=concepto,
        categoria=categoria,
        total=total,
        fecha=fecha,
    )

    _registrar_actividad(
        user=request.user,
        titulo="Egreso manual creado",
        descripcion=f"{request.user.username} registró el egreso manual '{concepto}' por ${total}.",
        url=f"/dashboard/finanzas/flujo/?anio={fecha.year}&mes={fecha.month}",
    )

    messages.success(request, "Egreso manual registrado correctamente.")
    return redirect(next_url or f"/dashboard/finanzas/flujo/?anio={fecha.year}&mes={fecha.month}")


@login_required
def egreso_manual_eliminar_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    egreso = get_object_or_404(Egreso, pk=pk)

    if not _egreso_es_manual(egreso):
        messages.error(request, "Solo se pueden eliminar egresos manuales desde esta pantalla.")
        return redirect("/dashboard/finanzas/flujo/")

    if request.method != "POST":
        return redirect(f"/dashboard/finanzas/flujo/?anio={egreso.fecha.year}&mes={egreso.fecha.month}")

    concepto = getattr(egreso, "concepto", "") or "Egreso manual"
    total = egreso.total
    fecha = egreso.fecha

    _registrar_actividad(
        user=request.user,
        titulo="Egreso manual eliminado",
        descripcion=f"{request.user.username} eliminó el egreso manual '{concepto}' por ${total}.",
        url=f"/dashboard/finanzas/flujo/?anio={fecha.year}&mes={fecha.month}",
    )

    egreso.delete()
    messages.success(request, "Egreso manual eliminado.")
    return redirect(f"/dashboard/finanzas/flujo/?anio={fecha.year}&mes={fecha.month}")


@login_required
def ingreso_manual_crear_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    if request.method != "POST":
        return redirect("/dashboard/finanzas/flujo/")

    concepto = (request.POST.get("concepto", "") or "").strip()
    total_str = (request.POST.get("total", "") or "").strip()
    fecha_str = (request.POST.get("fecha", "") or "").strip()
    next_url = (request.POST.get("next", "") or "").strip()

    if not concepto:
        messages.error(request, "Debes escribir un concepto para el ingreso.")
        return redirect(next_url or "/dashboard/finanzas/flujo/")

    try:
        total = float(total_str)
        if total <= 0:
            raise ValueError
    except Exception:
        messages.error(request, "Total inválido para el ingreso.")
        return redirect(next_url or "/dashboard/finanzas/flujo/")

    fecha = parse_date(fecha_str)
    if not fecha:
        messages.error(request, "Fecha inválida para el ingreso.")
        return redirect(next_url or "/dashboard/finanzas/flujo/")

    Ingreso.objects.create(
        concepto=concepto,
        total=total,
        fecha=fecha,
    )

    _registrar_actividad(
        user=request.user,
        titulo="Ingreso manual creado",
        descripcion=f"{request.user.username} registró el ingreso manual '{concepto}' por ${total}.",
        url=f"/dashboard/finanzas/flujo/?anio={fecha.year}&mes={fecha.month}",
    )

    messages.success(request, "Ingreso manual registrado correctamente.")
    return redirect(next_url or f"/dashboard/finanzas/flujo/?anio={fecha.year}&mes={fecha.month}")


@login_required
def ingreso_manual_eliminar_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    ingreso = get_object_or_404(Ingreso, pk=pk)

    if not _ingreso_es_manual(ingreso):
        messages.error(request, "Solo se pueden eliminar ingresos manuales desde esta pantalla.")
        return redirect("/dashboard/finanzas/flujo/")

    if request.method != "POST":
        return redirect(f"/dashboard/finanzas/flujo/?anio={ingreso.fecha.year}&mes={ingreso.fecha.month}")

    concepto = getattr(ingreso, "concepto", "") or "Ingreso manual"
    total = ingreso.total
    fecha = ingreso.fecha

    _registrar_actividad(
        user=request.user,
        titulo="Ingreso manual eliminado",
        descripcion=f"{request.user.username} eliminó el ingreso manual '{concepto}' por ${total}.",
        url=f"/dashboard/finanzas/flujo/?anio={fecha.year}&mes={fecha.month}",
    )

    ingreso.delete()
    messages.success(request, "Ingreso manual eliminado.")
    return redirect(f"/dashboard/finanzas/flujo/?anio={fecha.year}&mes={fecha.month}")


@login_required
def ingreso_list_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    ingresos = Ingreso.objects.all().order_by("-fecha", "-id")[:200]
    total = sum(float(i.total) for i in ingresos) if ingresos else 0

    return render(
        request,
        "dashboard/ingresos_list.html",
        {"ingresos": ingresos, "total": total, "es_admin": True},
    )


@login_required
def ingreso_crear_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    if request.method == "POST":
        concepto = request.POST.get("concepto", "").strip()
        total_str = request.POST.get("total", "").strip()
        fecha_str = request.POST.get("fecha", "").strip()

        if not concepto:
            messages.error(request, "Debes escribir un concepto.")
            return redirect("/dashboard/finanzas/ingresos/nuevo/")

        try:
            total = float(total_str)
            if total <= 0:
                raise ValueError
        except Exception:
            messages.error(request, "Total inválido.")
            return redirect("/dashboard/finanzas/ingresos/nuevo/")

        fecha = parse_date(fecha_str)
        if not fecha:
            messages.error(request, "Fecha inválida.")
            return redirect("/dashboard/finanzas/ingresos/nuevo/")

        ingreso = Ingreso.objects.create(concepto=concepto, total=total, fecha=fecha)
        _registrar_actividad(
            user=request.user,
            titulo="Ingreso creado",
            descripcion=f"{request.user.username} creó el ingreso '{concepto}' por ${total}.",
            url=f"/dashboard/finanzas/ingresos/{ingreso.pk}/editar/",
        )

        messages.success(request, "Ingreso creado correctamente.")
        return redirect("/dashboard/finanzas/ingresos/")

    return render(request, "dashboard/ingreso_form.html", {"modo": "crear", "es_admin": True})


@login_required
def ingreso_editar_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    ingreso = get_object_or_404(Ingreso, pk=pk)

    if request.method == "POST":
        concepto = request.POST.get("concepto", "").strip()
        total_str = request.POST.get("total", "").strip()
        fecha_str = request.POST.get("fecha", "").strip()

        if not concepto:
            messages.error(request, "Debes escribir un concepto.")
            return redirect(f"/dashboard/finanzas/ingresos/{pk}/editar/")

        try:
            total = float(total_str)
            if total <= 0:
                raise ValueError
        except Exception:
            messages.error(request, "Total inválido.")
            return redirect(f"/dashboard/finanzas/ingresos/{pk}/editar/")

        fecha = parse_date(fecha_str)
        if not fecha:
            messages.error(request, "Fecha inválido.")
            return redirect(f"/dashboard/finanzas/ingresos/{pk}/editar/")

        ingreso.concepto = concepto
        ingreso.total = total
        ingreso.fecha = fecha
        ingreso.save()

        _registrar_actividad(
            user=request.user,
            titulo="Ingreso actualizado",
            descripcion=f"{request.user.username} actualizó el ingreso '{concepto}' a ${total}.",
            url=f"/dashboard/finanzas/ingresos/{pk}/editar/",
        )

        messages.success(request, "Ingreso actualizado.")
        return redirect("/dashboard/finanzas/ingresos/")

    return render(
        request,
        "dashboard/ingreso_form.html",
        {"modo": "editar", "ingreso": ingreso, "es_admin": True},
    )


@login_required
def ingreso_eliminar_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    ingreso = get_object_or_404(Ingreso, pk=pk)

    if request.method == "POST":
        concepto = ingreso.concepto
        total = ingreso.total
        _registrar_actividad(
            user=request.user,
            titulo="Ingreso eliminado",
            descripcion=f"{request.user.username} eliminó el ingreso '{concepto}' por ${total}.",
            url="/dashboard/finanzas/ingresos/",
        )

        ingreso.delete()
        messages.success(request, "Ingreso eliminado.")
        return redirect("/dashboard/finanzas/ingresos/")

    return render(
        request,
        "dashboard/ingreso_eliminar.html",
        {"ingreso": ingreso, "es_admin": True},
    )


# -------------------
# Finanzas - Movimientos recurrentes
# -------------------
@login_required
def movimientos_recurrentes_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    notificar_movimientos_recurrentes_proximos()

    if request.method == "POST":
        tipo = (request.POST.get("tipo", "") or "").strip()
        concepto = (request.POST.get("concepto", "") or "").strip()
        monto_str = (request.POST.get("monto", "") or "").strip()
        frecuencia = (request.POST.get("frecuencia", "") or "").strip()
        proxima_fecha_str = (request.POST.get("proxima_fecha", "") or "").strip()
        activo = request.POST.get("activo") == "on"

        if tipo not in ["ingreso", "egreso"]:
            messages.error(request, "Tipo de movimiento inválido.")
            return redirect("/dashboard/finanzas/recurrentes/")

        if frecuencia not in ["mensual", "semanal"]:
            messages.error(request, "Frecuencia inválida.")
            return redirect("/dashboard/finanzas/recurrentes/")

        if not concepto:
            messages.error(request, "Debes escribir un concepto.")
            return redirect("/dashboard/finanzas/recurrentes/")

        try:
            monto = float(monto_str)
            if monto <= 0:
                raise ValueError
        except Exception:
            messages.error(request, "Monto inválido.")
            return redirect("/dashboard/finanzas/recurrentes/")

        proxima_fecha = parse_date(proxima_fecha_str)
        if not proxima_fecha:
            messages.error(request, "Fecha inválida.")
            return redirect("/dashboard/finanzas/recurrentes/")

        MovimientoRecurrente.objects.create(
            tipo=tipo,
            concepto=concepto,
            monto=monto,
            frecuencia=frecuencia,
            proxima_fecha=proxima_fecha,
            activo=activo,
        )

        _registrar_actividad(
            user=request.user,
            titulo="Movimiento recurrente creado",
            descripcion=f"{request.user.username} creó el movimiento recurrente '{concepto}' por ${monto}.",
            url="/dashboard/finanzas/recurrentes/",
        )

        messages.success(request, "Movimiento recurrente creado correctamente.")
        return redirect("/dashboard/finanzas/recurrentes/")

    movimientos = MovimientoRecurrente.objects.all().order_by("activo", "proxima_fecha", "-id")
    total_activos = movimientos.filter(activo=True).count()
    total_inactivos = movimientos.filter(activo=False).count()
    total_ingresos = movimientos.filter(tipo="ingreso", activo=True).count()
    total_egresos = movimientos.filter(tipo="egreso", activo=True).count()
    pendientes = movimientos.filter(activo=True, proxima_fecha__lte=date.today()).count()

    return render(
        request,
        "dashboard/movimientos_recurrentes.html",
        {
            "movimientos": movimientos,
            "total_activos": total_activos,
            "total_inactivos": total_inactivos,
            "total_ingresos": total_ingresos,
            "total_egresos": total_egresos,
            "pendientes": pendientes,
            "hoy": date.today(),
            "es_admin": True,
        },
    )


@login_required
@require_http_methods(["POST"])
def movimientos_recurrentes_procesar_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    resultado = procesar_movimientos_recurrentes()

    _registrar_actividad(
        user=request.user,
        titulo="Recurrentes procesados",
        descripcion=(
            f"{request.user.username} ejecutó los movimientos recurrentes: "
            f"{resultado['total_generados']} generados "
            f"({resultado['ingresos_generados']} ingresos, {resultado['egresos_generados']} egresos)."
        ),
        url="/dashboard/finanzas/recurrentes/",
    )

    if resultado["total_generados"] > 0:
        messages.success(
            request,
            f"Proceso completado: {resultado['total_generados']} movimientos generados "
            f"({resultado['ingresos_generados']} ingresos y {resultado['egresos_generados']} egresos)."
        )
    else:
        messages.info(request, "No había movimientos recurrentes pendientes por procesar.")

    return redirect("/dashboard/finanzas/recurrentes/")


@login_required
def movimiento_recurrente_editar_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    movimiento = get_object_or_404(MovimientoRecurrente, pk=pk)

    if request.method == "POST":
        tipo = (request.POST.get("tipo", "") or "").strip()
        concepto = (request.POST.get("concepto", "") or "").strip()
        monto_str = (request.POST.get("monto", "") or "").strip()
        frecuencia = (request.POST.get("frecuencia", "") or "").strip()
        proxima_fecha_str = (request.POST.get("proxima_fecha", "") or "").strip()
        activo = request.POST.get("activo") == "on"

        if tipo not in ["ingreso", "egreso"]:
            messages.error(request, "Tipo inválido.")
            return redirect(f"/dashboard/finanzas/recurrentes/{pk}/editar/")

        if frecuencia not in ["mensual", "semanal"]:
            messages.error(request, "Frecuencia inválida.")
            return redirect(f"/dashboard/finanzas/recurrentes/{pk}/editar/")

        if not concepto:
            messages.error(request, "Debes escribir un concepto.")
            return redirect(f"/dashboard/finanzas/recurrentes/{pk}/editar/")

        try:
            monto = float(monto_str)
            if monto <= 0:
                raise ValueError
        except Exception:
            messages.error(request, "Monto inválido.")
            return redirect(f"/dashboard/finanzas/recurrentes/{pk}/editar/")

        proxima_fecha = parse_date(proxima_fecha_str)
        if not proxima_fecha:
            messages.error(request, "Fecha inválida.")
            return redirect(f"/dashboard/finanzas/recurrentes/{pk}/editar/")

        movimiento.tipo = tipo
        movimiento.concepto = concepto
        movimiento.monto = monto
        movimiento.frecuencia = frecuencia
        movimiento.proxima_fecha = proxima_fecha
        movimiento.activo = activo
        movimiento.save()

        _registrar_actividad(
            user=request.user,
            titulo="Movimiento recurrente actualizado",
            descripcion=f"{request.user.username} actualizó el movimiento recurrente '{concepto}'.",
            url="/dashboard/finanzas/recurrentes/",
        )

        messages.success(request, "Movimiento recurrente actualizado correctamente.")
        return redirect("/dashboard/finanzas/recurrentes/")

    return render(
        request,
        "dashboard/movimiento_recurrente_form.html",
        {
            "movimiento": movimiento,
            "modo": "editar",
            "es_admin": True,
        },
    )


@login_required
@require_http_methods(["POST"])
def movimiento_recurrente_toggle_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    movimiento = get_object_or_404(MovimientoRecurrente, pk=pk)

    movimiento.activo = not movimiento.activo
    movimiento.save(update_fields=["activo"])

    estado = "activado" if movimiento.activo else "desactivado"

    _registrar_actividad(
        user=request.user,
        titulo="Movimiento recurrente actualizado",
        descripcion=f"{request.user.username} {estado} el movimiento recurrente '{movimiento.concepto}'.",
        url="/dashboard/finanzas/recurrentes/",
    )

    messages.success(request, f"Movimiento recurrente {estado} correctamente.")
    return redirect("/dashboard/finanzas/recurrentes/")


@login_required
def movimiento_recurrente_eliminar_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    movimiento = get_object_or_404(MovimientoRecurrente, pk=pk)

    if request.method == "POST":
        concepto = movimiento.concepto

        _registrar_actividad(
            user=request.user,
            titulo="Movimiento recurrente eliminado",
            descripcion=f"{request.user.username} eliminó el movimiento recurrente '{concepto}'.",
            url="/dashboard/finanzas/recurrentes/",
        )

        movimiento.delete()
        messages.success(request, "Movimiento recurrente eliminado correctamente.")
        return redirect("/dashboard/finanzas/recurrentes/")

    return render(
        request,
        "dashboard/movimiento_recurrente_form.html",
        {
            "movimiento": movimiento,
            "modo": "eliminar",
            "es_admin": True,
        },
    )


def offline_view(request):
    return render(request, "dashboard/offline.html")


@require_GET
@login_required
def unread_count_view(request):
    if es_admin(request.user):
        try:
            from finanzas.alertas_financieras import generar_alertas_financieras
            generar_alertas_financieras(enviar_push=True)
        except Exception:
            logger.exception("No se pudieron actualizar las alertas financieras.")

    if Notificacion is None:
        return JsonResponse({"count": 0})

    count = Notificacion.objects.filter(
        user_id=request.user.id,
        leida=False
    ).count()

    return JsonResponse({"count": count})

#======================
# INVENTARIO INTELIGENTE
#======================

from inventario.models import (
    Insumo, VentaInsumo, EntradaStock, MovimientoInventario,
    InventarioTrabajador, CompraInsumo, PresentacionInsumo, SolicitudReposicion,
)
from inventario.services import (
    convertir_a_base, entregar_a_trabajador, devolver_de_trabajador,
    consumir_trabajador, revertir_consumo, decimal_positivo,
)


def _inventario_valorizado():
    total = Decimal("0.00")
    for insumo in Insumo.objects.filter(activo=True):
        total += Decimal(insumo.stock or 0) * Decimal(insumo.costo or 0)
    return total


def _cantidad_texto(insumo, cantidad):
    cantidad = Decimal(cantidad or 0)
    if insumo.unidad_base == "kg":
        return f"{cantidad:.3f} kg"
    return f"{cantidad:.0f} unid."


@login_required
def inventario_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    insumos = list(Insumo.objects.filter(activo=True).prefetch_related("presentaciones").order_by("nombre"))
    trabajadores = list(Trabajador.objects.filter(activo=True).select_related("user").order_by("user__username"))
    inventarios_trabajadores = list(
        InventarioTrabajador.objects.filter(stock__gt=0)
        .select_related("trabajador__user", "insumo")
        .order_by("trabajador__user__username", "insumo__nombre")
    )

    total_insumos = len(insumos)
    bajo_stock = sum(1 for i in insumos if i.bajo_stock)
    sin_stock = sum(1 for i in insumos if i.sin_stock)
    stock_total_kg = sum((Decimal(i.stock) for i in insumos if i.unidad_base == "kg"), Decimal("0.000"))

    hoy = timezone.localdate()
    primer_dia_mes = hoy.replace(day=1)
    ventas_mes = VentaInsumo.objects.filter(fecha__gte=primer_dia_mes)
    total_ventas_mes = ventas_mes.aggregate(total=Sum("total")).get("total") or Decimal("0")
    ganancia_mes = ventas_mes.aggregate(total=Sum("ganancia")).get("total") or Decimal("0")
    total_compras_mes = CompraInsumo.objects.filter(fecha__gte=primer_dia_mes).aggregate(total=Sum("total")).get("total") or Decimal("0")
    consumo_mes = MovimientoInventario.objects.filter(tipo="mantenimiento", fecha__gte=primer_dia_mes)
    costo_consumo_mes = consumo_mes.aggregate(total=Sum("total_costo")).get("total") or Decimal("0")
    cantidad_consumo_mes = consumo_mes.aggregate(total=Sum("cantidad")).get("total") or Decimal("0")
    solicitudes_pendientes = SolicitudReposicion.objects.filter(estado="pendiente").select_related("trabajador__user", "insumo").order_by("-creada_en")

    movimientos_recientes = list(
        MovimientoInventario.objects.select_related("insumo", "trabajador__user", "mantenimiento__cliente")
        .all().order_by("-creado_en", "-id")[:12]
    )

    resumen_trabajadores = []
    por_trabajador = defaultdict(list)
    for inv in inventarios_trabajadores:
        por_trabajador[inv.trabajador].append(inv)
    for trabajador, stocks in por_trabajador.items():
        valorizado = sum((Decimal(x.stock) * Decimal(x.insumo.costo or 0) for x in stocks), Decimal("0.00"))
        resumen_trabajadores.append({"trabajador": trabajador, "stocks": stocks, "valor": valorizado})

    return render(request, "dashboard/inventario.html", {
        "insumos": insumos,
        "trabajadores": trabajadores,
        "inventarios_trabajadores": inventarios_trabajadores,
        "resumen_trabajadores": resumen_trabajadores,
        "total_insumos": total_insumos,
        "bajo_stock": bajo_stock,
        "sin_stock": sin_stock,
        "stock_total_kg": stock_total_kg,
        "valor_inventario": _inventario_valorizado(),
        "total_ventas_mes": total_ventas_mes,
        "ganancia_mes": ganancia_mes,
        "total_compras_mes": total_compras_mes,
        "costo_consumo_mes": costo_consumo_mes,
        "cantidad_consumo_mes": cantidad_consumo_mes,
        "solicitudes_pendientes": solicitudes_pendientes[:8],
        "solicitudes_pendientes_count": solicitudes_pendientes.count(),
        "movimientos_recientes": movimientos_recientes,
        "es_admin": True,
    })


@login_required
def compra_inventario_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    if request.method != "POST":
        return redirect("inventario")

    insumo = get_object_or_404(Insumo, pk=request.POST.get("insumo_id"))
    try:
        cantidad_base = convertir_a_base(insumo, request.POST.get("cantidad"), request.POST.get("unidad", "base"), request.POST.get("presentacion_id") or None)
        costo_unitario = decimal_positivo(request.POST.get("costo_unitario"), "Costo unitario")
    except (ValueError, PresentacionInsumo.DoesNotExist) as exc:
        messages.error(request, str(exc))
        return redirect("inventario")

    proveedor = (request.POST.get("proveedor") or "").strip()
    lote = (request.POST.get("lote") or "").strip()
    fecha_fabricacion = parse_date((request.POST.get("fecha_fabricacion") or "").strip()) or None
    fecha_vencimiento = parse_date((request.POST.get("fecha_vencimiento") or "").strip()) or None
    observacion = (request.POST.get("observacion") or "").strip()
    total = (cantidad_base * costo_unitario).quantize(Decimal("0.01"))

    with transaction.atomic():
        insumo = Insumo.objects.select_for_update().get(pk=insumo.pk)
        stock_anterior = Decimal(insumo.stock)
        stock_nuevo = stock_anterior + cantidad_base
        valor_anterior = stock_anterior * Decimal(insumo.costo or 0)
        valor_compra = cantidad_base * costo_unitario
        nuevo_costo = ((valor_anterior + valor_compra) / stock_nuevo) if stock_nuevo > 0 else costo_unitario
        insumo.stock = stock_nuevo
        insumo.costo = nuevo_costo.quantize(Decimal("0.0001"))
        insumo.save(update_fields=["stock", "costo"])

        egreso = Egreso.objects.create(
            concepto=f"Compra inventario: {insumo.nombre}", categoria="quimicos" if insumo.categoria == "quimicos" else "materiales",
            cantidad=1, costo_unitario=total, total=total, fecha=timezone.localdate(), proveedor=proveedor,
        )
        compra = CompraInsumo.objects.create(
            insumo=insumo, cantidad=cantidad_base, costo_unitario=costo_unitario,
            total=total, proveedor=proveedor, lote=lote, fecha_fabricacion=fecha_fabricacion,
            fecha_vencimiento=fecha_vencimiento, observacion=observacion, egreso=egreso,
        )
        MovimientoInventario.objects.create(
            insumo=insumo, tipo="compra", cantidad=cantidad_base,
            stock_anterior=stock_anterior, stock_resultante=insumo.stock,
            costo_unitario=costo_unitario, total_costo=total, usuario=request.user,
            observacion=observacion or f"Compra #{compra.pk} · {proveedor or 'sin proveedor'}",
        )

    _registrar_actividad(request.user, "Compra de inventario", f"Se ingresaron {_cantidad_texto(insumo, cantidad_base)} de {insumo.nombre} por ${total}.", "/dashboard/inventario/")
    messages.success(request, f"Compra registrada. Stock actual: {_cantidad_texto(insumo, insumo.stock)}")
    return redirect("inventario")


@login_required
def agregar_stock_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    if request.method != "POST":
        return redirect("inventario")
    insumo = get_object_or_404(Insumo, pk=request.POST.get("insumo_id"))
    try:
        cantidad_base = convertir_a_base(insumo, request.POST.get("cantidad"), request.POST.get("unidad", "base"))
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("inventario")
    observacion = (request.POST.get("observacion") or "").strip()
    with transaction.atomic():
        insumo = Insumo.objects.select_for_update().get(pk=insumo.pk)
        anterior = Decimal(insumo.stock)
        insumo.stock = anterior + cantidad_base
        insumo.save(update_fields=["stock"])
        EntradaStock.objects.create(insumo=insumo, cantidad=cantidad_base, observacion=observacion)
        MovimientoInventario.objects.create(
            insumo=insumo, tipo="entrada", cantidad=cantidad_base, stock_anterior=anterior,
            stock_resultante=insumo.stock, costo_unitario=insumo.costo or 0,
            total_costo=(cantidad_base * Decimal(insumo.costo or 0)).quantize(Decimal("0.01")),
            usuario=request.user, observacion=observacion or "Ajuste de entrada manual",
        )
    messages.success(request, "Stock agregado correctamente.")
    return redirect("inventario")


@login_required
def vender_insumo_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    if request.method != "POST":
        return redirect("inventario")

    insumo = get_object_or_404(Insumo, pk=request.POST.get("insumo_id"), puede_venderse=True)
    try:
        cantidad_base = convertir_a_base(insumo, request.POST.get("cantidad"), request.POST.get("unidad", "base"), request.POST.get("presentacion_id") or None)
    except (ValueError, PresentacionInsumo.DoesNotExist) as exc:
        messages.error(request, str(exc))
        return redirect("inventario")

    precio_unitario_raw = request.POST.get("precio_unitario")
    precio_unitario = Decimal(str(precio_unitario_raw).replace(",", ".")) if precio_unitario_raw else Decimal(insumo.precio)

    with transaction.atomic():
        insumo = Insumo.objects.select_for_update().get(pk=insumo.pk)
        if Decimal(insumo.stock) < cantidad_base:
            messages.error(request, f"Stock insuficiente. Disponible: {_cantidad_texto(insumo, insumo.stock)}")
            return redirect("inventario")
        anterior = Decimal(insumo.stock)
        costo_unitario = Decimal(insumo.costo or 0)
        total = (cantidad_base * precio_unitario).quantize(Decimal("0.01"))
        ganancia = (cantidad_base * (precio_unitario - costo_unitario)).quantize(Decimal("0.01"))
        insumo.stock = anterior - cantidad_base
        insumo.save(update_fields=["stock"])
        venta = VentaInsumo.objects.create(
            insumo=insumo, cantidad=cantidad_base, unidad_registro=request.POST.get("unidad", "base"),
            precio_unitario=precio_unitario, costo_unitario=costo_unitario, total=total, ganancia=ganancia,
        )
        MovimientoInventario.objects.create(
            insumo=insumo, tipo="venta", cantidad=cantidad_base,
            stock_anterior=anterior, stock_resultante=insumo.stock,
            costo_unitario=costo_unitario, total_costo=(cantidad_base * costo_unitario).quantize(Decimal("0.01")),
            usuario=request.user, observacion=f"Venta #{venta.pk} · ingreso ${total}",
        )
        Ingreso.objects.create(concepto=f"Venta de insumo: {insumo.nombre}", total=total, fecha=timezone.localdate())

    _registrar_actividad(request.user, "Venta de insumo registrada", f"Venta de {insumo.nombre} por ${total}. Ganancia estimada ${ganancia}.", "/dashboard/inventario/")
    messages.success(request, f"Venta registrada correctamente. Ganancia estimada: ${ganancia}")
    return redirect("inventario")


@login_required
def inventario_entrega_trabajador_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    if request.method != "POST":
        return redirect("inventario")
    insumo = get_object_or_404(Insumo, pk=request.POST.get("insumo_id"), puede_asignarse_trabajador=True)
    trabajador = get_object_or_404(Trabajador, pk=request.POST.get("trabajador_id"), activo=True)
    try:
        cantidad = convertir_a_base(insumo, request.POST.get("cantidad"), request.POST.get("unidad", "base"))
        entregar_a_trabajador(insumo=insumo, trabajador=trabajador, cantidad_base=cantidad, usuario=request.user, observacion=request.POST.get("observacion", ""))
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("inventario")
    messages.success(request, f"Entrega registrada a {trabajador}: {_cantidad_texto(insumo, cantidad)} de {insumo.nombre}.")
    return redirect("inventario")


@login_required
def inventario_devolucion_trabajador_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    if request.method != "POST":
        return redirect("inventario")
    insumo = get_object_or_404(Insumo, pk=request.POST.get("insumo_id"))
    trabajador = get_object_or_404(Trabajador, pk=request.POST.get("trabajador_id"))
    try:
        cantidad = convertir_a_base(insumo, request.POST.get("cantidad"), request.POST.get("unidad", "base"))
        devolver_de_trabajador(insumo=insumo, trabajador=trabajador, cantidad_base=cantidad, usuario=request.user, observacion=request.POST.get("observacion", ""))
    except (ValueError, InventarioTrabajador.DoesNotExist) as exc:
        messages.error(request, str(exc) if str(exc) else "El trabajador no tiene stock de ese producto.")
        return redirect("inventario")
    messages.success(request, "Devolución registrada correctamente.")
    return redirect("inventario")


def _codigo_producto_siguiente(categoria):
    prefijos = {
        "quimicos": "QUI", "repuestos": "REP", "herramientas": "HER",
        "equipos": "EQU", "construccion": "MAT", "otros": "OTR",
    }
    prefijo = prefijos.get(categoria, "PRO")
    existentes = Insumo.objects.filter(codigo__startswith=f"JVQ-{prefijo}-").values_list("codigo", flat=True)
    mayor = 0
    for codigo in existentes:
        try:
            mayor = max(mayor, int(str(codigo).rsplit("-", 1)[-1]))
        except (TypeError, ValueError):
            continue
    return f"JVQ-{prefijo}-{mayor + 1:04d}"


def _guardar_producto_desde_post(insumo, post):
    nombre = (post.get("nombre") or "").strip()
    if not nombre:
        raise ValueError("El nombre del producto es obligatorio.")
    categoria = (post.get("categoria") or "quimicos").strip()
    if categoria not in dict(Insumo.CATEGORIA_CHOICES):
        raise ValueError("Categoría inválida.")
    unidad = (post.get("unidad_base") or "kg").strip()
    if unidad not in dict(Insumo.UNIDAD_CHOICES):
        raise ValueError("Unidad base inválida.")

    def dec(nombre_campo, defecto="0"):
        raw = (post.get(nombre_campo) or defecto).strip().replace(",", ".")
        try:
            return Decimal(raw)
        except Exception:
            raise ValueError(f"Valor inválido en {nombre_campo.replace('_', ' ')}.")

    insumo.nombre = nombre
    insumo.categoria = categoria
    insumo.unidad_base = unidad
    insumo.codigo = (post.get("codigo") or "").strip() or insumo.codigo or _codigo_producto_siguiente(categoria)
    if Insumo.objects.filter(codigo=insumo.codigo).exclude(pk=insumo.pk).exists():
        raise ValueError("Ya existe otro producto con ese código.")
    insumo.marca = (post.get("marca") or "").strip()
    insumo.modelo = (post.get("modelo") or "").strip()
    insumo.descripcion = (post.get("descripcion") or "").strip()
    insumo.stock_minimo = dec("stock_minimo", "0")
    if insumo.stock_minimo < 0:
        raise ValueError("El stock mínimo no puede ser negativo.")
    maximo_raw = (post.get("stock_maximo") or "").strip().replace(",", ".")
    insumo.stock_maximo = Decimal(maximo_raw) if maximo_raw else None
    insumo.precio = dec("precio", "0")
    # El costo promedio no se altera al editar salvo que sea un producto nuevo sin movimientos.
    if not insumo.pk:
        insumo.costo = dec("costo", "0")
    insumo.activo = post.get("activo") == "on"
    insumo.puede_venderse = post.get("puede_venderse") == "on"
    insumo.puede_mantenimiento = post.get("puede_mantenimiento") == "on"
    insumo.puede_asignarse_trabajador = post.get("puede_asignarse_trabajador") == "on"
    insumo.puede_construccion = post.get("puede_construccion") == "on"
    insumo.controla_inventario = post.get("controla_inventario") == "on"
    insumo.save()
    return insumo


@login_required
def inventario_productos_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    q = (request.GET.get("q") or "").strip()
    categoria = (request.GET.get("categoria") or "").strip()
    estado = (request.GET.get("estado") or "activos").strip()
    qs = Insumo.objects.prefetch_related("presentaciones").order_by("nombre")
    if q:
        qs = qs.filter(models.Q(nombre__icontains=q) | models.Q(codigo__icontains=q) | models.Q(marca__icontains=q) | models.Q(modelo__icontains=q))
    if categoria in dict(Insumo.CATEGORIA_CHOICES):
        qs = qs.filter(categoria=categoria)
    if estado == "activos": qs = qs.filter(activo=True)
    elif estado == "inactivos": qs = qs.filter(activo=False)
    return render(request, "dashboard/inventario_productos.html", {
        "productos": qs, "q": q, "categoria": categoria, "estado": estado,
        "categorias": Insumo.CATEGORIA_CHOICES, "es_admin": True,
    })


@login_required
def inventario_producto_crear_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    insumo = Insumo(activo=True, puede_venderse=True, puede_mantenimiento=True, puede_asignarse_trabajador=True, controla_inventario=True)
    if request.method == "POST":
        try:
            _guardar_producto_desde_post(insumo, request.POST)
            _registrar_actividad(request.user, "Producto creado", f"Se creó {insumo.nombre} ({insumo.codigo}).", f"/dashboard/inventario/productos/{insumo.pk}/")
            messages.success(request, "Producto creado correctamente.")
            return redirect("inventario_producto_detalle", pk=insumo.pk)
        except (ValueError, Exception) as exc:
            # Los errores de integridad/campo se muestran sin perder el formulario.
            messages.error(request, str(exc))
    return render(request, "dashboard/inventario_producto_form.html", {
        "producto": insumo, "modo": "crear", "categorias": Insumo.CATEGORIA_CHOICES,
        "unidades": Insumo.UNIDAD_CHOICES, "es_admin": True,
    })


@login_required
def inventario_producto_editar_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    insumo = get_object_or_404(Insumo, pk=pk)
    if request.method == "POST":
        try:
            _guardar_producto_desde_post(insumo, request.POST)
            _registrar_actividad(request.user, "Producto actualizado", f"Se actualizó {insumo.nombre} ({insumo.codigo}).", f"/dashboard/inventario/productos/{insumo.pk}/")
            messages.success(request, "Producto actualizado correctamente.")
            return redirect("inventario_producto_detalle", pk=insumo.pk)
        except Exception as exc:
            messages.error(request, str(exc))
    return render(request, "dashboard/inventario_producto_form.html", {
        "producto": insumo, "modo": "editar", "categorias": Insumo.CATEGORIA_CHOICES,
        "unidades": Insumo.UNIDAD_CHOICES, "es_admin": True,
    })


@login_required
def inventario_producto_detalle_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    insumo = get_object_or_404(Insumo.objects.prefetch_related("presentaciones"), pk=pk)
    movimientos = list(MovimientoInventario.objects.filter(insumo=insumo).select_related("trabajador__user", "mantenimiento__cliente", "usuario").order_by("-creado_en")[:100])
    stocks = list(InventarioTrabajador.objects.filter(insumo=insumo, stock__gt=0).select_related("trabajador__user").order_by("trabajador__user__username"))
    hoy = timezone.localdate(); inicio = hoy.replace(day=1)
    consumo = MovimientoInventario.objects.filter(insumo=insumo, tipo="mantenimiento", fecha__gte=inicio).aggregate(c=Sum("cantidad"), costo=Sum("total_costo"))
    ventas = VentaInsumo.objects.filter(insumo=insumo, fecha__gte=inicio).aggregate(c=Sum("cantidad"), total=Sum("total"))
    ultima_compra = CompraInsumo.objects.filter(insumo=insumo).order_by("-creado_en").first()
    return render(request, "dashboard/inventario_producto_detalle.html", {
        "producto": insumo, "movimientos": movimientos, "stocks_trabajadores": stocks,
        "consumo_mes": consumo.get("c") or Decimal("0"), "costo_consumo_mes": consumo.get("costo") or Decimal("0"),
        "ventas_mes": ventas.get("c") or Decimal("0"), "ventas_total_mes": ventas.get("total") or Decimal("0"),
        "ultima_compra": ultima_compra, "valor_stock": Decimal(insumo.stock or 0) * Decimal(insumo.costo or 0), "es_admin": True,
    })


@login_required
@require_http_methods(["POST"])
def inventario_producto_toggle_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    insumo = get_object_or_404(Insumo, pk=pk)
    insumo.activo = not insumo.activo
    insumo.save(update_fields=["activo"])
    messages.success(request, f"Producto {'activado' if insumo.activo else 'desactivado'} correctamente.")
    return redirect("inventario_producto_detalle", pk=pk)


@login_required
@require_http_methods(["POST"])
def inventario_producto_eliminar_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    insumo = get_object_or_404(Insumo, pk=pk)
    tiene_historial = (
        MovimientoInventario.objects.filter(insumo=insumo).exists()
        or VentaInsumo.objects.filter(insumo=insumo).exists()
        or CompraInsumo.objects.filter(insumo=insumo).exists()
        or InventarioTrabajador.objects.filter(insumo=insumo, stock__gt=0).exists()
        or Decimal(insumo.stock or 0) != 0
    )
    if tiene_historial:
        insumo.activo = False
        insumo.save(update_fields=["activo"])
        messages.warning(request, "El producto tiene historial o existencias; se desactivó para conservar la trazabilidad.")
        return redirect("inventario_producto_detalle", pk=pk)
    nombre = insumo.nombre
    insumo.delete()
    messages.success(request, f"Producto {nombre} eliminado definitivamente.")
    return redirect("inventario_productos")


@login_required
@require_http_methods(["POST"])
def inventario_presentacion_agregar_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    insumo = get_object_or_404(Insumo, pk=pk)
    nombre = (request.POST.get("nombre") or "").strip()
    try:
        cantidad = decimal_positivo(request.POST.get("cantidad_base"), "Cantidad de la presentación")
        precio_raw = (request.POST.get("precio_venta") or "").strip().replace(",", ".")
        precio = Decimal(precio_raw) if precio_raw else None
        if not nombre: raise ValueError("Escribe el nombre de la presentación.")
        PresentacionInsumo.objects.create(insumo=insumo, nombre=nombre, cantidad_base=cantidad, precio_venta=precio, activa=True)
        messages.success(request, "Presentación agregada.")
    except Exception as exc:
        messages.error(request, str(exc))
    return redirect("inventario_producto_detalle", pk=pk)


@login_required
@require_http_methods(["POST"])
def inventario_presentacion_eliminar_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    presentacion = get_object_or_404(PresentacionInsumo, pk=pk)
    insumo_id = presentacion.insumo_id
    if presentacion.ventas.exists():
        presentacion.activa = False; presentacion.save(update_fields=["activa"])
        messages.warning(request, "La presentación tiene ventas históricas y fue desactivada.")
    else:
        presentacion.delete(); messages.success(request, "Presentación eliminada.")
    return redirect("inventario_producto_detalle", pk=insumo_id)


@login_required
def mi_inventario_trabajador_view(request):
    if not es_trabajador(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    trabajador = get_object_or_404(Trabajador.objects.select_related("user"), user=request.user)
    stocks = list(InventarioTrabajador.objects.filter(trabajador=trabajador, stock__gt=0).select_related("insumo").order_by("insumo__nombre"))
    movimientos = list(MovimientoInventario.objects.filter(trabajador=trabajador).select_related("insumo", "mantenimiento__cliente").order_by("-creado_en")[:100])
    hoy = timezone.localdate(); inicio = hoy.replace(day=1)
    consumos = {x["insumo_id"]: x["total"] or Decimal("0") for x in MovimientoInventario.objects.filter(trabajador=trabajador, tipo="mantenimiento", fecha__gte=inicio).values("insumo_id").annotate(total=Sum("cantidad"))}
    ultimas_entregas = {}
    for mov in MovimientoInventario.objects.filter(trabajador=trabajador, tipo="entrega").select_related("insumo").order_by("-creado_en"):
        ultimas_entregas.setdefault(mov.insumo_id, mov)
    items = []
    for stock in stocks:
        items.append({"stock": stock, "consumo_mes": consumos.get(stock.insumo_id, Decimal("0")), "ultima_entrega": ultimas_entregas.get(stock.insumo_id)})
    return render(request, "dashboard/mi_inventario_trabajador.html", {
        "trabajador": trabajador, "items": items, "movimientos": movimientos,
        "valor_estimado": sum((Decimal(x.stock) * Decimal(x.insumo.costo or 0) for x in stocks), Decimal("0")),
        "es_admin": False,
    })


@login_required
@require_http_methods(["POST"])
def solicitud_reposicion_crear_view(request, insumo_id):
    if not es_trabajador(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    trabajador = get_object_or_404(Trabajador, user=request.user)
    insumo = get_object_or_404(Insumo, pk=insumo_id, activo=True)
    stock = InventarioTrabajador.objects.filter(trabajador=trabajador, insumo=insumo).first()
    if SolicitudReposicion.objects.filter(trabajador=trabajador, insumo=insumo, estado="pendiente").exists():
        messages.info(request, "Ya existe una solicitud pendiente para este producto.")
        return redirect("mi_inventario_trabajador")
    cantidad = None
    raw = (request.POST.get("cantidad_sugerida") or "").strip()
    if raw:
        try: cantidad = convertir_a_base(insumo, raw, request.POST.get("unidad", "base"))
        except ValueError as exc:
            messages.error(request, str(exc)); return redirect("mi_inventario_trabajador")
    solicitud = SolicitudReposicion.objects.create(
        trabajador=trabajador, insumo=insumo, stock_al_solicitar=Decimal(stock.stock if stock else 0),
        cantidad_sugerida=cantidad, observacion=(request.POST.get("observacion") or "").strip(),
    )
    _notificar_admins("📦 Solicitud de reposición", f"{trabajador} solicitó reposición de {insumo.nombre}. Stock actual: {_cantidad_texto(insumo, solicitud.stock_al_solicitar)}.", "/dashboard/inventario/", enviar_push=True)
    messages.success(request, "Solicitud enviada a administración.")
    return redirect("mi_inventario_trabajador")


@login_required
@require_http_methods(["POST"])
def solicitud_reposicion_atender_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    solicitud = get_object_or_404(SolicitudReposicion, pk=pk)
    solicitud.estado = "atendida"
    solicitud.atendida_en = timezone.now()
    solicitud.atendida_por = request.user
    solicitud.save(update_fields=["estado", "atendida_en", "atendida_por"])
    messages.success(request, "Solicitud marcada como atendida.")
    return redirect("inventario")


@login_required
def inventario_productos_criticos_pdf_view(request):
    if not es_admin(request.user): return render(request, "dashboard/no_autorizado.html", status=403)
    filas = [["Producto", "Categoría", "Stock", "Mínimo", "Estado"]]
    for i in Insumo.objects.filter(activo=True).order_by("nombre"):
        if i.stock <= i.stock_minimo:
            filas.append([i.nombre, i.get_categoria_display(), _cantidad_texto(i, i.stock), _cantidad_texto(i, i.stock_minimo), "SIN STOCK" if i.sin_stock else "STOCK BAJO"])
    return _pdf_inventario_response("Productos críticos JVAQUA", filas, "productos_criticos.pdf")


@login_required
def inventario_consumo_trabajadores_pdf_view(request):
    if not es_admin(request.user): return render(request, "dashboard/no_autorizado.html", status=403)
    filas = [["Trabajador", "Producto", "Consumo", "Costo"]]
    qs = MovimientoInventario.objects.filter(tipo="mantenimiento").values("trabajador__user__first_name", "trabajador__user__last_name", "trabajador__user__username", "insumo__nombre", "insumo__unidad_base").annotate(cantidad=Sum("cantidad"), costo=Sum("total_costo")).order_by("trabajador__user__username", "insumo__nombre")
    for x in qs:
        nombre = (f"{x['trabajador__user__first_name']} {x['trabajador__user__last_name']}").strip() or x["trabajador__user__username"] or "—"
        unidad = "kg" if x["insumo__unidad_base"] == "kg" else "unid."
        filas.append([nombre, x["insumo__nombre"], f"{Decimal(x['cantidad'] or 0):.3f} {unidad}", f"${Decimal(x['costo'] or 0):,.2f}"])
    return _pdf_inventario_response("Consumo por trabajador JVAQUA", filas, "consumo_por_trabajador.pdf")


@login_required
def inventario_consumo_contratos_pdf_view(request):
    if not es_admin(request.user): return render(request, "dashboard/no_autorizado.html", status=403)
    filas = [["Cliente / contrato", "Producto", "Consumo", "Costo"]]
    qs = MovimientoInventario.objects.filter(tipo="mantenimiento", mantenimiento__isnull=False).values("mantenimiento__cliente__nombre", "insumo__nombre", "insumo__unidad_base").annotate(cantidad=Sum("cantidad"), costo=Sum("total_costo")).order_by("mantenimiento__cliente__nombre", "insumo__nombre")
    for x in qs:
        unidad = "kg" if x["insumo__unidad_base"] == "kg" else "unid."
        filas.append([x["mantenimiento__cliente__nombre"] or "—", x["insumo__nombre"], f"{Decimal(x['cantidad'] or 0):.3f} {unidad}", f"${Decimal(x['costo'] or 0):,.2f}"])
    return _pdf_inventario_response("Consumo por contrato JVAQUA", filas, "consumo_por_contrato.pdf")


@login_required
def inventario_historial_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    q = (request.GET.get("q") or "").strip()
    tipo = (request.GET.get("tipo") or "").strip()
    trabajador_id = (request.GET.get("trabajador") or "").strip()
    insumo_id = (request.GET.get("insumo") or "").strip()
    fecha_desde_str = (request.GET.get("fecha_desde") or "").strip()
    fecha_hasta_str = (request.GET.get("fecha_hasta") or "").strip()
    qs = MovimientoInventario.objects.select_related("insumo", "trabajador__user", "mantenimiento__cliente", "usuario").all()
    if tipo in dict(MovimientoInventario.TIPO_CHOICES): qs = qs.filter(tipo=tipo)
    if trabajador_id.isdigit(): qs = qs.filter(trabajador_id=int(trabajador_id))
    if insumo_id.isdigit(): qs = qs.filter(insumo_id=int(insumo_id))
    if fecha_desde_str and parse_date(fecha_desde_str): qs = qs.filter(fecha__gte=parse_date(fecha_desde_str))
    if fecha_hasta_str and parse_date(fecha_hasta_str): qs = qs.filter(fecha__lte=parse_date(fecha_hasta_str))
    if q: qs = qs.filter(models.Q(insumo__nombre__icontains=q) | models.Q(observacion__icontains=q) | models.Q(trabajador__user__username__icontains=q))
    qs = qs.order_by("-creado_en", "-id")
    paginator = Paginator(qs, 40)
    page_obj = paginator.get_page(request.GET.get("page"))
    qp = request.GET.copy(); qp.pop("page", None)
    return render(request, "dashboard/inventario_historial.html", {
        "page_obj": page_obj, "q": q, "tipo": tipo, "trabajador_id": trabajador_id, "insumo_id": insumo_id,
        "fecha_desde": fecha_desde_str, "fecha_hasta": fecha_hasta_str,
        "insumos_filtro": Insumo.objects.filter(activo=True).order_by("nombre"),
        "trabajadores": Trabajador.objects.filter(activo=True).select_related("user").order_by("user__username"),
        "tipos_movimiento": MovimientoInventario.TIPO_CHOICES,
        "querystring": qp.urlencode(), "total_movimientos": qs.count(), "es_admin": True,
    })


@login_required
def inventario_trabajador_detalle_view(request, trabajador_id):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    trabajador = get_object_or_404(Trabajador.objects.select_related("user"), pk=trabajador_id)
    stocks = list(InventarioTrabajador.objects.filter(trabajador=trabajador).select_related("insumo").order_by("insumo__nombre"))
    movimientos = MovimientoInventario.objects.filter(trabajador=trabajador).select_related("insumo", "mantenimiento__cliente").order_by("-creado_en")[:100]
    valor = sum((Decimal(x.stock) * Decimal(x.insumo.costo or 0) for x in stocks), Decimal("0.00"))
    return render(request, "dashboard/inventario_trabajador.html", {"trabajador": trabajador, "stocks": stocks, "movimientos": movimientos, "valor": valor, "es_admin": True})


def _pdf_inventario_response(titulo, filas, nombre_archivo, resumen=None):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.1*cm, leftMargin=1.1*cm, topMargin=1.1*cm, bottomMargin=1.1*cm)
    styles = getSampleStyleSheet(); story = [Paragraph(titulo, styles["Title"]), Spacer(1, 8)]
    story.append(Paragraph(f"Generado: {timezone.localdate().strftime('%d/%m/%Y')}", styles["Normal"]))
    if resumen:
        story.append(Spacer(1, 6)); story.append(Paragraph(resumen, styles["Normal"]))
    story.append(Spacer(1, 10))
    tabla = Table(filas, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#D9EAF7")), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), .45, colors.lightgrey), ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("FONTSIZE", (0,0), (-1,-1), 8.5), ("PADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(tabla); doc.build(story); buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response


@login_required
def inventario_general_pdf_view(request):
    if not es_admin(request.user): return render(request, "dashboard/no_autorizado.html", status=403)
    filas = [["Producto", "Categoría", "Stock", "Mínimo", "Costo/base", "Valor"]]
    for i in Insumo.objects.filter(activo=True).order_by("nombre"):
        filas.append([i.nombre, i.get_categoria_display(), _cantidad_texto(i, i.stock), _cantidad_texto(i, i.stock_minimo), f"${i.costo:,.4f}", f"${(Decimal(i.stock)*Decimal(i.costo or 0)):,.2f}"])
    return _pdf_inventario_response("Inventario General JVAQUA", filas, "inventario_general.pdf", f"Valor estimado del inventario: ${_inventario_valorizado():,.2f}")


@login_required
def inventario_trabajador_pdf_view(request, trabajador_id):
    if not es_admin(request.user): return render(request, "dashboard/no_autorizado.html", status=403)
    trabajador = get_object_or_404(Trabajador, pk=trabajador_id)
    filas = [["Producto", "Stock asignado", "Costo/base", "Valor"]]
    for x in InventarioTrabajador.objects.filter(trabajador=trabajador).select_related("insumo").order_by("insumo__nombre"):
        filas.append([x.insumo.nombre, _cantidad_texto(x.insumo, x.stock), f"${x.insumo.costo:,.4f}", f"${(Decimal(x.stock)*Decimal(x.insumo.costo or 0)):,.2f}"])
    return _pdf_inventario_response(f"Inventario de {trabajador}", filas, f"inventario_trabajador_{trabajador_id}.pdf")


@login_required
def inventario_kardex_pdf_view(request, insumo_id):
    if not es_admin(request.user): return render(request, "dashboard/no_autorizado.html", status=403)
    insumo = get_object_or_404(Insumo, pk=insumo_id)
    filas = [["Fecha", "Movimiento", "Cantidad", "Trabajador", "Mantenimiento", "Observación"]]
    for m in MovimientoInventario.objects.filter(insumo=insumo).select_related("trabajador__user", "mantenimiento__cliente").order_by("-creado_en"):
        filas.append([m.fecha.strftime("%d/%m/%Y"), m.get_tipo_display(), _cantidad_texto(insumo, m.cantidad), str(m.trabajador or "—"), str(m.mantenimiento or "—"), m.observacion or "—"])
    return _pdf_inventario_response(f"Kardex · {insumo.nombre}", filas, f"kardex_{insumo_id}.pdf", f"Stock general actual: {_cantidad_texto(insumo, insumo.stock)}")


@login_required
def inventario_movimientos_pdf_view(request):
    if not es_admin(request.user): return render(request, "dashboard/no_autorizado.html", status=403)
    filas = [["Fecha", "Producto", "Tipo", "Cantidad", "Trabajador", "Costo"]]
    for m in MovimientoInventario.objects.select_related("insumo", "trabajador__user").order_by("-creado_en")[:1000]:
        filas.append([m.fecha.strftime("%d/%m/%Y"), m.insumo.nombre, m.get_tipo_display(), _cantidad_texto(m.insumo, m.cantidad), str(m.trabajador or "—"), f"${m.total_costo:,.2f}"])
    return _pdf_inventario_response("Movimientos de Inventario JVAQUA", filas, "movimientos_inventario.pdf")


@login_required
def inventario_ventas_pdf_view(request):
    if not es_admin(request.user): return render(request, "dashboard/no_autorizado.html", status=403)
    filas = [["Fecha", "Producto", "Cantidad", "Venta", "Costo", "Ganancia"]]
    total = Decimal("0.00"); ganancia = Decimal("0.00")
    for v in VentaInsumo.objects.select_related("insumo").order_by("-creado_en")[:1000]:
        total += Decimal(v.total or 0); ganancia += Decimal(v.ganancia or 0)
        filas.append([v.fecha.strftime("%d/%m/%Y"), v.insumo.nombre, _cantidad_texto(v.insumo, v.cantidad), f"${v.total:,.2f}", f"${(Decimal(v.cantidad)*Decimal(v.costo_unitario or 0)):,.2f}", f"${v.ganancia:,.2f}"])
    return _pdf_inventario_response("Ventas de Insumos JVAQUA", filas, "ventas_inventario.pdf", f"Ventas: ${total:,.2f} · Ganancia estimada: ${ganancia:,.2f}")


@login_required
def inventario_compras_pdf_view(request):
    if not es_admin(request.user): return render(request, "dashboard/no_autorizado.html", status=403)
    filas = [["Fecha", "Producto", "Cantidad", "Costo/base", "Total", "Proveedor", "Lote / vencimiento"]]
    total = Decimal("0.00")
    for c in CompraInsumo.objects.select_related("insumo").order_by("-creado_en")[:1000]:
        total += Decimal(c.total or 0)
        lote_txt = c.lote or "—"
        if c.fecha_vencimiento:
            lote_txt += f" · {c.fecha_vencimiento.strftime('%d/%m/%Y')}"
        filas.append([c.fecha.strftime("%d/%m/%Y"), c.insumo.nombre, _cantidad_texto(c.insumo, c.cantidad), f"${c.costo_unitario:,.4f}", f"${c.total:,.2f}", c.proveedor or "—", lote_txt])
    return _pdf_inventario_response("Compras de Inventario JVAQUA", filas, "compras_inventario.pdf", f"Total compras registradas: ${total:,.2f}")


# ================================
# REPORTE DE GANANCIAS PRO
# ================================
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


def _iterar_meses(anio_inicio, mes_inicio, anio_fin, mes_fin):
    anio = anio_inicio
    mes = mes_inicio

    while (anio < anio_fin) or (anio == anio_fin and mes <= mes_fin):
        yield anio, mes
        if mes == 12:
            anio += 1
            mes = 1
        else:
            mes += 1


def _obtener_rango_grafico(fecha_inicio=None, fecha_fin=None):
    hoy = timezone.localdate()

    if fecha_inicio and fecha_fin:
        fecha_ini = parse_date(str(fecha_inicio))
        fecha_fin_real = parse_date(str(fecha_fin))

        if fecha_ini and fecha_fin_real:
            return fecha_ini, fecha_fin_real

    fecha_fin_real = hoy
    meses_atras = 5

    anio = hoy.year
    mes = hoy.month

    for _ in range(meses_atras):
        if mes == 1:
            anio -= 1
            mes = 12
        else:
            mes -= 1

    fecha_ini = date(anio, mes, 1)
    return fecha_ini, fecha_fin_real


def _obtener_serie_mensual_ganancias(fecha_inicio=None, fecha_fin=None):
    fecha_ini, fecha_fin_real = _obtener_rango_grafico(fecha_inicio, fecha_fin)

    labels = []
    ingresos_data = []
    egresos_data = []
    balance_data = []

    for anio, mes in _iterar_meses(fecha_ini.year, fecha_ini.month, fecha_fin_real.year, fecha_fin_real.month):
        primer_dia = date(anio, mes, 1)
        ultimo_dia = date(anio, mes, monthrange(anio, mes)[1])

        if anio == fecha_fin_real.year and mes == fecha_fin_real.month:
            ultimo_dia = min(ultimo_dia, fecha_fin_real)

        ingresos_mes = (
            Ingreso.objects.filter(fecha__range=(primer_dia, ultimo_dia))
            .aggregate(total=Sum("total"))
            .get("total")
            or Decimal("0")
        )
        egresos_mes = (
            Egreso.objects.filter(fecha__range=(primer_dia, ultimo_dia))
            .aggregate(total=Sum("total"))
            .get("total")
            or Decimal("0")
        )
        balance_mes = ingresos_mes - egresos_mes

        labels.append(f"{mes:02d}/{anio}")
        ingresos_data.append(float(ingresos_mes))
        egresos_data.append(float(egresos_mes))
        balance_data.append(float(balance_mes))

    return {
        "labels": labels,
        "ingresos": ingresos_data,
        "egresos": egresos_data,
        "balance": balance_data,
    }


def _obtener_datos_reporte_ganancias(fecha_inicio=None, fecha_fin=None):
    ingresos = Ingreso.objects.all().order_by("-fecha", "-id")
    egresos = (
        Egreso.objects.all()
        .select_related("insumo", "mantenimiento", "mantenimiento__cliente")
        .order_by("-fecha", "-id")
    )

    if fecha_inicio:
        ingresos = ingresos.filter(fecha__gte=fecha_inicio)
        egresos = egresos.filter(fecha__gte=fecha_inicio)

    if fecha_fin:
        ingresos = ingresos.filter(fecha__lte=fecha_fin)
        egresos = egresos.filter(fecha__lte=fecha_fin)

    total_ingresos = ingresos.aggregate(total=Sum("total"))["total"] or Decimal("0")
    total_egresos = egresos.aggregate(total=Sum("total"))["total"] or Decimal("0")
    ganancia = total_ingresos - total_egresos

    movimientos = []

    for i in ingresos:
        movimientos.append({
            "tipo": "Ingreso",
            "concepto": getattr(i, "concepto", "") or "-",
            "monto": i.total or Decimal("0"),
            "fecha": i.fecha,
        })

    for e in egresos:
        if getattr(e, "insumo", None):
            concepto = str(e.insumo)
        else:
            concepto = getattr(e, "concepto", "") or "Egreso"

        movimientos.append({
            "tipo": "Egreso",
            "concepto": concepto,
            "monto": e.total or Decimal("0"),
            "fecha": e.fecha,
        })

    movimientos.sort(key=lambda x: (x["fecha"], x["tipo"]), reverse=True)

    return {
        "ingresos": ingresos,
        "egresos": egresos,
        "movimientos": movimientos,
        "total_ingresos": total_ingresos,
        "total_egresos": total_egresos,
        "ganancia": ganancia,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
    }


@login_required
def reporte_ganancias_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    fecha_inicio = request.GET.get("fecha_inicio") or None
    fecha_fin = request.GET.get("fecha_fin") or None

    context = _obtener_datos_reporte_ganancias(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )

    serie = _obtener_serie_mensual_ganancias(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )

    mejor_mes_valor = max(serie["balance"]) if serie["balance"] else 0
    peor_mes_valor = min(serie["balance"]) if serie["balance"] else 0

    context.update({
        "grafico_ganancias": json.dumps(serie),
        "total_movimientos": len(context["movimientos"]),
        "ticket_promedio": (
            float(context["ganancia"]) / len(context["movimientos"])
            if context["movimientos"] else 0
        ),
        "mejor_mes_valor": mejor_mes_valor,
        "peor_mes_valor": peor_mes_valor,
    })

    return render(request, "dashboard/reporte_ganancias.html", context)


@login_required
def exportar_ganancias_excel(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    fecha_inicio = request.GET.get("fecha_inicio") or None
    fecha_fin = request.GET.get("fecha_fin") or None

    data = _obtener_datos_reporte_ganancias(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Ganancias"

    fill_titulo = PatternFill("solid", fgColor="1F4E78")
    fill_encabezado = PatternFill("solid", fgColor="D9EAF7")
    fill_ingreso = PatternFill("solid", fgColor="E2F0D9")
    fill_egreso = PatternFill("solid", fgColor="FDE9E7")

    font_blanco = Font(color="FFFFFF", bold=True, size=12)
    font_negrita = Font(bold=True)

    ws.merge_cells("A1:D1")
    ws["A1"] = "REPORTE DE GANANCIAS"
    ws["A1"].fill = fill_titulo
    ws["A1"].font = font_blanco
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A3"] = "Fecha inicio"
    ws["B3"] = data["fecha_inicio"] or "Todas"
    ws["C3"] = "Fecha fin"
    ws["D3"] = data["fecha_fin"] or "Todas"

    ws["A5"] = "Total ingresos"
    ws["B5"] = float(data["total_ingresos"])
    ws["C5"] = "Total egresos"
    ws["D5"] = float(data["total_egresos"])

    ws["A6"] = "Ganancia neta"
    ws["B6"] = float(data["ganancia"])

    for cell in ("A5", "C5", "A6"):
        ws[cell].font = font_negrita

    encabezados = ["Tipo", "Concepto", "Monto", "Fecha"]
    fila_inicio_tabla = 8

    for col, encabezado in enumerate(encabezados, start=1):
        cell = ws.cell(row=fila_inicio_tabla, column=col, value=encabezado)
        cell.fill = fill_encabezado
        cell.font = font_negrita
        cell.alignment = Alignment(horizontal="center")

    fila = fila_inicio_tabla + 1
    for mov in data["movimientos"]:
        ws.cell(row=fila, column=1, value=mov["tipo"])
        ws.cell(row=fila, column=2, value=mov["concepto"])
        ws.cell(row=fila, column=3, value=float(mov["monto"]))
        ws.cell(row=fila, column=4, value=mov["fecha"].strftime("%d/%m/%Y") if mov["fecha"] else "")

        if mov["tipo"] == "Ingreso":
            ws.cell(row=fila, column=1).fill = fill_ingreso
        else:
            ws.cell(row=fila, column=1).fill = fill_egreso

        fila += 1

    for row in ws.iter_rows(min_row=5, max_row=fila, min_col=2, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$ #,##0.00'

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_ganancias.xlsx"'

    wb.save(response)
    return response


@login_required
def exportar_ganancias_pdf(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    fecha_inicio = request.GET.get("fecha_inicio") or None
    fecha_fin = request.GET.get("fecha_fin") or None

    data = _obtener_datos_reporte_ganancias(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )

    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Reporte de Ganancias", styles["Title"]))
    story.append(Spacer(1, 10))

    rango_texto = f"Desde: {data['fecha_inicio'] or 'Todas'} &nbsp;&nbsp;&nbsp; Hasta: {data['fecha_fin'] or 'Todas'}"
    story.append(Paragraph(rango_texto, styles["Normal"]))
    story.append(Spacer(1, 10))

    resumen = [
        ["Total ingresos", f"${data['total_ingresos']:,.2f}"],
        ["Total egresos", f"${data['total_egresos']:,.2f}"],
        ["Ganancia neta", f"${data['ganancia']:,.2f}"],
    ]

    tabla_resumen = Table(resumen, colWidths=[7 * cm, 5 * cm])
    tabla_resumen.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
        ("BOX", (0, 0), (-1, -1), 0.75, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(tabla_resumen)
    story.append(Spacer(1, 14))

    detalle = [["Tipo", "Concepto", "Monto", "Fecha"]]
    for mov in data["movimientos"]:
        detalle.append([
            mov["tipo"],
            mov["concepto"],
            f"${mov['monto']:,.2f}",
            mov["fecha"].strftime("%d/%m/%Y") if mov["fecha"] else "",
        ])

    tabla_detalle = Table(detalle, colWidths=[3 * cm, 8 * cm, 3.2 * cm, 3.2 * cm])
    tabla_detalle.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOX", (0, 0), (-1, -1), 0.75, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("ALIGN", (2, 1), (2, -1), "RIGHT"),
        ("ALIGN", (3, 1), (3, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]))

    for idx, mov in enumerate(data["movimientos"], start=1):
        if mov["tipo"] == "Ingreso":
            tabla_detalle.setStyle(TableStyle([
                ("BACKGROUND", (0, idx), (0, idx), colors.HexColor("#E2F0D9"))
            ]))
        else:
            tabla_detalle.setStyle(TableStyle([
                ("BACKGROUND", (0, idx), (0, idx), colors.HexColor("#FDE9E7"))
            ]))

    story.append(tabla_detalle)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_ganancias.pdf"'
    response.write(pdf)
    return response


@login_required
def calculadora_quimicos_view(request):
    if not es_trabajador(request.user) and not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    return render(
        request,
        "dashboard/calculadora_quimicos.html",
    )


# ================================
# CONTRATOS
# ================================

FRECUENCIAS_CONTRATO_VALIDAS = {"1_semanal", "2_semanales", "3_semanales", "quincenal", "personalizado"}
FORMAS_PAGO_CONTRATO_VALIDAS = {valor for valor, _ in Contrato.FORMA_PAGO_CHOICES}
PROGRAMACIONES_COBRO_VALIDAS = {valor for valor, _ in Contrato.PROGRAMACION_COBRO_CHOICES}
MOMENTOS_FACTURACION_VALIDOS = {valor for valor, _ in Contrato.MOMENTO_FACTURACION_CHOICES}


def _entero_post(request, nombre, minimo=0, maximo=None, obligatorio=False):
    valor = (request.POST.get(nombre) or "").strip()
    if not valor:
        if obligatorio:
            raise ValueError
        return None
    numero = int(valor)
    if numero < minimo or (maximo is not None and numero > maximo):
        raise ValueError
    return numero


def _validar_datos_contrato(request):
    cliente_id = (request.POST.get("cliente") or "").strip()
    frecuencia = (request.POST.get("frecuencia") or "").strip()
    frecuencia_personalizada = (request.POST.get("frecuencia_personalizada") or "").strip()
    forma_pago = (request.POST.get("forma_pago") or "").strip()
    forma_pago_personalizada = (request.POST.get("forma_pago_personalizada") or "").strip()
    programacion_cobro = (request.POST.get("programacion_cobro") or "").strip()
    programacion_personalizada = (request.POST.get("programacion_cobro_personalizada") or "").strip()
    precio_mensual_str = (request.POST.get("precio_mensual") or "").strip()
    valor_tecnico_str = (request.POST.get("valor_tecnico_mensual") or "0").strip()
    fecha_inicio_str = (request.POST.get("fecha_inicio") or "").strip()
    activo = request.POST.get("activo") == "on"
    generacion_automatica = request.POST.get("generacion_automatica") == "on"
    tecnico_id = (request.POST.get("tecnico_designado") or "").strip()
    dias_visita = normalizar_dias(request.POST.getlist("dias_visita"))
    requiere_factura = request.POST.get("requiere_factura") == "on"
    notificar_facturacion = request.POST.get("notificar_facturacion") == "on"
    momento_facturacion = (request.POST.get("momento_facturacion") or "").strip()
    errores = []

    cliente = Cliente.objects.filter(pk=int(cliente_id)).first() if cliente_id.isdigit() else None
    if cliente is None: errores.append("Debes seleccionar un cliente válido.")
    if frecuencia not in FRECUENCIAS_CONTRATO_VALIDAS: errores.append("Debes seleccionar una frecuencia válida.")
    if frecuencia == "personalizado" and not frecuencia_personalizada: errores.append("Debes escribir la frecuencia personalizada.")
    if forma_pago not in FORMAS_PAGO_CONTRATO_VALIDAS: errores.append("Debes seleccionar una forma de pago válida.")
    if forma_pago == "personalizado" and not forma_pago_personalizada: errores.append("Debes escribir la forma de pago personalizada.")
    if programacion_cobro not in PROGRAMACIONES_COBRO_VALIDAS: errores.append("Debes seleccionar una programación de cobro válida.")
    if programacion_cobro == "personalizado" and not programacion_personalizada: errores.append("Describe la programación de cobro personalizada.")

    campos_enteros = {}
    configuracion = [
        ("periodo_dia_inicio", 1, 31, True), ("cobro_mes_desfase", 0, 2, True),
        ("cobro_dia_1", 1, 31, programacion_cobro in {"dia_fijo", "dos_pagos"}),
        ("cobro_dia_2", 1, 31, programacion_cobro == "dos_pagos"),
        ("cobro_rango_desde", 1, 31, programacion_cobro == "rango_dias"),
        ("cobro_rango_hasta", 1, 31, programacion_cobro == "rango_dias"),
        ("cobro_dias_despues_cierre", 0, 365, programacion_cobro == "despues_cierre"),
        ("facturacion_dia", 1, 31, requiere_factura and momento_facturacion in {"dia_fijo", "personalizado"}),
        ("facturacion_dias_antes", 0, 365, False),
        ("notificacion_factura_dias_antes", 0, 365, False),
    ]
    for nombre, minimo, maximo, obligatorio in configuracion:
        try: campos_enteros[nombre] = _entero_post(request, nombre, minimo, maximo, obligatorio)
        except (TypeError, ValueError): errores.append(f"Revisa el valor de {nombre.replace('_', ' ')}.")
    if programacion_cobro == "rango_dias" and campos_enteros.get("cobro_rango_desde") and campos_enteros.get("cobro_rango_hasta") and campos_enteros["cobro_rango_hasta"] < campos_enteros["cobro_rango_desde"]:
        errores.append("El último día del rango no puede ser anterior al primero.")

    try:
        porcentaje_primer_pago = Decimal((request.POST.get("porcentaje_primer_pago") or "50").strip())
        if not Decimal("0.01") <= porcentaje_primer_pago <= Decimal("99.99"): raise ValueError
    except Exception:
        porcentaje_primer_pago = Decimal("50.00"); errores.append("El porcentaje del primer pago debe estar entre 0.01 y 99.99.")

    if requiere_factura and momento_facturacion not in MOMENTOS_FACTURACION_VALIDOS:
        errores.append("Selecciona cuándo debe emitirse la factura.")

    tecnico_designado = None
    if tecnico_id:
        tecnico_designado = Trabajador.objects.filter(pk=int(tecnico_id), activo=True).select_related("user").first() if tecnico_id.isdigit() else None
        if tecnico_designado is None: errores.append("El técnico seleccionado no existe o está inactivo.")
    errores.extend(validar_programacion(frecuencia, dias_visita, tecnico_designado, automatica=generacion_automatica and activo))
    try:
        precio_mensual = Decimal(precio_mensual_str)
        if precio_mensual <= 0: raise ValueError
    except Exception:
        precio_mensual = None; errores.append("El precio mensual debe ser un valor mayor que cero.")
    try:
        valor_tecnico_mensual = Decimal(valor_tecnico_str or "0")
        if valor_tecnico_mensual < 0: raise ValueError
    except Exception:
        valor_tecnico_mensual = Decimal("0.00"); errores.append("El valor mensual del técnico debe ser cero o mayor.")
    if valor_tecnico_mensual and not tecnico_designado: errores.append("Debes seleccionar un técnico para asignarle un valor mensual.")
    fecha_inicio = parse_date(fecha_inicio_str)
    if not fecha_inicio: errores.append("Debes seleccionar una fecha de inicio válida.")

    return {
        "errores": errores, "cliente": cliente, "frecuencia": frecuencia,
        "frecuencia_personalizada": frecuencia_personalizada, "forma_pago": forma_pago,
        "forma_pago_personalizada": forma_pago_personalizada, "precio_mensual": precio_mensual,
        "valor_tecnico_mensual": valor_tecnico_mensual, "fecha_inicio": fecha_inicio, "activo": activo,
        "generacion_automatica": generacion_automatica, "tecnico_designado": tecnico_designado,
        "tecnico_id": tecnico_id, "dias_visita": dias_visita, "programacion_cobro": programacion_cobro,
        "programacion_cobro_personalizada": programacion_personalizada, "porcentaje_primer_pago": porcentaje_primer_pago,
        "requiere_factura": requiere_factura, "momento_facturacion": momento_facturacion if requiere_factura else "",
        "notificar_facturacion": notificar_facturacion if requiere_factura else False,
        "observaciones_facturacion": (request.POST.get("observaciones_facturacion") or "").strip(),
        **campos_enteros,
    }



def _normalizar_telefono_cliente(valor):
    return "".join(caracter for caracter in (valor or "") if caracter.isdigit())


def _datos_cliente_desde_request(request):
    return {
        "nombre": (request.POST.get("nombre") or "").strip(),
        "telefono": (request.POST.get("telefono") or "").strip(),
        "email": (request.POST.get("email") or "").strip(),
        "ciudad": (request.POST.get("ciudad") or "").strip(),
        "sector_urbanizacion": (request.POST.get("sector_urbanizacion") or "").strip(),
        "direccion": (request.POST.get("direccion") or "").strip(),
        "enlace_google_maps": (request.POST.get("enlace_google_maps") or "").strip(),
    }


def _validar_cliente(datos, cliente_actual=None):
    errores = []
    for campo, etiqueta in (
        ("nombre", "El nombre"),
        ("telefono", "El teléfono principal"),
        ("ciudad", "La ciudad"),
        ("sector_urbanizacion", "El sector o urbanización"),
        ("direccion", "La dirección"),
    ):
        if not datos.get(campo):
            errores.append(f"{etiqueta} es obligatorio.")

    telefono_normalizado = _normalizar_telefono_cliente(datos.get("telefono"))
    if datos.get("telefono") and len(telefono_normalizado) < 7:
        errores.append("Ingresa un teléfono principal válido.")

    if telefono_normalizado:
        candidatos = Cliente.objects.all()
        if cliente_actual:
            candidatos = candidatos.exclude(pk=cliente_actual.pk)
        duplicado = next(
            (
                cliente for cliente in candidatos.only("id", "nombre", "telefono")
                if _normalizar_telefono_cliente(cliente.telefono) == telefono_normalizado
            ),
            None,
        )
        if duplicado:
            errores.append(
                f"Ya existe el cliente {duplicado.nombre} con este teléfono."
            )

    return errores


@login_required
def cliente_list_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    q = (request.GET.get("q") or "").strip()
    estado = (request.GET.get("estado") or "").strip().lower()
    clientes = Cliente.objects.annotate(
        total_contratos=Count("contratos", distinct=True),
        contratos_activos=Count(
            "contratos",
            filter=models.Q(contratos__activo=True),
            distinct=True,
        ),
    ).order_by("nombre", "id")

    if q:
        clientes = clientes.filter(
            models.Q(nombre__icontains=q)
            | models.Q(telefono__icontains=q)
            | models.Q(email__icontains=q)
            | models.Q(ciudad__icontains=q)
            | models.Q(sector_urbanizacion__icontains=q)
            | models.Q(direccion__icontains=q)
        )
    if estado == "activo":
        clientes = clientes.filter(activo=True)
    elif estado == "inactivo":
        clientes = clientes.filter(activo=False)

    total_clientes = clientes.count()
    total_activos = clientes.filter(activo=True).count()
    con_contrato = clientes.filter(contratos_activos__gt=0).count()
    sin_contrato = clientes.filter(total_contratos=0).count()

    paginator = Paginator(clientes, 20)
    page_obj = paginator.get_page(request.GET.get("page"))
    query_params = request.GET.copy()
    query_params.pop("page", None)

    return render(request, "dashboard/cliente_list.html", {
        "page_obj": page_obj,
        "q": q,
        "estado": estado,
        "total_clientes": total_clientes,
        "total_activos": total_activos,
        "con_contrato": con_contrato,
        "sin_contrato": sin_contrato,
        "querystring": query_params.urlencode(),
        "es_admin": True,
    })


@login_required
def cliente_crear_view(request):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    datos = {
        "nombre": "", "telefono": "", "email": "", "ciudad": "",
        "sector_urbanizacion": "", "direccion": "", "enlace_google_maps": "",
    }
    if request.method == "POST":
        datos = _datos_cliente_desde_request(request)
        errores = _validar_cliente(datos)
        if errores:
            for error in errores:
                messages.error(request, error)
        else:
            cliente = Cliente.objects.create(**datos)
            _registrar_actividad(
                user=request.user,
                titulo="Cliente creado",
                descripcion=f"{request.user.username} registró a {cliente.nombre}.",
                url=f"/dashboard/clientes/{cliente.pk}/",
            )
            messages.success(request, "Cliente creado correctamente.")
            if request.POST.get("accion") == "guardar_y_contrato":
                return redirect(f"/dashboard/contratos/nuevo/?cliente={cliente.pk}")
            return redirect(f"/dashboard/clientes/{cliente.pk}/")

    return render(request, "dashboard/cliente_form.html", {
        "modo": "crear", "cliente": None, "datos": datos, "es_admin": True,
    })


@login_required
def cliente_editar_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    cliente = get_object_or_404(Cliente, pk=pk)
    datos = {
        "nombre": cliente.nombre,
        "telefono": cliente.telefono,
        "email": cliente.email or "",
        "ciudad": cliente.ciudad,
        "sector_urbanizacion": cliente.sector_urbanizacion,
        "direccion": cliente.direccion,
        "enlace_google_maps": cliente.enlace_google_maps,
    }
    if request.method == "POST":
        datos = _datos_cliente_desde_request(request)
        errores = _validar_cliente(datos, cliente_actual=cliente)
        if errores:
            for error in errores:
                messages.error(request, error)
        else:
            for campo, valor in datos.items():
                setattr(cliente, campo, valor)
            cliente.save()
            _registrar_actividad(
                user=request.user,
                titulo="Cliente actualizado",
                descripcion=f"{request.user.username} actualizó a {cliente.nombre}.",
                url=f"/dashboard/clientes/{cliente.pk}/",
            )
            messages.success(request, "Cliente actualizado correctamente.")
            return redirect(f"/dashboard/clientes/{cliente.pk}/")

    return render(request, "dashboard/cliente_form.html", {
        "modo": "editar", "cliente": cliente, "datos": datos, "es_admin": True,
    })


@login_required
def cliente_detalle_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)
    cliente = get_object_or_404(Cliente, pk=pk)
    contratos = cliente.contratos.select_related("tecnico_designado__user").order_by("-activo", "-id")
    mantenimientos = Mantenimiento.objects.filter(cliente=cliente).order_by("-fecha", "-id")
    proximo_mantenimiento = mantenimientos.filter(
        estado="pendiente", fecha__gte=timezone.localdate()
    ).order_by("fecha").first()
    return render(request, "dashboard/cliente_detalle.html", {
        "cliente": cliente,
        "contratos": contratos,
        "total_contratos": contratos.count(),
        "contratos_activos": contratos.filter(activo=True).count(),
        "total_mantenimientos": mantenimientos.count(),
        "proximo_mantenimiento": proximo_mantenimiento,
        "mantenimientos_recientes": mantenimientos[:8],
        "es_admin": True,
    })


@login_required
@require_http_methods(["POST"])
def cliente_crear_rapido_view(request):
    if not es_admin(request.user):
        return JsonResponse({"ok": False, "errores": ["No autorizado."]}, status=403)
    datos = _datos_cliente_desde_request(request)
    errores = _validar_cliente(datos)
    if errores:
        return JsonResponse({"ok": False, "errores": errores}, status=400)
    cliente = Cliente.objects.create(**datos)
    _registrar_actividad(
        user=request.user,
        titulo="Cliente creado",
        descripcion=f"{request.user.username} registró rápidamente a {cliente.nombre}.",
        url=f"/dashboard/clientes/{cliente.pk}/",
    )
    return JsonResponse({
        "ok": True,
        "cliente": {"id": cliente.pk, "nombre": cliente.nombre, "telefono": cliente.telefono},
    })


@login_required
def contrato_list_view(request):
    if not es_admin(request.user):
        return render(
            request,
            "dashboard/no_autorizado.html",
            status=403,
        )

    q = (request.GET.get("q") or "").strip()
    estado = (request.GET.get("estado") or "").strip().lower()
    frecuencia = (
        request.GET.get("frecuencia") or ""
    ).strip()

    contratos = (
        Contrato.objects
        .select_related("cliente")
        .annotate(
            total_mantenimientos=Count(
                "mantenimiento",
                distinct=True,
            )
        )
        .order_by("-activo", "cliente__nombre", "id")
    )

    if q:
        contratos = contratos.filter(
            models.Q(cliente__nombre__icontains=q)
            | models.Q(cliente__telefono__icontains=q)
            | models.Q(cliente__email__icontains=q)
            | models.Q(frecuencia_personalizada__icontains=q)
            | models.Q(forma_pago_personalizada__icontains=q)
        )

    if estado == "activo":
        contratos = contratos.filter(activo=True)
    elif estado == "inactivo":
        contratos = contratos.filter(activo=False)

    if frecuencia in FRECUENCIAS_CONTRATO_VALIDAS:
        contratos = contratos.filter(frecuencia=frecuencia)

    total_contratos = contratos.count()
    total_activos = contratos.filter(activo=True).count()
    total_inactivos = contratos.filter(activo=False).count()

    ingreso_mensual_estimado = (
        contratos
        .filter(activo=True)
        .aggregate(total=Sum("precio_mensual"))
        .get("total")
        or Decimal("0.00")
    )

    paginator = Paginator(contratos, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    if "page" in query_params:
        query_params.pop("page")
    querystring = query_params.urlencode()

    return render(
        request,
        "dashboard/contrato_list.html",
        {
            "page_obj": page_obj,
            "q": q,
            "estado": estado,
            "frecuencia": frecuencia,
            "frecuencias": Contrato.FRECUENCIA_CHOICES,
            "total_contratos": total_contratos,
            "total_activos": total_activos,
            "total_inactivos": total_inactivos,
            "ingreso_mensual_estimado": ingreso_mensual_estimado,
            "querystring": querystring,
            "es_admin": True,
        },
    )


@login_required
def contrato_crear_view(request):
    if not es_admin(request.user):
        return render(
            request,
            "dashboard/no_autorizado.html",
            status=403,
        )

    clientes = (
        Cliente.objects
        .filter(activo=True)
        .order_by("nombre")
    )
    trabajadores = (
        Trabajador.objects
        .filter(activo=True)
        .select_related("user")
        .order_by("user__username")
    )

    datos_formulario = {
        "cliente_id": (request.GET.get("cliente") or "").strip(),
        "frecuencia": "",
        "frecuencia_personalizada": "",
        "forma_pago": "",
        "forma_pago_personalizada": "",
        "periodo_dia_inicio": timezone.localdate().day,
        "programacion_cobro": "inicio_periodo",
        "cobro_mes_desfase": 0,
        "cobro_dia_1": "", "cobro_dia_2": "",
        "cobro_rango_desde": "", "cobro_rango_hasta": "",
        "cobro_dias_despues_cierre": 0, "porcentaje_primer_pago": "50.00",
        "programacion_cobro_personalizada": "",
        "requiere_factura": False, "momento_facturacion": "", "facturacion_dia": "",
        "facturacion_dias_antes": 0, "notificar_facturacion": False,
        "notificacion_factura_dias_antes": 1, "observaciones_facturacion": "",
        "precio_mensual": "",
        "valor_tecnico_mensual": "",
        "fecha_inicio": timezone.localdate().isoformat(),
        "activo": True,
        "generacion_automatica": True,
        "tecnico_id": "",
        "dias_visita": [],
    }

    if request.method == "POST":
        validacion = _validar_datos_contrato(request)

        datos_formulario = {
            "cliente_id": request.POST.get("cliente", ""),
            "frecuencia": validacion["frecuencia"],
            "frecuencia_personalizada": (
                validacion["frecuencia_personalizada"]
            ),
            "forma_pago": validacion["forma_pago"],
            "forma_pago_personalizada": (
                validacion["forma_pago_personalizada"]
            ),
            **{campo: validacion.get(campo) or "" for campo in (
                "periodo_dia_inicio", "programacion_cobro", "cobro_mes_desfase", "cobro_dia_1", "cobro_dia_2",
                "cobro_rango_desde", "cobro_rango_hasta", "cobro_dias_despues_cierre", "porcentaje_primer_pago",
                "programacion_cobro_personalizada", "momento_facturacion", "facturacion_dia", "facturacion_dias_antes",
                "notificacion_factura_dias_antes", "observaciones_facturacion")},
            "requiere_factura": validacion["requiere_factura"],
            "notificar_facturacion": validacion["notificar_facturacion"],
            "precio_mensual": request.POST.get("precio_mensual", ""),
            "valor_tecnico_mensual": request.POST.get("valor_tecnico_mensual", ""),
            "fecha_inicio": request.POST.get(
                "fecha_inicio",
                "",
            ),
            "activo": validacion["activo"],
            "generacion_automatica": validacion["generacion_automatica"],
            "tecnico_id": validacion["tecnico_id"],
            "dias_visita": validacion["dias_visita"],
        }

        if validacion["errores"]:
            for error in validacion["errores"]:
                messages.error(request, error)
        else:
            contrato = Contrato.objects.create(
                cliente=validacion["cliente"],
                tipo="variable",
                frecuencia=validacion["frecuencia"],
                frecuencia_personalizada=(
                    validacion["frecuencia_personalizada"]
                ),
                forma_pago=validacion["forma_pago"],
                forma_pago_personalizada=(
                    validacion["forma_pago_personalizada"]
                ),
                periodo_dia_inicio=validacion["periodo_dia_inicio"],
                programacion_cobro=validacion["programacion_cobro"],
                cobro_mes_desfase=validacion["cobro_mes_desfase"] or 0,
                cobro_dia_1=validacion["cobro_dia_1"], cobro_dia_2=validacion["cobro_dia_2"],
                cobro_rango_desde=validacion["cobro_rango_desde"], cobro_rango_hasta=validacion["cobro_rango_hasta"],
                cobro_dias_despues_cierre=validacion["cobro_dias_despues_cierre"] or 0,
                porcentaje_primer_pago=validacion["porcentaje_primer_pago"],
                programacion_cobro_personalizada=validacion["programacion_cobro_personalizada"],
                requiere_factura=validacion["requiere_factura"], momento_facturacion=validacion["momento_facturacion"],
                facturacion_dia=validacion["facturacion_dia"], facturacion_dias_antes=validacion["facturacion_dias_antes"] or 0,
                notificar_facturacion=validacion["notificar_facturacion"],
                notificacion_factura_dias_antes=validacion["notificacion_factura_dias_antes"] or 1,
                observaciones_facturacion=validacion["observaciones_facturacion"],
                precio_mensual=validacion["precio_mensual"],
                valor_tecnico_mensual=validacion["valor_tecnico_mensual"],
                fecha_inicio=validacion["fecha_inicio"],
                activo=validacion["activo"],
                tecnico_designado=validacion["tecnico_designado"],
                dias_visita=validacion["dias_visita"],
                generacion_automatica=validacion["generacion_automatica"],
            )

            resultado_programacion = generar_mantenimientos_contrato(contrato)

            _registrar_actividad(
                user=request.user,
                titulo="Contrato creado",
                descripcion=(
                    f"{request.user.username} creó el contrato "
                    f"de {contrato.cliente} por "
                    f"${contrato.precio_mensual} mensuales."
                ),
                url=f"/dashboard/contratos/{contrato.pk}/",
            )

            creados = resultado_programacion.get("creados", 0)
            messages.success(
                request,
                f"Contrato creado correctamente. Se programaron {creados} mantenimientos.",
            )
            return redirect(
                f"/dashboard/contratos/{contrato.pk}/"
            )

    return render(
        request,
        "dashboard/contrato_form.html",
        {
            "modo": "crear",
            "contrato": None,
            "clientes": clientes,
            "frecuencias": Contrato.FRECUENCIA_CHOICES,
            "formas_pago": Contrato.FORMA_PAGO_CHOICES,
            "programaciones_cobro": Contrato.PROGRAMACION_COBRO_CHOICES,
            "momentos_facturacion": Contrato.MOMENTO_FACTURACION_CHOICES,
            "trabajadores": trabajadores,
            "dias_semana": DIAS_SEMANA.items(),
            "dias_mes": range(1, 32),
            "datos_formulario": datos_formulario,
            "es_admin": True,
        },
    )


@login_required
def contrato_editar_view(request, pk):
    if not es_admin(request.user):
        return render(
            request,
            "dashboard/no_autorizado.html",
            status=403,
        )

    contrato = get_object_or_404(
        Contrato.objects.select_related("cliente"),
        pk=pk,
    )

    clientes = Cliente.objects.order_by(
        "-activo",
        "nombre",
    )
    trabajadores = (
        Trabajador.objects
        .filter(activo=True)
        .select_related("user")
        .order_by("user__username")
    )

    datos_formulario = {
        "cliente_id": str(contrato.cliente_id),
        "frecuencia": contrato.frecuencia,
        "frecuencia_personalizada": (
            contrato.frecuencia_personalizada
        ),
        "forma_pago": contrato.forma_pago,
        "forma_pago_personalizada": (
            contrato.forma_pago_personalizada
        ),
        "periodo_dia_inicio": contrato.periodo_dia_inicio,
        "programacion_cobro": contrato.programacion_cobro,
        "cobro_mes_desfase": contrato.cobro_mes_desfase,
        "cobro_dia_1": contrato.cobro_dia_1 or "", "cobro_dia_2": contrato.cobro_dia_2 or "",
        "cobro_rango_desde": contrato.cobro_rango_desde or "", "cobro_rango_hasta": contrato.cobro_rango_hasta or "",
        "cobro_dias_despues_cierre": contrato.cobro_dias_despues_cierre,
        "porcentaje_primer_pago": contrato.porcentaje_primer_pago,
        "programacion_cobro_personalizada": contrato.programacion_cobro_personalizada,
        "requiere_factura": contrato.requiere_factura, "momento_facturacion": contrato.momento_facturacion,
        "facturacion_dia": contrato.facturacion_dia or "", "facturacion_dias_antes": contrato.facturacion_dias_antes,
        "notificar_facturacion": contrato.notificar_facturacion,
        "notificacion_factura_dias_antes": contrato.notificacion_factura_dias_antes,
        "observaciones_facturacion": contrato.observaciones_facturacion,
        "precio_mensual": contrato.precio_mensual,
        "valor_tecnico_mensual": contrato.valor_tecnico_mensual,
        "fecha_inicio": contrato.fecha_inicio.isoformat(),
        "activo": contrato.activo,
        "generacion_automatica": contrato.generacion_automatica,
        "tecnico_id": str(contrato.tecnico_designado_id or ""),
        "dias_visita": normalizar_dias(contrato.dias_visita),
    }

    if request.method == "POST":
        validacion = _validar_datos_contrato(request)

        datos_formulario = {
            "cliente_id": request.POST.get("cliente", ""),
            "frecuencia": validacion["frecuencia"],
            "frecuencia_personalizada": (
                validacion["frecuencia_personalizada"]
            ),
            "forma_pago": validacion["forma_pago"],
            "forma_pago_personalizada": (
                validacion["forma_pago_personalizada"]
            ),
            **{campo: validacion.get(campo) or "" for campo in (
                "periodo_dia_inicio", "programacion_cobro", "cobro_mes_desfase", "cobro_dia_1", "cobro_dia_2",
                "cobro_rango_desde", "cobro_rango_hasta", "cobro_dias_despues_cierre", "porcentaje_primer_pago",
                "programacion_cobro_personalizada", "momento_facturacion", "facturacion_dia", "facturacion_dias_antes",
                "notificacion_factura_dias_antes", "observaciones_facturacion")},
            "requiere_factura": validacion["requiere_factura"],
            "notificar_facturacion": validacion["notificar_facturacion"],
            "precio_mensual": request.POST.get("precio_mensual", ""),
            "valor_tecnico_mensual": request.POST.get("valor_tecnico_mensual", ""),
            "fecha_inicio": request.POST.get(
                "fecha_inicio",
                "",
            ),
            "activo": validacion["activo"],
            "generacion_automatica": validacion["generacion_automatica"],
            "tecnico_id": validacion["tecnico_id"],
            "dias_visita": validacion["dias_visita"],
        }

        if validacion["errores"]:
            for error in validacion["errores"]:
                messages.error(request, error)
        else:
            contrato.cliente = validacion["cliente"]
            contrato.frecuencia = validacion["frecuencia"]
            contrato.frecuencia_personalizada = (
                validacion["frecuencia_personalizada"]
            )
            contrato.forma_pago = validacion["forma_pago"]
            contrato.forma_pago_personalizada = (
                validacion["forma_pago_personalizada"]
            )
            for campo in (
                "periodo_dia_inicio", "programacion_cobro", "cobro_mes_desfase", "cobro_dia_1", "cobro_dia_2",
                "cobro_rango_desde", "cobro_rango_hasta", "cobro_dias_despues_cierre", "porcentaje_primer_pago",
                "programacion_cobro_personalizada", "requiere_factura", "momento_facturacion", "facturacion_dia",
                "facturacion_dias_antes", "notificar_facturacion", "notificacion_factura_dias_antes",
                "observaciones_facturacion"):
                setattr(contrato, campo, validacion[campo])
            contrato.precio_mensual = validacion["precio_mensual"]
            contrato.valor_tecnico_mensual = validacion["valor_tecnico_mensual"]
            contrato.fecha_inicio = validacion["fecha_inicio"]
            contrato.activo = validacion["activo"]
            contrato.tecnico_designado = validacion["tecnico_designado"]
            contrato.dias_visita = validacion["dias_visita"]
            contrato.generacion_automatica = validacion["generacion_automatica"]
            contrato.save()

            if contrato.activo and contrato.generacion_automatica:
                resultado_programacion = generar_mantenimientos_contrato(
                    contrato,
                    reconciliar=True,
                )
            else:
                eliminados = cancelar_programacion_futura(contrato)
                resultado_programacion = {"creados": 0, "eliminados": eliminados}

            _registrar_actividad(
                user=request.user,
                titulo="Contrato actualizado",
                descripcion=(
                    f"{request.user.username} actualizó el contrato "
                    f"de {contrato.cliente}."
                ),
                url=f"/dashboard/contratos/{contrato.pk}/",
            )

            messages.success(
                request,
                "Contrato actualizado correctamente. La programación futura fue sincronizada.",
            )
            return redirect(
                f"/dashboard/contratos/{contrato.pk}/"
            )

    return render(
        request,
        "dashboard/contrato_form.html",
        {
            "modo": "editar",
            "contrato": contrato,
            "clientes": clientes,
            "frecuencias": Contrato.FRECUENCIA_CHOICES,
            "formas_pago": Contrato.FORMA_PAGO_CHOICES,
            "programaciones_cobro": Contrato.PROGRAMACION_COBRO_CHOICES,
            "momentos_facturacion": Contrato.MOMENTO_FACTURACION_CHOICES,
            "trabajadores": trabajadores,
            "dias_semana": DIAS_SEMANA.items(),
            "dias_mes": range(1, 32),
            "datos_formulario": datos_formulario,
            "es_admin": True,
        },
    )


@login_required
def contrato_detalle_view(request, pk):
    if not es_admin(request.user):
        return render(
            request,
            "dashboard/no_autorizado.html",
            status=403,
        )

    contrato = get_object_or_404(
        Contrato.objects.select_related("cliente", "tecnico_designado__user"),
        pk=pk,
    )

    mantenimientos = (
        Mantenimiento.objects
        .filter(contrato=contrato)
        .select_related("cliente", "contrato")
        .prefetch_related("trabajadores")
        .order_by("-fecha", "-id")
    )

    total_mantenimientos = mantenimientos.count()
    total_realizados = mantenimientos.filter(
        estado="realizado"
    ).count()
    total_pendientes = mantenimientos.filter(
        estado="pendiente"
    ).count()
    total_atrasados = mantenimientos.filter(
        estado="pendiente",
        fecha__lt=timezone.localdate(),
    ).count()

    mantenimientos_recientes = mantenimientos[:10]

    facturas = (
        Factura.objects
        .filter(contrato=contrato)
        .order_by("-periodo_anio", "-periodo_mes", "-id")[:10]
    )

    total_facturado = (
        Factura.objects
        .filter(contrato=contrato)
        .aggregate(total=Sum("total"))
        .get("total")
        or Decimal("0.00")
    )

    total_pagado = (
        Factura.objects
        .filter(
            contrato=contrato,
            estado=Factura.ESTADO_PAGADA,
        )
        .aggregate(total=Sum("total"))
        .get("total")
        or Decimal("0.00")
    )

    total_pendiente = (
        Factura.objects
        .filter(
            contrato=contrato,
            estado__in=[
                Factura.ESTADO_PENDIENTE,
                Factura.ESTADO_VENCIDA,
            ],
        )
        .aggregate(total=Sum("total"))
        .get("total")
        or Decimal("0.00")
    )

    mantenimientos_futuros = mantenimientos.filter(
        estado="pendiente",
        fecha__gte=timezone.localdate(),
    ).count()
    proximo_mantenimiento = mantenimientos.filter(
        estado="pendiente",
        fecha__gte=timezone.localdate(),
    ).order_by("fecha").first()

    hoy = timezone.localdate()
    calendario_cobros = contrato.calendario_cobros(hoy.year, hoy.month)
    periodo_inicio, periodo_fin = contrato.periodo_servicio(hoy.year, hoy.month)
    fecha_facturacion_programada = contrato.fecha_programada_facturacion(hoy.year, hoy.month)

    return render(
        request,
        "dashboard/contrato_detalle.html",
        {
            "contrato": contrato,
            "mantenimientos_recientes": mantenimientos_recientes,
            "facturas": facturas,
            "total_mantenimientos": total_mantenimientos,
            "total_realizados": total_realizados,
            "total_pendientes": total_pendientes,
            "total_atrasados": total_atrasados,
            "total_facturado": total_facturado,
            "total_pagado": total_pagado,
            "total_pendiente": total_pendiente,
            "mantenimientos_futuros": mantenimientos_futuros,
            "proximo_mantenimiento": proximo_mantenimiento,
            "periodo_inicio_actual": periodo_inicio,
            "periodo_fin_actual": periodo_fin,
            "calendario_cobros": calendario_cobros,
            "fecha_facturacion_programada": fecha_facturacion_programada,
            "es_admin": True,
        },
    )


@login_required
@require_http_methods(["POST"])
def contrato_regenerar_programacion_view(request, pk):
    if not es_admin(request.user):
        return render(request, "dashboard/no_autorizado.html", status=403)

    contrato = get_object_or_404(
        Contrato.objects.select_related("cliente", "tecnico_designado"),
        pk=pk,
    )
    resultado = generar_mantenimientos_contrato(contrato, reconciliar=True)
    errores = resultado.get("errores") or []
    if errores:
        for error in errores:
            messages.error(request, error)
    else:
        messages.success(
            request,
            f"Programación actualizada: {resultado['creados']} mantenimientos creados.",
        )
    return redirect(f"/dashboard/contratos/{contrato.pk}/")


@login_required
@require_http_methods(["POST"])
def contrato_toggle_view(request, pk):
    if not es_admin(request.user):
        return render(
            request,
            "dashboard/no_autorizado.html",
            status=403,
        )

    contrato = get_object_or_404(Contrato, pk=pk)

    contrato.activo = not contrato.activo
    contrato.save(update_fields=["activo"])

    if contrato.activo and contrato.generacion_automatica:
        generar_mantenimientos_contrato(contrato, reconciliar=True)
    elif not contrato.activo:
        cancelar_programacion_futura(contrato)

    estado_texto = (
        "activado"
        if contrato.activo
        else "desactivado"
    )

    _registrar_actividad(
        user=request.user,
        titulo="Estado de contrato actualizado",
        descripcion=(
            f"{request.user.username} {estado_texto} "
            f"el contrato de {contrato.cliente}."
        ),
        url=f"/dashboard/contratos/{contrato.pk}/",
    )

    messages.success(
        request,
        f"Contrato {estado_texto} correctamente.",
    )

    return redirect(
        request.POST.get(
            "next",
            f"/dashboard/contratos/{contrato.pk}/",
        )
    )